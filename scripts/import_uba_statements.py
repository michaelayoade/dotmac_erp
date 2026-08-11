#!/usr/bin/env python
"""
Import UBA Bank Statements from Excel files.

Handles password-protected UBA statement files.

Usage:
    poetry run python scripts/import_uba_statements.py --org-id <uuid> [--dry-run]
"""

from __future__ import annotations

import argparse
import io
import logging
import os
import sys
import uuid
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path

import msoffcrypto
import openpyxl

from app.db.session_context import session_for_org
from app.models.batch_operation import BatchOperationType
from app.services.batch_operation import batch_operation, file_manifest_digest
from app.services.finance.banking.bank_statement import (
    BankStatementService,
)
from app.services.finance.banking.statement_import import (
    AccountProfile,
    BankProfile,
    ensure_bank_account,
    find_bank_account,
    to_statement_lines,
)
from app.services.finance.banking.statement_parsing import (
    extract_account_number,
    parse_date,
    parse_decimal,
    parse_period,
)

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

BANK = BankProfile(bank_name="United Bank for Africa", bank_code="033")

SYSTEM_ACTOR_ID = uuid.UUID("00000000-0000-0000-0000-000000000000")

# Where the statement workbooks are read from. Overridable, with no
# environment-specific path baked in.
STATEMENT_DIR = Path(
    os.environ.get("UBA_STATEMENT_DIR", "~/.dotmac/statements/uba")
).expanduser()


def _statement_password(account_number: str) -> str:
    """The workbook password for one UBA account, from the environment.

    These were committed literals until 2026-08-11. They are bank-issued
    passwords for the statement workbooks, and no default is provided on
    purpose: a script that silently falls back to "no password" would report
    a parse failure rather than a missing credential, which is the harder
    thing to diagnose.

    Set `UBA_STATEMENT_PASSWORD_<account-number>` in the environment, sourced
    from OpenBao rather than typed. Deliberately NOT derived in code from the
    account number — that is UBA's scheme to change, not ours to hardcode.
    """
    variable = f"UBA_STATEMENT_PASSWORD_{account_number}"
    password = os.environ.get(variable)
    if not password:
        raise SystemExit(
            f"{variable} is not set — the workbook for account {account_number} "
            "cannot be opened. Load it from OpenBao into the environment; it is "
            "deliberately not stored in this file."
        )
    return password


# Account configuration. Passwords are NOT here — see `_statement_password`.
ACCOUNT_CONFIG = {
    "1018904696": {
        "name": "UBA 96 (Main)",
        "currency": "NGN",
        "gl_account_code": "1202",  # UBA GL account
        "files": [
            "101xxxxx96.xlsx",
            "101xxxxx96 (1).xlsx",
        ],
    },
    "3004154294": {
        "name": "UBA USD",
        "currency": "USD",
        "gl_account_code": "1202",  # UBA GL account
        "files": [
            "300xxxxx94.xlsx",
        ],
    },
}


@dataclass
class ParsedStatement:
    """Parsed statement data from Excel file."""

    account_number: str
    account_name: str
    currency: str
    period_start: date
    period_end: date
    opening_balance: Decimal
    closing_balance: Decimal
    total_debit: Decimal
    total_credit: Decimal
    transactions: list[dict]
    source_file: str


def open_workbook(filepath: Path, password: str | None = None):
    """Open Excel workbook, handling password protection."""
    try:
        # Try without password first
        return openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        if password:
            with open(filepath, "rb") as file:
                decrypted = io.BytesIO()
                office_file = msoffcrypto.OfficeFile(file)
                office_file.load_key(password=password)
                office_file.decrypt(decrypted)
                decrypted.seek(0)
                return openpyxl.load_workbook(decrypted, data_only=True)
        raise


def parse_uba_statement(
    filepath: Path, password: str | None = None
) -> ParsedStatement | None:
    """
    Parse UBA statement format.

    Structure:
    - Row 4: Account Number
    - Row 5: Account Name
    - Row 8: Opening Balance
    - Row 9: Total Credit
    - Row 10: Total Debit
    - Row 11: Closing Balance
    - Row 12: Currency
    - Row 13: Period
    - Row 17: Headers (Tran Date, Value Date, Narration, Chq. No, Debit, Credit, Balance)
    - Row 18+: Data
    """
    logger.info(f"Parsing UBA statement: {filepath.name}")

    try:
        wb = open_workbook(filepath, password)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))

        # Extract header info (0-indexed)
        account_number = extract_account_number(str(rows[3][1])) if rows[3][1] else None
        account_name = str(rows[4][1]).strip() if rows[4][1] else ""
        opening_balance = parse_decimal(rows[7][1])
        total_credit = parse_decimal(rows[8][1])
        total_debit = parse_decimal(rows[9][1])
        closing_balance = parse_decimal(rows[10][1])
        currency = str(rows[11][1]).strip() if rows[11][1] else "NGN"
        period_start, period_end = parse_period(str(rows[12][1]) if rows[12][1] else "")

        if not account_number:
            logger.warning(f"Could not extract account number from {filepath.name}")
            return None

        # Parse transactions (start from row 18, index 17)
        transactions = []
        for _i, row in enumerate(rows[17:], start=18):
            tran_date = parse_date(row[0])
            if not tran_date:
                continue

            value_date = parse_date(row[1])
            narration = str(row[2]).strip() if row[2] else ""
            chq_no = str(row[3]).strip() if row[3] else None
            debit = parse_decimal(row[4])
            credit = parse_decimal(row[5])
            balance = parse_decimal(row[6])

            # Skip opening balance line
            if "Opening Balance" in narration:
                continue

            # Skip lines with both debit and credit as 0
            if (not debit or debit == 0) and (not credit or credit == 0):
                continue

            transactions.append(
                {
                    "line_number": len(transactions) + 1,
                    "date_posted": tran_date,
                    "value_date": value_date or tran_date,
                    "description": narration,
                    "reference": chq_no,
                    "debit": debit if debit and debit > 0 else None,
                    "credit": credit if credit and credit > 0 else None,
                    "balance": balance,
                }
            )

        logger.info(f"  Parsed {len(transactions)} transactions from {filepath.name}")

        return ParsedStatement(
            account_number=account_number,
            account_name=account_name,
            currency=currency,
            period_start=period_start,
            period_end=period_end,
            opening_balance=opening_balance or Decimal("0"),
            closing_balance=closing_balance or Decimal("0"),
            total_debit=total_debit or Decimal("0"),
            total_credit=total_credit or Decimal("0"),
            transactions=transactions,
            source_file=filepath.name,
        )

    except Exception as e:
        logger.error(f"Error parsing {filepath.name}: {e}")
        import traceback

        traceback.print_exc()
        return None


def _configured_files() -> list[Path]:
    """Every statement file this run will read, in configuration order."""
    return [
        STATEMENT_DIR / filename
        for config in ACCOUNT_CONFIG.values()
        for filename in config["files"]
    ]


def import_statements(
    *, organization_id: uuid.UUID, actor_id: uuid.UUID, dry_run: bool = False
) -> int:
    """Import every configured UBA account. Returns the number of failures."""

    mode = "DRY RUN" if dry_run else "EXECUTE"
    logger.info("=" * 60)
    logger.info("UBA Bank Statement Import (%s) — org %s", mode, organization_id)
    logger.info("=" * 60)

    checksum, per_file = file_manifest_digest(_configured_files())
    logger.info("Input manifest: %d file(s), digest %s", len(per_file), checksum[:12])

    with (
        session_for_org(organization_id) as db,
        batch_operation(
            db,
            organization_id=organization_id,
            operation_type=BatchOperationType.IMPORT,
            operation_name="import_uba_statements",
            started_by_id=actor_id,
            description=f"UBA statement import ({mode})",
            source_file=str(STATEMENT_DIR),
            source_checksum=checksum,
            metadata={"bank": BANK.bank_code, "files": per_file},
        ) as tally,
    ):
        org_id = organization_id
        total_imported = 0
        total_skipped = 0

        for account_number, config in ACCOUNT_CONFIG.items():
            logger.info("")
            logger.info(f"Processing account: {account_number} ({config['name']})")
            logger.info("-" * 40)

            # Ensure bank account exists
            if not dry_run:
                bank_account = ensure_bank_account(
                    db,
                    organization_id=org_id,
                    bank=BANK,
                    account=AccountProfile(
                        account_number=account_number,
                        account_name=config["name"],
                        currency_code=config["currency"],
                        gl_account_code=config["gl_account_code"],
                    ),
                )
            else:
                bank_account = find_bank_account(
                    db, organization_id=org_id, account_number=account_number
                )

            # Parse all statement files for this account
            all_transactions = []
            first_opening = None
            last_closing = None
            earliest_date = None
            latest_date = None

            for filename in config["files"]:
                filepath = STATEMENT_DIR / filename
                if not filepath.exists():
                    logger.warning(f"  File not found: {filename}")
                    continue

                parsed = parse_uba_statement(
                    filepath, _statement_password(account_number)
                )

                if parsed:
                    all_transactions.extend(parsed.transactions)

                    # Track opening/closing from first/last file
                    if first_opening is None:
                        first_opening = parsed.opening_balance
                        earliest_date = parsed.period_start
                    last_closing = parsed.closing_balance
                    latest_date = parsed.period_end

            if not all_transactions:
                logger.warning(f"  No transactions parsed for {account_number}")
                continue

            # Sort by date and remove duplicates
            all_transactions.sort(
                key=lambda x: (x["date_posted"], x["description"][:30])
            )

            # Deduplicate based on date + amount + description prefix
            seen = set()
            unique_transactions = []
            for txn in all_transactions:
                key = (
                    txn["date_posted"],
                    str(txn["debit"] or txn["credit"]),
                    txn["description"][:40],
                )
                if key not in seen:
                    seen.add(key)
                    unique_transactions.append(txn)

            logger.info(f"  Total transactions: {len(all_transactions)}")
            logger.info(f"  Unique transactions: {len(unique_transactions)}")

            # Determine period from transactions if not from headers
            dates = [t["date_posted"] for t in unique_transactions if t["date_posted"]]
            if dates:
                period_start = earliest_date or min(dates)
                period_end = latest_date or max(dates)
            else:
                continue

            opening_balance = first_opening or Decimal("0")
            closing_balance = last_closing or Decimal("0")

            # Create statement number
            statement_number = f"UBA-{account_number}-{period_start.strftime('%Y%m%d')}-{period_end.strftime('%Y%m%d')}"

            logger.info(f"  Period: {period_start} to {period_end}")
            logger.info(f"  Opening Balance: {opening_balance:,.2f}")
            logger.info(f"  Closing Balance: {closing_balance:,.2f}")
            logger.info(f"  Statement Number: {statement_number}")

            if dry_run:
                logger.info("  [DRY RUN] Would import statement")
                tally.skipped += len(unique_transactions)
                continue

            if not bank_account:
                logger.error("  Bank account not found")
                tally.failed += 1
                continue

            # Convert to statement lines
            lines = to_statement_lines(unique_transactions)

            # Import using BankStatementService
            service = BankStatementService()

            try:
                result = service.import_statement(
                    db=db,
                    organization_id=org_id,
                    bank_account_id=bank_account.bank_account_id,
                    statement_number=statement_number,
                    statement_date=period_end,
                    period_start=period_start,
                    period_end=period_end,
                    opening_balance=opening_balance,
                    closing_balance=closing_balance,
                    lines=lines,
                    import_source="uba_excel",
                    import_filename=", ".join(config["files"]),
                    check_duplicates=True,
                    skip_duplicates=True,
                )

                logger.info(f"  Imported: {result.lines_imported} lines")
                logger.info(f"  Skipped: {result.lines_skipped} lines")
                logger.info(f"  Duplicates: {result.duplicates_found}")

                total_imported += result.lines_imported
                total_skipped += result.lines_skipped
                tally.created += result.lines_imported
                tally.skipped += result.lines_skipped
                tally.track("bank_statement", result.statement.bank_statement_id)

                if result.errors:
                    tally.failed += len(result.errors)
                    for err in result.errors[:5]:
                        logger.warning(f"    Error: {err}")

                if result.warnings:
                    for warn in result.warnings[:5]:
                        logger.info(f"    Warning: {warn}")

            except Exception as e:
                # One unreadable account must not abandon the others, but the
                # run is NOT a success: the failure count reaches the batch
                # record and the exit status, so "Import Complete" stops being
                # printed over a partial import.
                logger.error(f"  Failed to import: {e}", exc_info=True)
                tally.failed += 1

        # NOTE: `batch_operation` commits when it marks the run COMPLETED, so a
        # dry run is safe only because the body writes nothing — the
        # `ensure_bank_account` call above is guarded and `import_statement` is
        # never reached. Do not add an unguarded write below this line expecting
        # the old "skip the commit and let the session roll back" behaviour;
        # that behaviour is gone.
        if not dry_run:
            db.commit()

        logger.info("")
        logger.info("=" * 60)
        logger.info(
            "Import %s: %d imported, %d skipped, %d failed",
            "complete" if not tally.failed else "FINISHED WITH FAILURES",
            total_imported,
            total_skipped,
            tally.failed,
        )
        logger.info("=" * 60)
        return tally.failed


def main() -> int:
    parser = argparse.ArgumentParser(description="Import UBA bank statements")
    parser.add_argument(
        "--org-id",
        required=True,
        type=uuid.UUID,
        help="Organization to import into (no default: this is multi-tenant)",
    )
    parser.add_argument(
        "--actor-id",
        type=uuid.UUID,
        default=SYSTEM_ACTOR_ID,
        help="Recorded as who ran this, on the BatchOperation record",
    )
    parser.add_argument(
        "--dry-run", action="store_true", help="Parse only, don't import"
    )
    args = parser.parse_args()

    failures = import_statements(
        organization_id=args.org_id, actor_id=args.actor_id, dry_run=args.dry_run
    )
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
