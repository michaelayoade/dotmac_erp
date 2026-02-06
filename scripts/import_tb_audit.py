#!/usr/bin/env python3
"""
Import audited Trial Balance data for Dotmac.

This script imports the TB closing balances for 2022, 2023, and 2024
as proper journal entries using the LedgerPostingService.
"""

import os
import sys
from datetime import date
from decimal import Decimal
from uuid import uuid4, UUID

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.db import SessionLocal
from app.models.finance.gl.account import Account, AccountType, NormalBalance
from app.models.finance.gl.account_category import AccountCategory, IFRSCategory
from app.models.finance.gl.fiscal_period import FiscalPeriod, PeriodStatus
from app.models.finance.gl.fiscal_year import FiscalYear
from app.models.finance.gl.journal_entry import JournalEntry, JournalStatus, JournalType
from app.models.finance.gl.journal_entry_line import JournalEntryLine
from app.models.finance.core_org.organization import Organization
from app.services.finance.gl.ledger_posting import (
    LedgerPostingService,
    PostingRequest,
)

# Dotmac org ID
ORG_ID = UUID("11d770b8-dc09-497c-9d67-c1b6b088dda5")
# Admin user (person_id from user_credentials)
ADMIN_USER_ID = UUID("b0dc82b1-6aba-4ad1-b931-dbe47e95a81c")
DOCS_PATH = "/app/docs"

# Account mapping: TB Name -> (account_code, category_code, normal_balance)
# Categories: FA=Fixed Assets, AD=Accum Dep, CA=Current Assets, CL=Current Liab,
#             LTL=Long Term Liab, EQ=Equity, REV=Revenue, COS=Cost of Sales, EXP=Expenses
ACCOUNT_MAP = {
    # Fixed Assets
    "Office Equipment": ("1110", "FA", "DEBIT"),
    "Motor Vehicle": ("1120", "FA", "DEBIT"),
    "Furniture & Fittings": ("1130", "FA", "DEBIT"),
    "Plant & Machinery": ("1140", "FA", "DEBIT"),
    # Accumulated Depreciation
    "Accumulated Depreciation - Office Equipment": ("1211", "AD", "CREDIT"),
    "Office Equipment:Depreciation": ("1211", "AD", "CREDIT"),  # alias
    "Accumulated Depreciation - Motor Vehicle": ("1212", "AD", "CREDIT"),
    "Accumulated Depreciation - Furniture & Fittings": ("1213", "AD", "CREDIT"),
    "Accumulated Depreciation - Plant & Machinery": ("1214", "AD", "CREDIT"),
    # Bank & Cash
    "Zenith Bank": ("1310", "BANK", "DEBIT"),
    "Heritage Bank": ("1320", "BANK", "DEBIT"),
    "UBA": ("1330", "BANK", "DEBIT"),
    "Paystack": ("1340", "BANK", "DEBIT"),
    "Paystack OPEX Account": ("1341", "BANK", "DEBIT"),
    "Flutterwave OPEX": ("1350", "BANK", "DEBIT"),
    "Fultterwave": ("1351", "BANK", "DEBIT"),  # typo in TB
    "First Bank": ("1360", "BANK", "DEBIT"),
    "Quick Teller": ("1370", "BANK", "DEBIT"),
    "Cash at Hand": ("1380", "CASH", "DEBIT"),
    "Undeposited Funds": ("1390", "CASH", "DEBIT"),
    # Receivables
    "Trade Receivables": ("1410", "AR", "DEBIT"),
    "Staff Loan": ("1420", "AR", "DEBIT"),
    "Prepayment": ("1430", "AR", "DEBIT"),
    "Goods in Transit": ("1440", "AR", "DEBIT"),
    # Tax Assets
    "Input VAT": ("1510", "TAX_A", "DEBIT"),
    "Withholding Taxes": ("1520", "TAX_A", "DEBIT"),
    "VAT Paid": ("1530", "TAX_A", "DEBIT"),
    # Inventory
    "Materials": ("1610", "INV", "DEBIT"),
    "Customer Terminal Devices": ("1620", "INV", "DEBIT"),
    # Payables
    "Trade Payables": ("2110", "AP", "CREDIT"),
    "Accurued Expenses": ("2120", "AP", "CREDIT"),
    "Payables & Other Liabilities": ("2130", "AP", "CREDIT"),
    "Employee Reimbursables": ("2140", "AP", "CREDIT"),
    "Directors Current Account": ("2150", "AP", "CREDIT"),
    "Unearned Revenue": ("2160", "AP", "CREDIT"),
    # Tax Liabilities
    "VAT (Out Put)": ("2210", "TAX_L", "CREDIT"),
    "VAT Payables": ("2220", "TAX_L", "CREDIT"),
    "Value Added Tax Withheld": ("2230", "TAX_L", "CREDIT"),
    "Payee": ("2240", "TAX_L", "CREDIT"),
    "Income Tax": ("2250", "TAX_L", "CREDIT"),
    "Education Tax": ("2260", "TAX_L", "CREDIT"),
    "IT Levy": ("2270", "TAX_L", "CREDIT"),
    "Tax Audit Liability": ("2280", "TAX_L", "CREDIT"),
    "NHF Payables": ("2290", "TAX_L", "CREDIT"),
    "Pension": ("2295", "TAX_L", "CREDIT"),
    # Long-term Liabilities
    "Long Term Borrowings": ("2510", "LTL", "CREDIT"),
    # Equity
    "Share Capital": ("3110", "EQ", "CREDIT"),
    "Issued and Fully Paid": ("3110", "EQ", "CREDIT"),  # alias
    "Retained Earnings": ("3210", "EQ", "CREDIT"),
    "Other Componrnt Equity": ("3310", "EQ", "CREDIT"),
    # Revenue
    "Revenue": ("4110", "REV", "CREDIT"),
    "Internet  Revenue": ("4120", "REV", "CREDIT"),
    "Other Business Revenue": ("4130", "REV", "CREDIT"),
    "Discount": ("4140", "REV", "CREDIT"),  # contra
    # Cost of Sales
    "Cost of Sales": ("5110", "COS", "DEBIT"),
    "Purchase of bandwitdh and Interconnect": ("5120", "COS", "DEBIT"),
    "Purchases": ("5130", "COS", "DEBIT"),
    "Discounts given - COS": ("5140", "COS", "DEBIT"),
    # Operating Expenses
    "Depreciation": ("6110", "EXP", "DEBIT"),
    "Staff Salaries & Wage": ("6120", "EXP", "DEBIT"),
    "Rent or Lease Payment": ("6130", "EXP", "DEBIT"),
    "Utilities": ("6140", "EXP", "DEBIT"),
    "Telephone bills": ("6150", "EXP", "DEBIT"),
    "IT & Internet Expenses": ("6160", "EXP", "DEBIT"),
    "Base Station Repairs and Maintenance": ("6170", "EXP", "DEBIT"),
    "Motor Vehichle Repairs & Maintenance": ("6180", "EXP", "DEBIT"),
    "Office Repairs & Maintenance": ("6190", "EXP", "DEBIT"),
    "Fuel & Lubricant": ("6200", "EXP", "DEBIT"),
    "Transportation & Travelling Expenses": ("6210", "EXP", "DEBIT"),
    "Accomodation Expenses": ("6220", "EXP", "DEBIT"),
    "Entertament": ("6230", "EXP", "DEBIT"),
    "Advertising Expenses": ("6240", "EXP", "DEBIT"),
    "Insurance Expenses": ("6250", "EXP", "DEBIT"),
    "Security Expenses": ("6260", "EXP", "DEBIT"),
    "Janitorial Expenses": ("6270", "EXP", "DEBIT"),
    "Printing & stationery": ("6280", "EXP", "DEBIT"),
    "Medical Expenses": ("6290", "EXP", "DEBIT"),
    "Staff Training": ("6300", "EXP", "DEBIT"),
    "Consultancy": ("6310", "EXP", "DEBIT"),
    "Legal & Professional Fee": ("6320", "EXP", "DEBIT"),
    "Audit Fee": ("6330", "EXP", "DEBIT"),
    "Commission & Fees": ("6340", "EXP", "DEBIT"),
    "Subscription & Renewal": ("6350", "EXP", "DEBIT"),
    "Membership dues": ("6360", "EXP", "DEBIT"),
    "Contract Tender Fees": ("6370", "EXP", "DEBIT"),
    "PAYE expenses": ("6380", "EXP", "DEBIT"),
    "NHF Charges/Expenses": ("6390", "EXP", "DEBIT"),
    "Statutory Expenses": ("6400", "EXP", "DEBIT"),
    "NCC Licence Fees": ("6410", "EXP", "DEBIT"),
    "NCC Operating Licence": ("6420", "EXP", "DEBIT"),
    "Installation and maintenance of fiber optic Network": ("6430", "EXP", "DEBIT"),
    "Equipment Rental": ("6440", "EXP", "DEBIT"),
    "Equipment rental": ("6440", "EXP", "DEBIT"),  # case variant
    "Other Expenses": ("6450", "EXP", "DEBIT"),
    "Bad Debt": ("6460", "EXP", "DEBIT"),
    "Shipping & Delivery Expenses": ("6470", "EXP", "DEBIT"),
    # Finance & Tax Expenses
    "Finance Cost": ("6510", "FIN", "DEBIT"),
    "Exchange gain or Loss": ("6520", "FIN", "DEBIT"),
    "Tax Expenses": ("6530", "FIN", "DEBIT"),
    "Tax Audit Expense": ("6540", "FIN", "DEBIT"),
    "Reconciliation Discrepancies": ("6550", "FIN", "DEBIT"),
    "Stampduty Deducted At Source": ("6560", "FIN", "DEBIT"),
}

# Category definitions: code -> (name, ifrs_category)
CATEGORY_DEFS = {
    "FA": ("Fixed Assets", IFRSCategory.ASSETS),
    "AD": ("Accumulated Depreciation", IFRSCategory.ASSETS),
    "BANK": ("Bank Accounts", IFRSCategory.ASSETS),
    "CASH": ("Cash", IFRSCategory.ASSETS),
    "AR": ("Accounts Receivable", IFRSCategory.ASSETS),
    "TAX_A": ("Tax Assets", IFRSCategory.ASSETS),
    "INV": ("Inventory", IFRSCategory.ASSETS),
    "AP": ("Accounts Payable", IFRSCategory.LIABILITIES),
    "TAX_L": ("Tax Liabilities", IFRSCategory.LIABILITIES),
    "LTL": ("Long-term Liabilities", IFRSCategory.LIABILITIES),
    "EQ": ("Equity", IFRSCategory.EQUITY),
    "REV": ("Revenue", IFRSCategory.REVENUE),
    "COS": ("Cost of Sales", IFRSCategory.EXPENSES),
    "EXP": ("Operating Expenses", IFRSCategory.EXPENSES),
    "FIN": ("Finance & Tax Expenses", IFRSCategory.EXPENSES),
}


def parse_decimal(val) -> Decimal:
    """Parse a value to Decimal, handling None and formulas."""
    if val is None:
        return Decimal("0")
    if isinstance(val, str):
        if val.startswith("="):
            return Decimal("0")
        val = val.replace(",", "").strip()
        if not val:
            return Decimal("0")
    try:
        return Decimal(str(val))
    except:
        return Decimal("0")


def read_tb(year: str) -> dict[str, Decimal]:
    """
    Read TB file and return closing balances.
    Returns dict of account_name -> closing_balance (positive=debit, negative=credit)
    """
    filepath = os.path.join(DOCS_PATH, f"{year} TB.xlsx")
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    balances = {}
    current_section = None  # Track section context

    for row in ws.iter_rows(values_only=True):
        if not row[0]:
            continue
        name = str(row[0]).strip()

        # Track section headers
        if name == "Accumulated Depreciation":
            current_section = "ACCUM_DEP"
            continue
        elif name in [
            "FIXED ASSETS",
            "Current Assets",
            "Payables & Other Liabilities",
            "EQUITY",
            "REVENUE",
            "COST OF SALES",
            "EXPENSES",
        ]:
            current_section = name
            continue

        # Skip headers, section titles, and notes
        if name.upper() == name:  # All caps = header
            continue
        if name.endswith(":"):  # Section title
            continue
        if len(name) > 80:  # Likely a note
            continue
        if name in ["ITEMS", "None", ""]:
            continue

        # Handle Accumulated Depreciation section - prepend to account name
        effective_name = name
        if current_section == "ACCUM_DEP":
            effective_name = f"Accumulated Depreciation - {name}"

        # Skip if not in our mapping
        if effective_name not in ACCOUNT_MAP:
            # Try original name as fallback
            if name not in ACCOUNT_MAP:
                continue
            effective_name = name

        # Parse: B=Opening Dr, C=Opening Cr, D=Curr Dr, E=Curr Cr, F=Adj Dr, G=Adj Cr
        opening_dr = parse_decimal(row[1])
        opening_cr = parse_decimal(row[2])
        curr_dr = parse_decimal(row[3])
        curr_cr = parse_decimal(row[4])
        adj_dr = parse_decimal(row[5])
        adj_cr = parse_decimal(row[6])

        # Calculate closing balance (positive = debit balance, negative = credit balance)
        closing = (opening_dr + curr_dr + adj_dr) - (opening_cr + curr_cr + adj_cr)

        # Filter out near-zero balances (floating point precision issues in Excel)
        if abs(closing) > Decimal("0.01"):
            # Accumulate if same account appears multiple times (e.g., Accrued Expenses)
            if effective_name in balances:
                balances[effective_name] += closing
            else:
                balances[effective_name] = closing

    wb.close()
    return balances


def ensure_categories(db, org_id: UUID) -> dict[str, UUID]:
    """Ensure all account categories exist. Returns code -> category_id mapping."""
    category_ids = {}

    for code, (name, ifrs_cat) in CATEGORY_DEFS.items():
        existing = (
            db.query(AccountCategory)
            .filter(
                AccountCategory.organization_id == org_id,
                AccountCategory.category_code == code,
            )
            .first()
        )

        if existing:
            category_ids[code] = existing.category_id
        else:
            cat = AccountCategory(
                category_id=uuid4(),
                organization_id=org_id,
                category_code=code,
                category_name=name,
                ifrs_category=ifrs_cat,
                is_active=True,
            )
            db.add(cat)
            db.flush()
            category_ids[code] = cat.category_id
            print(f"  Created category: {code} - {name}")

    return category_ids


def ensure_accounts(db, org_id: UUID, category_ids: dict[str, UUID]) -> dict[str, UUID]:
    """Ensure all accounts exist. Returns account_name -> account_id mapping."""
    account_ids = {}

    for acct_name, (acct_code, cat_code, normal_bal) in ACCOUNT_MAP.items():
        existing = (
            db.query(Account)
            .filter(
                Account.organization_id == org_id,
                Account.account_code == acct_code,
            )
            .first()
        )

        if existing:
            account_ids[acct_name] = existing.account_id
        else:
            cat_id = category_ids.get(cat_code)
            if not cat_id:
                print(f"  Warning: Category {cat_code} not found for {acct_name}")
                continue

            acct = Account(
                account_id=uuid4(),
                organization_id=org_id,
                account_code=acct_code,
                account_name=acct_name,
                category_id=cat_id,
                account_type=AccountType.POSTING,
                normal_balance=NormalBalance.DEBIT
                if normal_bal == "DEBIT"
                else NormalBalance.CREDIT,
                is_active=True,
                is_posting_allowed=True,
            )
            db.add(acct)
            db.flush()
            account_ids[acct_name] = acct.account_id
            print(f"  Created account: {acct_code} - {acct_name}")

    db.commit()
    return account_ids


def reopen_period(db, org_id: UUID, year: int, month: int) -> UUID:
    """Reopen a fiscal period and return its ID."""
    period = (
        db.query(FiscalPeriod)
        .join(FiscalYear)
        .filter(
            FiscalYear.organization_id == org_id,
            FiscalPeriod.start_date >= date(year, month, 1),
            FiscalPeriod.start_date < date(year, month + 1 if month < 12 else 1, 1),
        )
        .first()
    )

    if not period:
        raise ValueError(f"Period not found for {year}-{month:02d}")

    if period.status == PeriodStatus.HARD_CLOSED:
        period.status = PeriodStatus.OPEN
        db.flush()
        print(f"  Reopened period: {period.period_name}")

    return period.fiscal_period_id


def create_opening_balance_journal(
    db,
    org_id: UUID,
    year: int,
    balances: dict[str, Decimal],
    account_ids: dict[str, UUID],
) -> UUID:
    """Create an opening balance journal entry for a year."""
    # Get the January period for this year
    period_id = reopen_period(db, org_id, year, 1)

    entry_dt = date(year, 1, 1)

    # Create journal entry
    journal = JournalEntry(
        journal_entry_id=uuid4(),
        organization_id=org_id,
        fiscal_period_id=period_id,
        journal_type=JournalType.OPENING,
        journal_number=f"OB-{year}",
        entry_date=entry_dt,
        posting_date=entry_dt,
        description=f"Opening Balance {year} (from audited TB)",
        reference=f"TB-{year}",
        status=JournalStatus.DRAFT,
        currency_code="NGN",
        total_debit=Decimal("0"),
        total_credit=Decimal("0"),
        created_by_user_id=ADMIN_USER_ID,
    )
    db.add(journal)
    db.flush()

    # Create journal lines
    total_debit = Decimal("0")
    total_credit = Decimal("0")
    line_num = 0

    for acct_name, balance in balances.items():
        if acct_name not in account_ids:
            print(f"  Warning: Account not found: {acct_name}")
            continue

        # Skip zero balances
        if balance == Decimal("0"):
            continue

        line_num += 1
        acct_id = account_ids[acct_name]

        # Positive balance = debit, negative = credit
        if balance > 0:
            debit = balance
            credit = Decimal("0")
            total_debit += debit
        else:
            debit = Decimal("0")
            credit = abs(balance)
            total_credit += credit

        line = JournalEntryLine(
            line_id=uuid4(),
            journal_entry_id=journal.journal_entry_id,
            line_number=line_num,
            account_id=acct_id,
            description=f"Opening balance - {acct_name}",
            debit_amount=debit,
            credit_amount=credit,
            debit_amount_functional=debit,
            credit_amount_functional=credit,
        )
        db.add(line)

    # Check balance and add balancing entry to Retained Earnings if needed
    diff = total_debit - total_credit
    if abs(diff) > Decimal("1"):
        re_account_id = account_ids.get("Retained Earnings")
        if re_account_id:
            line_num += 1
            if diff > 0:
                # More debits (expenses) than credits (revenue) = Loss -> debit RE less
                # Actually: More debits means we need more credits to balance
                balance_line = JournalEntryLine(
                    line_id=uuid4(),
                    journal_entry_id=journal.journal_entry_id,
                    line_number=line_num,
                    account_id=re_account_id,
                    description="Current Year Profit (P&L closing to RE)",
                    debit_amount=Decimal("0"),
                    credit_amount=diff,
                    debit_amount_functional=Decimal("0"),
                    credit_amount_functional=diff,
                )
                total_credit += diff
            else:
                # More credits than debits = need more debits
                balance_line = JournalEntryLine(
                    line_id=uuid4(),
                    journal_entry_id=journal.journal_entry_id,
                    line_number=line_num,
                    account_id=re_account_id,
                    description="Current Year Loss (P&L closing to RE)",
                    debit_amount=abs(diff),
                    credit_amount=Decimal("0"),
                    debit_amount_functional=abs(diff),
                    credit_amount_functional=Decimal("0"),
                )
                total_debit += abs(diff)
            db.add(balance_line)
            print(f"  Added P&L closing entry to Retained Earnings: {diff:,.2f}")

    # Update journal totals
    journal.total_debit = total_debit
    journal.total_credit = total_credit
    db.flush()

    print(
        f"  Created journal {journal.journal_number}: Dr {total_debit:,.2f} / Cr {total_credit:,.2f}"
    )

    return journal.journal_entry_id


def post_journal(db, org_id: UUID, journal_id: UUID, posting_date: date):
    """Post a journal entry using LedgerPostingService."""
    request = PostingRequest(
        organization_id=org_id,
        journal_entry_id=journal_id,
        posting_date=posting_date,
        idempotency_key=f"ob-{posting_date.year}-{journal_id}",
        source_module="tb_import",
    )

    result = LedgerPostingService.post_journal_entry(db, request)

    if result.success:
        print(f"  Posted: {result.posted_lines} lines, Batch ID: {result.batch_id}")
    else:
        print(f"  ERROR posting: {result.message}")

    return result.success


def main():
    print("=" * 70)
    print("Importing Audited Trial Balance for Dotmac")
    print("=" * 70)

    db = SessionLocal()

    try:
        # Verify org exists
        org = (
            db.query(Organization)
            .filter(Organization.organization_id == ORG_ID)
            .first()
        )
        if not org:
            print(f"ERROR: Organization {ORG_ID} not found!")
            return

        print(f"\nOrganization: {org.legal_name}")

        # Ensure categories exist
        print("\n1. Ensuring account categories...")
        category_ids = ensure_categories(db, ORG_ID)
        db.commit()

        # Ensure accounts exist
        print("\n2. Ensuring accounts...")
        account_ids = ensure_accounts(db, ORG_ID, category_ids)

        # Process each year
        for year in [2022, 2023, 2024]:
            print(f"\n{'=' * 70}")
            print(f"Processing {year} Trial Balance")
            print("=" * 70)

            # Read TB
            print(f"\n3. Reading {year} TB...")
            balances = read_tb(str(year))
            print(f"  Found {len(balances)} accounts with balances")

            # Create opening balance journal
            print(f"\n4. Creating opening balance journal for {year}...")
            journal_id = create_opening_balance_journal(
                db, ORG_ID, year, balances, account_ids
            )
            db.commit()

            # Post journal
            print("\n5. Posting journal...")
            success = post_journal(db, ORG_ID, journal_id, date(year, 1, 1))
            db.commit()

            if not success:
                print("  Failed to post - check errors above")

        print("\n" + "=" * 70)
        print("Import complete!")
        print("=" * 70)

    except Exception as e:
        db.rollback()
        print(f"\nERROR: {e}")
        import traceback

        traceback.print_exc()
    finally:
        db.close()


if __name__ == "__main__":
    main()
