#!/usr/bin/env python
"""
Import Zenith Bank Statements from Excel files.

This script handles both Zenith statement formats:
- "Account Statement - Soft Copy" (older format, 2022-Sep 2024)
- "BOP_CBA_003_Report" (newer format, Oct 2024-Dec 2025)

Usage:
    poetry run python scripts/import_zenith_statements.py --org-id <uuid> [--dry-run]
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
import uuid
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl

from app.db.session_context import session_for_org
from app.models.batch_operation import BatchOperationType
from app.services.batch_operation import batch_operation, file_manifest_digest
from app.services.finance.banking.bank_statement import (
    BankStatementService,
)
from app.services.finance.banking.statement_import import (
    AccountProfile,
    BankProfile,
    ensure_bank_account,
    find_bank_account,
    to_statement_lines,
)
from app.services.finance.banking.statement_parsing import (
    extract_account_number,
    parse_date,
    parse_decimal,
    parse_period,
)

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

BANK = BankProfile(bank_name="Zenith Bank", bank_code="057")

SYSTEM_ACTOR_ID = uuid.UUID("00000000-0000-0000-0000-000000000000")

# Zenith's control account in the chart of accounts. UBA is 1202; these two are
# the pair `finish_interbank_matching` matches on (DR 1200 / CR 1202).
ZENITH_GL_ACCOUNT_CODE = "1200"

# Where the statement workbooks are read from. This was `/root/.dotmac/zenith
# statement` — an absolute, root-owned path no non-root run could read, and the
# kind of environment-specific literal `AGENTS.md` § "Everything by config"
# exists to keep out. Same knob shape as the UBA importer.
STATEMENT_DIR = Path(
    os.environ.get("ZENITH_STATEMENT_DIR", "~/.dotmac/statements/zenith")
).expanduser()

# Account mapping: account_number -> (name, currency, file_pairs)
# file_pairs = [(old_format_file, new_format_file), ...]
ACCOUNT_CONFIG = {
    "1011649523": {
        "name": "Zenith 523 (Main)",
        "account_name": "DOTMAC TECHNOLOGIES LTD",
        "currency": "NGN",
        "files": [
            "Account Statement - Soft Copy.xlsx",
            "BOP_CBA_003_Report.xlsx",
        ],
    },
    "1016946461": {
        "name": "Zenith 461 (Services)",
        "account_name": "DOTMAC TECHNOLOGIES LTD SERVICES",
        "currency": "NGN",
        "files": [
            "Account Statement - Soft Copy (2).xlsx",
            "BOP_CBA_003_Report (2).xlsx",
        ],
    },
    "1016946454": {
        "name": "Zenith 454 (Int Project)",
        "account_name": "DOTMAC TECHNOLOGIES INT PROJECT",
        "currency": "NGN",
        "files": [
            "Account Statement - Soft Copy (1).xlsx",
            "BOP_CBA_003_Report (1).xlsx",
        ],
    },
    "5070061296": {
        "name": "Zenith USD",
        "account_name": "DOTMAC TECHNOLOGIES",
        "currency": "USD",
        "files": [
            "Account Statement - Soft Copy (4).xlsx",
            "BOP_CBA_003_Report (5).xlsx",
        ],
    },
}


@dataclass
class ParsedStatement:
    """Parsed statement data from Excel file."""

    account_number: str
    account_name: str
    currency: str
    period_start: date
    period_end: date
    opening_balance: Decimal
    closing_balance: Decimal
    total_debit: Decimal
    total_credit: Decimal
    transactions: list[dict]
    source_file: str


def parse_old_format_statement(filepath: Path) -> ParsedStatement | None:
    """
    Parse older Zenith statement format (Account Statement - Soft Copy).

    Structure:
    - Row 5: Account name in col 0, "Account Number:" in col 6, account in col 8
    - Row 6: Currency in col 8
    - Row 7: Opening balance in col 8
    - Row 8: Total debit in col 8
    - Row 9: Total credit in col 8
    - Row 10: Closing balance in col 8
    - Row 11: Period in col 8
    - Row 13: Headers (DATE POSTED, VALUE DATE, DESCRIPTION, DEBIT, CREDIT, BALANCE)
    - Row 15+: Data (skip opening balance row)
    """
    logger.info(f"Parsing old format: {filepath.name}")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))

        # Extract header info
        account_name = str(rows[4][0]).strip() if rows[4][0] else ""
        account_number = extract_account_number(str(rows[4][8])) if rows[4][8] else None
        currency = str(rows[5][8]).strip() if rows[5][8] else "NGN"
        opening_balance = parse_decimal(rows[6][8])
        total_debit = parse_decimal(rows[7][8])
        total_credit = parse_decimal(rows[8][8])
        closing_balance = parse_decimal(rows[9][8])
        period_start, period_end = parse_period(str(rows[10][8]) if rows[10][8] else "")

        if not account_number:
            logger.warning(f"Could not extract account number from {filepath.name}")
            return None

        # Parse transactions (skip header rows and opening balance row)
        transactions = []
        for _i, row in enumerate(rows[15:], start=16):
            date_posted = parse_date(row[0])
            if not date_posted:
                continue  # Skip non-transaction rows

            value_date = parse_date(row[2])
            description = str(row[3]).strip() if row[3] else ""
            debit = parse_decimal(row[6])
            credit = parse_decimal(row[7])
            balance = parse_decimal(row[9])

            # Skip opening balance line
            if "OPENING BALANCE" in description.upper():
                continue

            transactions.append(
                {
                    "line_number": len(transactions) + 1,
                    "date_posted": date_posted,
                    "value_date": value_date or date_posted,
                    "description": description,
                    "debit": abs(debit) if debit else None,
                    "credit": credit,
                    "balance": balance,
                }
            )

        logger.info(f"  Parsed {len(transactions)} transactions from {filepath.name}")

        return ParsedStatement(
            account_number=account_number,
            account_name=account_name,
            currency=currency,
            period_start=period_start,
            period_end=period_end,
            opening_balance=opening_balance or Decimal("0"),
            closing_balance=closing_balance or Decimal("0"),
            total_debit=abs(total_debit) if total_debit else Decimal("0"),
            total_credit=total_credit or Decimal("0"),
            transactions=transactions,
            source_file=filepath.name,
        )

    except Exception as e:
        logger.error(f"Error parsing {filepath.name}: {e}")
        return None


def parse_new_format_statement(filepath: Path) -> ParsedStatement | None:
    """
    Parse newer Zenith statement format (BOP_CBA_003_Report).

    Structure:
    - Row 9: Account name in col 0, "Account Number:" in col 6, account in col 8
    - Row 10: Currency in col 8
    - Row 11: Opening balance in col 8
    - Row 12: Total debit in col 8
    - Row 13: Total credit in col 8
    - Row 14: Closing balance in col 8
    - Row 15: Period in col 8
    - Row 17: Headers
    - Row 20: Opening balance line (skip)
    - Row 21+: Data
    """
    logger.info(f"Parsing new format: {filepath.name}")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))

        # Extract header info (0-indexed, so row 9 = index 8)
        account_name = str(rows[8][0]).strip() if rows[8][0] else ""
        account_number = extract_account_number(str(rows[8][8])) if rows[8][8] else None
        currency = str(rows[9][8]).strip() if rows[9][8] else "NGN"
        opening_balance = parse_decimal(rows[10][8])
        total_debit = parse_decimal(rows[11][8])
        total_credit = parse_decimal(rows[12][8])
        closing_balance = parse_decimal(rows[13][8])
        period_start, period_end = parse_period(str(rows[14][8]) if rows[14][8] else "")

        if not account_number:
            logger.warning(f"Could not extract account number from {filepath.name}")
            return None

        # Parse transactions (start from row 21, index 20)
        transactions = []
        for _i, row in enumerate(rows[20:], start=21):
            # Parse date from string format like ' 16/10/2024'
            date_str = str(row[0]).strip() if row[0] else ""
            date_posted = parse_date(date_str)

            if not date_posted:
                continue  # Skip non-transaction rows

            value_date_str = str(row[2]).strip() if row[2] else ""
            value_date = parse_date(value_date_str)
            description = str(row[3]).strip() if row[3] else ""

            # In new format: debit in col 6, credit in col 7
            debit = parse_decimal(row[6])
            credit = parse_decimal(row[7])
            balance_str = str(row[9]).strip().replace(",", "") if row[9] else ""
            balance = parse_decimal(balance_str)

            # Skip opening balance line
            if "Opening Balance" in description:
                continue

            # Convert: debit > 0 means debit, credit > 0 means credit
            actual_debit = debit if debit and debit > 0 else None
            actual_credit = credit if credit and credit > 0 else None

            transactions.append(
                {
                    "line_number": len(transactions) + 1,
                    "date_posted": date_posted,
                    "value_date": value_date or date_posted,
                    "description": description,
                    "debit": actual_debit,
                    "credit": actual_credit,
                    "balance": balance,
                }
            )

        logger.info(f"  Parsed {len(transactions)} transactions from {filepath.name}")

        return ParsedStatement(
            account_number=account_number,
            account_name=account_name,
            currency=currency,
            period_start=period_start,
            period_end=period_end,
            opening_balance=opening_balance or Decimal("0"),
            closing_balance=closing_balance or Decimal("0"),
            total_debit=total_debit or Decimal("0"),
            total_credit=total_credit or Decimal("0"),
            transactions=transactions,
            source_file=filepath.name,
        )

    except Exception as e:
        logger.error(f"Error parsing {filepath.name}: {e}")
        import traceback

        traceback.print_exc()
        return None


def _configured_files() -> list[Path]:
    """Every statement file this run will read, in configuration order."""
    return [
        STATEMENT_DIR / filename
        for config in ACCOUNT_CONFIG.values()
        for filename in config["files"]
    ]


def import_statements(
    *, organization_id: uuid.UUID, actor_id: uuid.UUID, dry_run: bool = False
) -> int:
    """Import every configured Zenith account. Returns the number of failures."""

    mode = "DRY RUN" if dry_run else "EXECUTE"
    logger.info("=" * 60)
    logger.info("Zenith Bank Statement Import (%s) — org %s", mode, organization_id)
    logger.info("=" * 60)

    checksum, per_file = file_manifest_digest(_configured_files())
    logger.info("Input manifest: %d file(s), digest %s", len(per_file), checksum[:12])

    with (
        session_for_org(organization_id) as db,
        batch_operation(
            db,
            organization_id=organization_id,
            operation_type=BatchOperationType.IMPORT,
            operation_name="import_zenith_statements",
            started_by_id=actor_id,
            description=f"Zenith statement import ({mode})",
            source_file=str(STATEMENT_DIR),
            source_checksum=checksum,
            metadata={"bank": BANK.bank_code, "files": per_file},
        ) as tally,
    ):
        org_id = organization_id
        total_imported = 0
        total_skipped = 0

        for account_number, config in ACCOUNT_CONFIG.items():
            logger.info("")
            logger.info(f"Processing account: {account_number} ({config['name']})")
            logger.info("-" * 40)

            # Creating the bank account is a WRITE, so a dry run must not do it.
            # This call used to be unguarded and was harmless only by accident:
            # dry runs skipped the final `db.commit()` and the session closed
            # with a rollback. `batch_operation` commits when it completes the
            # run, so that accident is gone and the guard is now load-bearing.
            if dry_run:
                bank_account = find_bank_account(
                    db, organization_id=org_id, account_number=account_number
                )
            else:
                bank_account = ensure_bank_account(
                    db,
                    organization_id=org_id,
                    bank=BANK,
                    account=AccountProfile(
                        account_number=account_number,
                        # The registered name at the bank, not the internal
                        # label in `name` — this is what appears on the account.
                        account_name=config["account_name"],
                        currency_code=config["currency"],
                        gl_account_code=ZENITH_GL_ACCOUNT_CODE,
                    ),
                )

            # Parse all statement files for this account
            all_transactions = []

            for filename in config["files"]:
                filepath = STATEMENT_DIR / filename
                if not filepath.exists():
                    logger.warning(f"  File not found: {filename}")
                    continue

                # Determine format and parse
                if filename.startswith("Account Statement"):
                    parsed = parse_old_format_statement(filepath)
                else:
                    parsed = parse_new_format_statement(filepath)

                if parsed:
                    all_transactions.extend(parsed.transactions)

            if not all_transactions:
                logger.warning(f"  No transactions parsed for {account_number}")
                continue

            # Sort by date and remove duplicates
            all_transactions.sort(key=lambda x: (x["date_posted"], x["description"]))

            # Deduplicate based on date + amount + description
            seen = set()
            unique_transactions = []
            for txn in all_transactions:
                key = (
                    txn["date_posted"],
                    str(txn["debit"] or txn["credit"]),
                    txn["description"][:50],
                )
                if key not in seen:
                    seen.add(key)
                    unique_transactions.append(txn)

            logger.info(f"  Total transactions: {len(all_transactions)}")
            logger.info(f"  Unique transactions: {len(unique_transactions)}")

            # Determine overall period
            dates = [t["date_posted"] for t in unique_transactions if t["date_posted"]]
            if not dates:
                continue

            period_start = min(dates)
            period_end = max(dates)

            # Calculate opening/closing from first/last transaction
            first_txn = unique_transactions[0]
            last_txn = unique_transactions[-1]

            # Calculate opening balance
            first_amount = first_txn["debit"] or first_txn["credit"] or Decimal("0")
            first_balance = first_txn.get("balance") or Decimal("0")
            if first_txn["debit"]:
                opening_balance = first_balance + first_amount
            else:
                opening_balance = first_balance - first_amount

            closing_balance = last_txn.get("balance") or Decimal("0")

            # Create statement number
            statement_number = f"ZENITH-{account_number}-{period_start.strftime('%Y%m%d')}-{period_end.strftime('%Y%m%d')}"

            logger.info(f"  Period: {period_start} to {period_end}")
            logger.info(f"  Opening Balance: {opening_balance:,.2f}")
            logger.info(f"  Closing Balance: {closing_balance:,.2f}")
            logger.info(f"  Statement Number: {statement_number}")

            if dry_run:
                logger.info("  [DRY RUN] Would import statement")
                tally.skipped += len(unique_transactions)
                continue

            if not bank_account:
                logger.error("  Bank account not found")
                tally.failed += 1
                continue

            # Convert to statement lines
            lines = to_statement_lines(unique_transactions)

            # Import using BankStatementService
            service = BankStatementService()

            try:
                result = service.import_statement(
                    db=db,
                    organization_id=org_id,
                    bank_account_id=bank_account.bank_account_id,
                    statement_number=statement_number,
                    statement_date=period_end,
                    period_start=period_start,
                    period_end=period_end,
                    opening_balance=opening_balance,
                    closing_balance=closing_balance,
                    lines=lines,
                    import_source="zenith_excel",
                    import_filename=", ".join(config["files"]),
                    check_duplicates=True,
                    skip_duplicates=True,
                )

                logger.info(f"  Imported: {result.lines_imported} lines")
                logger.info(f"  Skipped: {result.lines_skipped} lines")
                logger.info(f"  Duplicates: {result.duplicates_found}")

                total_imported += result.lines_imported
                total_skipped += result.lines_skipped
                tally.created += result.lines_imported
                tally.skipped += result.lines_skipped
                tally.track("bank_statement", result.statement.bank_statement_id)

                if result.errors:
                    tally.failed += len(result.errors)
                    for err in result.errors[:5]:
                        logger.warning(f"    Error: {err}")

                if result.warnings:
                    for warn in result.warnings[:5]:
                        logger.info(f"    Warning: {warn}")

            except Exception as e:
                # One unreadable account must not abandon the others, but the
                # run is NOT a success: the failure count reaches the batch
                # record and the exit status.
                logger.error(f"  Failed to import: {e}", exc_info=True)
                tally.failed += 1

        # NOTE: `batch_operation` commits when it marks the run COMPLETED, so a
        # dry run is safe only because the body writes nothing. That is why
        # `ensure_bank_account` above is guarded — see the comment there.
        if not dry_run:
            db.commit()

        logger.info("")
        logger.info("=" * 60)
        logger.info(
            "Import %s: %d imported, %d skipped, %d failed",
            "complete" if not tally.failed else "FINISHED WITH FAILURES",
            total_imported,
            total_skipped,
            tally.failed,
        )
        logger.info("=" * 60)
        return tally.failed


def main() -> int:
    parser = argparse.ArgumentParser(description="Import Zenith bank statements")
    parser.add_argument(
        "--org-id",
        required=True,
        type=uuid.UUID,
        help="Organization to import into (no default: this is multi-tenant)",
    )
    parser.add_argument(
        "--actor-id",
        type=uuid.UUID,
        default=SYSTEM_ACTOR_ID,
        help="Recorded as who ran this, on the BatchOperation record",
    )
    parser.add_argument(
        "--dry-run", action="store_true", help="Parse only, don't import"
    )
    args = parser.parse_args()

    failures = import_statements(
        organization_id=args.org_id, actor_id=args.actor_id, dry_run=args.dry_run
    )
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
