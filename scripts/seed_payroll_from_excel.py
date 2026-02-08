"""
Seed Payroll Data from Excel - January 2026 PAYE Spreadsheet.

This script:
1. Creates a batch operation record for full audit trail
2. Clears existing payroll data (slips, assignments, structures, components, entries)
3. Creates salary components (Basic, Housing, Transport, etc.)
4. Creates salary structures (Permanent Staff, Contract Staff)
5. Looks up existing employees by name OR creates new ones
6. Creates salary assignments linking employees to structures
7. Seeds NTA 2025 tax bands

Usage:
    poetry run python scripts/seed_payroll_from_excel.py

    # To match existing employees only (no new creations):
    poetry run python scripts/seed_payroll_from_excel.py --match-only

    # To see what would happen without making changes:
    poetry run python scripts/seed_payroll_from_excel.py --dry-run
"""

import argparse
import hashlib
import sys
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from datetime import date
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from uuid import UUID

import openpyxl
from sqlalchemy import func, select, text
from sqlalchemy.orm import Session

from app.db import SessionLocal
from app.models.batch_operation import BatchOperation, BatchOperationType
from app.models.people.hr.department import Department
from app.models.people.hr.designation import Designation
from app.models.people.hr.employee import Employee, EmployeeStatus
from app.models.people.hr.employment_type import EmploymentType
from app.models.people.payroll.salary_assignment import SalaryStructureAssignment
from app.models.people.payroll.salary_component import (
    SalaryComponent,
    SalaryComponentType,
)
from app.models.people.payroll.salary_structure import (
    PayrollFrequency,
    SalaryStructure,
    SalaryStructureDeduction,
    SalaryStructureEarning,
)
from app.models.person import Person, PersonStatus
from app.services.people.payroll.paye_calculator import PAYECalculator

# Excel file path - works both locally and in Docker
EXCEL_PATH = (
    Path("/app/jan paye.xlsx")
    if Path("/app/jan paye.xlsx").exists()
    else Path("/root/.dotmac/jan paye.xlsx")
)

# Division to Department mapping
DIVISION_MAP = {
    "1. Executive": "Executive",
    "2. Technology & NOC": "Technology & NOC",
    "3. Customer Experience": "Customer Experience",
    "4. Commercial": "Commercial",
    "5. Service Delivery": "Service Delivery",
    "6. Finance": "Finance",
    "7. Corporate Services": "Corporate Services",
}


def get_file_checksum(file_path: Path) -> str:
    """Calculate SHA256 checksum of a file."""
    sha256_hash = hashlib.sha256()
    with open(file_path, "rb") as f:
        for byte_block in iter(lambda: f.read(4096), b""):
            sha256_hash.update(byte_block)
    return sha256_hash.hexdigest()


def get_org_id(db: Session) -> UUID:
    """Get the first organization ID."""
    result = db.execute(
        text("SELECT organization_id FROM core_org.organization LIMIT 1")
    )
    row = result.fetchone()
    if not row:
        raise ValueError("No organization found. Please seed organization first.")
    return row[0]


def get_admin_user_id(db: Session) -> UUID:
    """Get admin user ID for audit fields."""
    result = db.execute(
        text(
            "SELECT person_id FROM public.user_credentials WHERE username = 'admin' LIMIT 1"
        )
    )
    row = result.fetchone()
    if row:
        return row[0]
    # Fallback to first user
    result = db.execute(text("SELECT person_id FROM public.user_credentials LIMIT 1"))
    row = result.fetchone()
    if not row:
        raise ValueError("No users found.")
    return row[0]


def clear_payroll_data(db: Session, org_id: UUID):
    """Clear all existing payroll data for the organization."""
    print("Clearing existing payroll data...")

    # Delete in order of dependencies
    db.execute(
        text("""
        DELETE FROM payroll.salary_slip_deduction
        WHERE slip_id IN (SELECT slip_id FROM payroll.salary_slip WHERE organization_id = :org_id)
    """),
        {"org_id": org_id},
    )

    db.execute(
        text("""
        DELETE FROM payroll.salary_slip_earning
        WHERE slip_id IN (SELECT slip_id FROM payroll.salary_slip WHERE organization_id = :org_id)
    """),
        {"org_id": org_id},
    )

    db.execute(
        text("DELETE FROM payroll.salary_slip WHERE organization_id = :org_id"),
        {"org_id": org_id},
    )
    db.execute(
        text("DELETE FROM payroll.payroll_entry WHERE organization_id = :org_id"),
        {"org_id": org_id},
    )
    db.execute(
        text(
            "DELETE FROM payroll.salary_structure_assignment WHERE organization_id = :org_id"
        ),
        {"org_id": org_id},
    )

    db.execute(
        text("""
        DELETE FROM payroll.salary_structure_earning
        WHERE structure_id IN (SELECT structure_id FROM payroll.salary_structure WHERE organization_id = :org_id)
    """),
        {"org_id": org_id},
    )

    db.execute(
        text("""
        DELETE FROM payroll.salary_structure_deduction
        WHERE structure_id IN (SELECT structure_id FROM payroll.salary_structure WHERE organization_id = :org_id)
    """),
        {"org_id": org_id},
    )

    db.execute(
        text("DELETE FROM payroll.salary_structure WHERE organization_id = :org_id"),
        {"org_id": org_id},
    )
    db.execute(
        text("DELETE FROM payroll.salary_component WHERE organization_id = :org_id"),
        {"org_id": org_id},
    )
    db.execute(
        text(
            "DELETE FROM payroll.employee_tax_profile WHERE organization_id = :org_id"
        ),
        {"org_id": org_id},
    )
    db.execute(
        text("DELETE FROM payroll.tax_band WHERE organization_id = :org_id"),
        {"org_id": org_id},
    )

    db.commit()
    print("  Done.")


def create_components(
    db: Session, org_id: UUID, user_id: UUID
) -> dict[str, SalaryComponent]:
    """Create salary components."""
    print("Creating salary components...")

    components_def = [
        # Earnings
        ("BASIC", "Basic Salary", "Basic", SalaryComponentType.EARNING, True, 1),
        ("HOUSING", "Housing Allowance", "Hsg", SalaryComponentType.EARNING, True, 2),
        (
            "TRANSPORT",
            "Transport Allowance",
            "Trsp",
            SalaryComponentType.EARNING,
            True,
            3,
        ),
        ("OTHER", "Other Allowances", "Other", SalaryComponentType.EARNING, True, 4),
        # Deductions
        (
            "PENSION",
            "Employee Pension (8%)",
            "Pen",
            SalaryComponentType.DEDUCTION,
            False,
            10,
        ),
        (
            "NHF",
            "National Housing Fund (2.5%)",
            "NHF",
            SalaryComponentType.DEDUCTION,
            False,
            11,
        ),
        ("PAYE", "PAYE Tax", "PAYE", SalaryComponentType.DEDUCTION, False, 12),
    ]

    components = {}
    for code, name, abbr, comp_type, is_taxable, order in components_def:
        comp = SalaryComponent(
            organization_id=org_id,
            component_code=code,
            component_name=name,
            abbr=abbr,
            component_type=comp_type,
            is_tax_applicable=is_taxable,
            is_statutory=(code in ("PENSION", "NHF", "PAYE")),
            depends_on_payment_days=(comp_type == SalaryComponentType.EARNING),
            display_order=order,
            created_by_id=user_id,
        )
        db.add(comp)
        components[code] = comp

    db.flush()
    print(f"  Created {len(components)} components.")
    return components


def create_structures(
    db: Session, org_id: UUID, user_id: UUID, components: dict[str, SalaryComponent]
) -> tuple[SalaryStructure, SalaryStructure]:
    """Create salary structures for permanent and contract staff."""
    print("Creating salary structures...")

    # Permanent Staff Structure - with formula-based earnings and deductions
    perm_struct = SalaryStructure(
        organization_id=org_id,
        structure_code="PERM-STAFF",
        structure_name="Permanent Staff",
        description="Full-time employees with PAYE, Pension, and NHF deductions",
        payroll_frequency=PayrollFrequency.MONTHLY,
        currency_code="NGN",
        is_active=True,
        created_by_id=user_id,
    )
    db.add(perm_struct)
    db.flush()

    # Add earnings to permanent structure (formula-based)
    perm_earnings = [
        (components["BASIC"], "base * 0.30", 1),  # 30% of gross
        (components["HOUSING"], "base * 0.15", 2),  # 15% of gross
        (components["TRANSPORT"], "base * 0.10", 3),  # 10% of gross
        (components["OTHER"], "base * 0.45", 4),  # 45% of gross
    ]
    for comp, formula, order in perm_earnings:
        earning = SalaryStructureEarning(
            structure_id=perm_struct.structure_id,
            component_id=comp.component_id,
            amount_based_on_formula=True,
            formula=formula,
            display_order=order,
        )
        db.add(earning)

    # Add deductions to permanent structure
    perm_deductions = [
        (components["PENSION"], "basic * 0.08", 10),  # 8% of basic
        (components["NHF"], "basic * 0.025", 11),  # 2.5% of basic
        (components["PAYE"], None, 12),  # Calculated by PAYE calculator
    ]
    for comp, formula, order in perm_deductions:
        deduction = SalaryStructureDeduction(
            structure_id=perm_struct.structure_id,
            component_id=comp.component_id,
            amount_based_on_formula=bool(formula),
            formula=formula,
            display_order=order,
        )
        db.add(deduction)

    # Contract Staff Structure - earnings only, no deductions
    contract_struct = SalaryStructure(
        organization_id=org_id,
        structure_code="CONTRACT-STAFF",
        structure_name="Contract Staff",
        description="Contract workers - flat pay, no statutory deductions",
        payroll_frequency=PayrollFrequency.MONTHLY,
        currency_code="NGN",
        is_active=True,
        created_by_id=user_id,
    )
    db.add(contract_struct)
    db.flush()

    # Contract staff: single earning (base = gross)
    contract_earning = SalaryStructureEarning(
        structure_id=contract_struct.structure_id,
        component_id=components["BASIC"].component_id,
        amount_based_on_formula=True,
        formula="base",  # Full amount as basic
        display_order=1,
    )
    db.add(contract_earning)

    db.flush()
    print(
        f"  Created structures: {perm_struct.structure_name}, {contract_struct.structure_name}"
    )
    return perm_struct, contract_struct


def get_or_create_department(
    db: Session, org_id: UUID, name: str, user_id: UUID
) -> Department:
    """Get or create a department."""
    dept = (
        db.query(Department)
        .filter(
            Department.organization_id == org_id,
            Department.department_name == name,
        )
        .first()
    )

    if not dept:
        dept = Department(
            organization_id=org_id,
            department_code=name.upper().replace(" ", "-")[:20],
            department_name=name,
            is_active=True,
            created_by_id=user_id,
        )
        db.add(dept)
        db.flush()

    return dept


def get_or_create_designation(
    db: Session, org_id: UUID, title: str, user_id: UUID
) -> Designation:
    """Get or create a designation."""
    desig = (
        db.query(Designation)
        .filter(
            Designation.organization_id == org_id,
            Designation.designation_name == title,
        )
        .first()
    )

    if not desig:
        desig = Designation(
            organization_id=org_id,
            designation_code=title.upper().replace(" ", "-")[:20],
            designation_name=title,
            created_by_id=user_id,
        )
        db.add(desig)
        db.flush()

    return desig


def get_or_create_employment_type(
    db: Session, org_id: UUID, type_name: str, user_id: UUID
) -> EmploymentType:
    """Get or create an employment type."""
    emp_type = (
        db.query(EmploymentType)
        .filter(
            EmploymentType.organization_id == org_id,
            EmploymentType.type_name == type_name,
        )
        .first()
    )

    if not emp_type:
        emp_type = EmploymentType(
            organization_id=org_id,
            type_code=type_name.upper().replace(" ", "_"),
            type_name=type_name,
            created_by_id=user_id,
        )
        db.add(emp_type)
        db.flush()

    return emp_type


def excel_to_decimal(
    value: float | int | str | None, decimal_places: int = 2
) -> Decimal:
    """
    Safely convert Excel cell value to Decimal with proper rounding.

    Excel stores numbers as IEEE 754 floats, which can lead to precision issues
    (e.g., 200000 stored as 199999.99999999997). This function rounds
    the value before converting to Decimal to avoid floating-point artifacts.

    Args:
        value: The cell value from Excel (can be float, int, str, or None)
        decimal_places: Number of decimal places to round to (default 2 for currency)

    Returns:
        Decimal value, properly rounded

    Raises:
        ValueError: If the value cannot be converted to a valid number
    """
    quantizer = Decimal(10) ** -decimal_places

    if value is None:
        return Decimal("0").quantize(quantizer)

    # Handle string values (may have commas, currency symbols, etc.)
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return Decimal("0").quantize(quantizer)

        # Remove common formatting: commas, currency symbols, spaces
        cleaned = cleaned.replace(",", "").replace("₦", "").replace("$", "").strip()

        # Check for non-numeric placeholder values
        if cleaned.upper() in ("N/A", "TBD", "-", "NIL", "NONE"):
            raise ValueError(f"Non-numeric value in salary field: '{value}'")

        try:
            # Convert to float first to handle scientific notation, then round
            numeric = round(float(cleaned), decimal_places)
            return Decimal(str(numeric)).quantize(quantizer, rounding=ROUND_HALF_UP)
        except (ValueError, InvalidOperation) as e:
            raise ValueError(f"Cannot convert '{value}' to Decimal: {e}") from e

    # For numbers (float/int), round first to avoid floating-point artifacts
    # Example: 199999.99999999997 → round → 200000.0 → Decimal("200000.00")
    rounded = round(float(value), decimal_places)
    return Decimal(str(rounded)).quantize(quantizer, rounding=ROUND_HALF_UP)


def parse_excel_data(excel_path: Path) -> tuple[list[dict], list[dict]]:
    """Parse the Excel file and extract employee data."""
    print(f"Reading Excel file: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    permanent_staff = []
    contract_staff = []

    # Parse Permanent Staff sheet
    ws = wb["Payroll(January 2026)"]
    for row in ws.iter_rows(min_row=4, values_only=True):  # Skip headers
        if row[1] and row[1] != "TOTALS" and isinstance(row[0], (int, float)):
            name = str(row[1]).strip() if row[1] else ""
            if not name:
                continue

            division = str(row[2]).strip() if row[2] else ""
            role = str(row[3]).strip() if row[3] else ""
            category = str(row[4]).strip() if row[4] else "Staff"
            monthly_gross = excel_to_decimal(row[6])

            permanent_staff.append(
                {
                    "name": name,
                    "division": division,
                    "role": role,
                    "category": category,
                    "monthly_gross": monthly_gross,
                    "is_nysc": category.upper() == "NYSC",
                }
            )

    # Parse Contract Staff sheet
    ws = wb["Payroll (January Contract)"]
    for row in ws.iter_rows(min_row=4, values_only=True):  # Skip headers
        if row[1] and isinstance(row[0], (int, float)):
            name = str(row[1]).strip() if row[1] else ""
            if not name:
                continue

            division = str(row[2]).strip() if row[2] else ""
            role = str(row[3]).strip() if row[3] else ""
            category = str(row[4]).strip() if row[4] else ""
            # Contract staff: use column 6 if available, else column 5 (Current Take-Home)
            monthly_gross = (
                excel_to_decimal(row[6]) if row[6] else excel_to_decimal(row[5])
            )

            contract_staff.append(
                {
                    "name": name,
                    "division": division,
                    "role": role,
                    "category": category,
                    "monthly_gross": monthly_gross,
                    "is_nysc": category.upper() == "NYSC",
                }
            )

    print(
        f"  Found {len(permanent_staff)} permanent staff, {len(contract_staff)} contract staff"
    )
    return permanent_staff, contract_staff


def normalize_name(name: str) -> str:
    """Normalize a name for comparison (lowercase, no extra spaces)."""
    return " ".join(name.lower().split())


def find_existing_employee(
    db: Session,
    org_id: UUID,
    name: str,
) -> Employee | None:
    """
    Find an existing employee by name within the organization.

    Matches by first_name + last_name on the linked Person record.
    """
    name_parts = name.split()
    first_name = name_parts[0] if name_parts else ""
    last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""

    # Try exact match first
    stmt = (
        select(Employee)
        .join(Person, Employee.person_id == Person.id)
        .where(
            Employee.organization_id == org_id,
            Employee.is_deleted == False,
            func.lower(Person.first_name) == first_name.lower(),
            func.lower(Person.last_name) == last_name.lower(),
        )
    )
    employee = db.scalar(stmt)
    if employee:
        return employee

    # Try matching by email pattern
    email_name = name.lower().replace(" ", ".").replace("-", "")
    email = f"{email_name}@dotmac.ng"

    stmt = (
        select(Employee)
        .join(Person, Employee.person_id == Person.id)
        .where(
            Employee.organization_id == org_id,
            Employee.is_deleted == False,
            func.lower(Person.email) == email.lower(),
        )
    )
    employee = db.scalar(stmt)
    if employee:
        return employee

    return None


def find_or_create_employee(
    db: Session,
    org_id: UUID,
    user_id: UUID,
    data: dict,
    emp_type: EmploymentType,
    employee_code: str,
    batch_operation_id: UUID,
    match_only: bool = False,
) -> tuple[Employee | None, str]:
    """
    Find existing employee or create new one.

    Returns:
        Tuple of (Employee or None, status) where status is one of:
        - "found": Existing employee matched
        - "created": New employee created
        - "skipped": Could not find and match_only=True
    """
    # First, try to find existing employee
    existing_emp = find_existing_employee(db, org_id, data["name"])

    if existing_emp:
        return existing_emp, "found"

    if match_only:
        return None, "skipped"

    # Create new employee
    name_parts = data["name"].split()
    first_name = name_parts[0] if name_parts else "Unknown"
    last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else "Staff"

    # Generate email from name
    email_name = data["name"].lower().replace(" ", ".").replace("-", "")
    email = f"{email_name}@dotmac.ng"

    # Check if person with this email already exists IN THIS ORGANIZATION
    existing_person = (
        db.query(Person)
        .filter(
            Person.email == email,
            Person.organization_id == org_id,  # CRITICAL: Filter by org_id
        )
        .first()
    )

    if existing_person:
        # Check if employee exists for this person IN THIS ORGANIZATION
        existing_emp = (
            db.query(Employee)
            .filter(
                Employee.person_id == existing_person.id,
                Employee.organization_id == org_id,  # CRITICAL: Filter by org_id
            )
            .first()
        )
        if existing_emp:
            return existing_emp, "found"
        person = existing_person
    else:
        # Check if email exists globally (different org)
        global_person = db.query(Person).filter(Person.email == email).first()
        if global_person:
            # Email exists in different org - make unique
            counter = 1
            while True:
                new_email = f"{email_name}.{counter}@dotmac.ng"
                if not db.query(Person).filter(Person.email == new_email).first():
                    email = new_email
                    print(f"    ⚠ Email collision, using: {email}")
                    break
                counter += 1

        # Create Person
        person = Person(
            organization_id=org_id,
            first_name=first_name,
            last_name=last_name,
            email=email,
            email_verified=True,
            status=PersonStatus.active,
            is_active=True,
            batch_operation_id=batch_operation_id,  # Track which batch created this
        )
        db.add(person)
        db.flush()

    # Get department
    dept_name = DIVISION_MAP.get(data["division"], "General")
    dept = get_or_create_department(db, org_id, dept_name, user_id)

    # Get designation
    desig = None
    if data["role"]:
        desig = get_or_create_designation(db, org_id, data["role"], user_id)

    emp = Employee(
        organization_id=org_id,
        person_id=person.id,
        employee_code=employee_code,
        status=EmployeeStatus.ACTIVE,
        date_of_joining=date(2024, 1, 1),  # Default
        department_id=dept.department_id,
        designation_id=desig.designation_id if desig else None,
        employment_type_id=emp_type.employment_type_id,
        created_by_id=user_id,
        batch_operation_id=batch_operation_id,  # Track which batch created this
    )
    db.add(emp)
    db.flush()

    return emp, "created"


def create_assignment(
    db: Session,
    org_id: UUID,
    user_id: UUID,
    employee: Employee,
    structure: SalaryStructure,
    monthly_gross: Decimal,
    batch_operation_id: UUID,
):
    """Create a salary structure assignment."""
    assignment = SalaryStructureAssignment(
        organization_id=org_id,
        employee_id=employee.employee_id,
        structure_id=structure.structure_id,
        from_date=date(2026, 1, 1),
        base=monthly_gross,
        created_by_id=user_id,
        batch_operation_id=batch_operation_id,  # Track which batch created this
    )
    db.add(assignment)


def seed_tax_bands(db: Session, org_id: UUID, user_id: UUID):
    """Seed NTA 2025 tax bands."""
    print("Seeding NTA 2025 tax bands...")
    calculator = PAYECalculator(db)
    bands = calculator.seed_nta_2025_bands(org_id, date(2026, 1, 1), user_id)
    print(f"  Created {len(bands)} tax bands.")


def main():
    parser = argparse.ArgumentParser(description="Seed payroll data from Excel")
    parser.add_argument(
        "--match-only",
        action="store_true",
        help="Only match existing employees, don't create new ones",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Show what would happen without making changes",
    )
    args = parser.parse_args()

    print("=" * 60)
    print("Payroll Data Seeding from Excel")
    if args.match_only:
        print("  Mode: MATCH ONLY (no new employees will be created)")
    if args.dry_run:
        print("  Mode: DRY RUN (no changes will be made)")
    print("=" * 60)

    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    db = SessionLocal()

    # Statistics
    stats = {
        "created": 0,
        "found": 0,
        "skipped": 0,
        "errors": [],
    }

    try:
        org_id = get_org_id(db)
        user_id = get_admin_user_id(db)
        print(f"Organization ID: {org_id}")
        print(f"Admin User ID: {user_id}")

        # Create batch operation record for audit trail
        file_checksum = get_file_checksum(EXCEL_PATH)
        batch = BatchOperation(
            organization_id=org_id,
            operation_type=BatchOperationType.SCRIPT,
            operation_name="seed_payroll_from_excel",
            description=f"Seeding payroll data from {EXCEL_PATH.name}",
            source_file=str(EXCEL_PATH),
            source_checksum=file_checksum,
            started_by_id=user_id,
            metadata_={"match_only": args.match_only, "dry_run": args.dry_run},
        )
        db.add(batch)
        db.flush()
        print(f"Batch Operation ID: {batch.id}")

        if args.dry_run:
            print("\n[DRY RUN - No changes will be committed]\n")

        # Step 1: Clear existing data (skip in dry run)
        if not args.dry_run:
            clear_payroll_data(db, org_id)

        # Step 2: Create components
        components = create_components(db, org_id, user_id)

        # Step 3: Create structures
        perm_struct, contract_struct = create_structures(
            db, org_id, user_id, components
        )

        # Step 4: Seed tax bands
        seed_tax_bands(db, org_id, user_id)

        # Step 5: Parse Excel data
        permanent_staff, contract_staff = parse_excel_data(EXCEL_PATH)

        # Step 6: Get/create employment types
        perm_emp_type = get_or_create_employment_type(db, org_id, "Permanent", user_id)
        contract_emp_type = get_or_create_employment_type(
            db, org_id, "Contract", user_id
        )
        nysc_emp_type = get_or_create_employment_type(db, org_id, "NYSC", user_id)

        # Step 7: Process employees and create assignments
        print("\nProcessing employees...")
        emp_counter = 1

        for data in permanent_staff:
            emp_type = nysc_emp_type if data["is_nysc"] else perm_emp_type
            employee_code = f"EMP{emp_counter:04d}"

            emp, status = find_or_create_employee(
                db,
                org_id,
                user_id,
                data,
                emp_type,
                employee_code,
                batch.id,
                match_only=args.match_only,
            )

            if status == "found":
                print(f"  ✓ Found: {data['name']} (code: {emp.employee_code})")
                stats["found"] += 1
                create_assignment(
                    db,
                    org_id,
                    user_id,
                    emp,
                    perm_struct,
                    data["monthly_gross"],
                    batch.id,
                )
            elif status == "created":
                print(f"  + Created: {data['name']} (code: {employee_code})")
                stats["created"] += 1
                batch.track_created("employee", emp.employee_id)
                batch.track_created("person", emp.person_id)
                create_assignment(
                    db,
                    org_id,
                    user_id,
                    emp,
                    perm_struct,
                    data["monthly_gross"],
                    batch.id,
                )
            else:
                print(f"  ⚠ Skipped: {data['name']} (not found, match-only mode)")
                stats["skipped"] += 1

            emp_counter += 1

        for data in contract_staff:
            emp_type = nysc_emp_type if data["is_nysc"] else contract_emp_type
            employee_code = f"EMP{emp_counter:04d}"

            emp, status = find_or_create_employee(
                db,
                org_id,
                user_id,
                data,
                emp_type,
                employee_code,
                batch.id,
                match_only=args.match_only,
            )

            if status == "found":
                print(f"  ✓ Found: {data['name']} (code: {emp.employee_code})")
                stats["found"] += 1
                create_assignment(
                    db,
                    org_id,
                    user_id,
                    emp,
                    contract_struct,
                    data["monthly_gross"],
                    batch.id,
                )
            elif status == "created":
                print(f"  + Created: {data['name']} (code: {employee_code})")
                stats["created"] += 1
                batch.track_created("employee", emp.employee_id)
                batch.track_created("person", emp.person_id)
                create_assignment(
                    db,
                    org_id,
                    user_id,
                    emp,
                    contract_struct,
                    data["monthly_gross"],
                    batch.id,
                )
            else:
                print(f"  ⚠ Skipped: {data['name']} (not found, match-only mode)")
                stats["skipped"] += 1

            emp_counter += 1

        # Update batch operation with statistics
        batch.mark_completed(
            created=stats["created"],
            updated=stats["found"],  # Found employees got new assignments
            skipped=stats["skipped"],
            failed=len(stats["errors"]),
        )

        if args.dry_run:
            print("\n[DRY RUN - Rolling back all changes]")
            db.rollback()
        else:
            db.commit()

        print("\n" + "=" * 60)
        print("SUMMARY")
        print("=" * 60)
        print(f"  Employees found (matched):  {stats['found']}")
        print(f"  Employees created (new):    {stats['created']}")
        print(f"  Employees skipped:          {stats['skipped']}")
        print(f"  Errors:                     {len(stats['errors'])}")
        print(f"\n  Batch Operation ID: {batch.id}")
        print("=" * 60)

        if not args.dry_run:
            print("\nSUCCESS: Payroll data seeded successfully!")
            print("\nTo rollback this batch, run:")
            print(
                f"  DELETE FROM payroll.salary_structure_assignment WHERE batch_operation_id = '{batch.id}';"
            )
            print(f"  DELETE FROM hr.employee WHERE batch_operation_id = '{batch.id}';")
            print(
                f"  DELETE FROM public.people WHERE batch_operation_id = '{batch.id}';"
            )

    except Exception as e:
        db.rollback()
        print(f"\nERROR: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    main()
