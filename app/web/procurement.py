"""
Procurement Web Routes.

Server-rendered HTML routes for procurement management.
"""

import csv
import json
import math
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Dict, List, Optional, Tuple
from urllib.parse import quote
from uuid import UUID

from fastapi import APIRouter, Depends, File, Form, Query, Request, UploadFile
from pydantic import ValidationError as PydanticValidationError

from fastapi.responses import HTMLResponse, RedirectResponse, Response
from sqlalchemy.orm import Session, selectinload
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError, DataError

from app.services.common import NotFoundError, ValidationError
from app.services.procurement.procurement_plan import ProcurementPlanService
from app.services.procurement.contract import ContractService
from app.services.procurement.rfq import RFQService
from app.services.procurement.requisition import RequisitionService
from app.services.procurement.vendor import VendorPrequalificationService
from app.services.procurement.web.procurement_web import ProcurementWebService
from app.models.procurement.enums import (
    ProcurementMethod,
    ProcurementPlanStatus,
    RequisitionStatus,
    RFQStatus,
    UrgencyLevel,
    ContractStatus,
)
from app.models.procurement.procurement_plan import ProcurementPlan
from app.models.procurement.rfq import RequestForQuotation
from app.models.procurement.purchase_requisition import PurchaseRequisition
from app.models.procurement.procurement_contract import ProcurementContract
from app.schemas.procurement.procurement_plan import (
    PlanItemCreate,
    ProcurementPlanCreate,
)
from app.schemas.procurement.requisition import RequisitionCreate, RequisitionLineCreate
from app.schemas.procurement.rfq import RFQCreate
from app.schemas.procurement.contract import ContractCreate
from app.schemas.procurement.vendor import PrequalificationCreate
from app.web.deps import (
    WebAuthContext,
    base_context,
    get_db,
    require_procurement_access,
    templates,
)

router = APIRouter(prefix="/procurement", tags=["procurement-web"])


IMPORT_REQUIRED_COLUMNS = [
    "plan_number",
    "fiscal_year",
    "title",
    "line_number",
    "description",
    "estimated_value",
    "planned_quarter",
]
IMPORT_OPTIONAL_COLUMNS = [
    "currency_code",
    "budget_line_code",
    "budget_id",
    "procurement_method",
    "category",
]
IMPORT_ALL_COLUMNS = IMPORT_REQUIRED_COLUMNS + IMPORT_OPTIONAL_COLUMNS


REQUISITION_REQUIRED_COLUMNS = [
    "requisition_number",
    "requisition_date",
    "requester_id",
    "line_number",
    "description",
    "quantity",
    "estimated_unit_price",
    "estimated_amount",
]
REQUISITION_OPTIONAL_COLUMNS = [
    "department_id",
    "urgency",
    "justification",
    "currency_code",
    "material_request_id",
    "plan_item_id",
    "item_id",
    "uom",
    "expense_account_id",
    "cost_center_id",
    "project_id",
    "delivery_date",
]
REQUISITION_ALL_COLUMNS = REQUISITION_REQUIRED_COLUMNS + REQUISITION_OPTIONAL_COLUMNS


RFQ_REQUIRED_COLUMNS = [
    "rfq_number",
    "title",
    "rfq_date",
    "closing_date",
]
RFQ_OPTIONAL_COLUMNS = [
    "procurement_method",
    "requisition_id",
    "plan_item_id",
    "evaluation_criteria",
    "terms_and_conditions",
    "estimated_value",
    "currency_code",
]
RFQ_ALL_COLUMNS = RFQ_REQUIRED_COLUMNS + RFQ_OPTIONAL_COLUMNS


CONTRACT_REQUIRED_COLUMNS = [
    "contract_number",
    "title",
    "supplier_id",
    "contract_date",
    "start_date",
    "end_date",
    "contract_value",
]
CONTRACT_OPTIONAL_COLUMNS = [
    "rfq_id",
    "evaluation_id",
    "currency_code",
    "bpp_clearance_number",
    "bpp_clearance_date",
    "payment_terms",
    "terms_and_conditions",
    "performance_bond_required",
    "performance_bond_amount",
    "retention_percentage",
]
CONTRACT_ALL_COLUMNS = CONTRACT_REQUIRED_COLUMNS + CONTRACT_OPTIONAL_COLUMNS


def _xlsx_available() -> bool:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return False
    return True


def _normalize_column(name: str) -> str:
    normalized = name.strip().lower()
    normalized = normalized.replace(" ", "_").replace("-", "_")
    return "".join(ch for ch in normalized if ch.isalnum() or ch == "_")


def _is_empty(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _parse_decimal(value: object) -> Decimal:
    if _is_empty(value):
        raise InvalidOperation("Missing decimal value")
    return Decimal(str(value).strip())


def _parse_int(value: object) -> int:
    if _is_empty(value):
        raise InvalidOperation("Missing integer value")
    if isinstance(value, bool):
        raise InvalidOperation("Invalid integer value")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        raise InvalidOperation("Invalid integer value")
    text = str(value).strip()
    num = Decimal(text)
    if num != num.to_integral_value():
        raise InvalidOperation("Invalid integer value")
    return int(num)


def _parse_date(value: object) -> date:
    if _is_empty(value):
        raise InvalidOperation("Missing date value")
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError as exc:
        raise InvalidOperation("Invalid date format (expected YYYY-MM-DD)") from exc


def _parse_uuid(value: object) -> Optional[UUID]:
    if _is_empty(value):
        return None
    try:
        return UUID(str(value).strip())
    except (ValueError, TypeError) as exc:
        raise InvalidOperation("Invalid UUID value") from exc


def _parse_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if _is_empty(value):
        return False
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y", "on"}:
        return True
    if text in {"false", "0", "no", "n", "off"}:
        return False
    raise InvalidOperation("Invalid boolean value")


def _load_import_rows(content: bytes, fmt: str) -> Tuple[List[Dict[str, object]], List[str]]:
    if fmt == "csv":
        reader = csv.DictReader(StringIO(content.decode("utf-8-sig")))
        if not reader.fieldnames:
            return [], []
        rows = [dict(row) for row in reader]
        return rows, list(reader.fieldnames)
    if fmt == "xlsx":
        try:
            import openpyxl
        except ImportError as exc:
            raise RuntimeError("openpyxl not installed") from exc
        workbook = openpyxl.load_workbook(
            BytesIO(content), data_only=True, read_only=True
        )
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], []
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        rows: List[Dict[str, object]] = []
        for row in rows_iter:
            row_dict: Dict[str, object] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = row[idx] if idx < len(row) else None
                row_dict[header] = value
            rows.append(row_dict)
        return rows, headers
    raise ValueError("Unsupported import format")


# =============================================================================
# Dashboard
# =============================================================================


@router.get("", response_class=HTMLResponse)
def procurement_dashboard(
    request: Request,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Procurement management dashboard."""
    context = base_context(request, auth, "Procurement", "procurement", db=db)
    web_service = ProcurementWebService(db)
    context.update(web_service.dashboard_context(auth.organization_id))
    return templates.TemplateResponse(request, "procurement/dashboard.html", context)


# =============================================================================
# Plans
# =============================================================================


@router.get("/plans", response_class=HTMLResponse)
def plan_list(
    request: Request,
    status: Optional[str] = None,
    fiscal_year: Optional[str] = None,
    offset: int = Query(0, ge=0),
    limit: int = Query(25, ge=1, le=100),
    success: Optional[str] = None,
    error: Optional[str] = None,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """List procurement plans."""
    context = base_context(request, auth, "Procurement Plans", "procurement", db=db)
    web_service = ProcurementWebService(db)
    context.update(
        web_service.plan_list_context(
            auth.organization_id,
            status=status,
            fiscal_year=fiscal_year,
            offset=offset,
            limit=limit,
        )
    )
    context["xlsx_available"] = _xlsx_available()
    context["success"] = success
    context["error"] = error
    return templates.TemplateResponse(request, "procurement/plans/list.html", context)


@router.get("/plans/new", response_class=HTMLResponse)
def plan_new(
    request: Request,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """New procurement plan form."""
    context = base_context(request, auth, "New Plan", "procurement", db=db)
    web_service = ProcurementWebService(db)
    context.update(web_service.plan_form_context(auth.organization_id))
    return templates.TemplateResponse(request, "procurement/plans/form.html", context)


@router.get("/plans/template")
def plan_import_template(
    format: str = Query("csv"),
    auth: WebAuthContext = Depends(require_procurement_access),
):
    """Download a procurement plan import template."""
    fmt = (format or "csv").lower()
    if fmt not in {"csv", "xlsx"}:
        fmt = "csv"
    if fmt == "xlsx" and not _xlsx_available():
        fmt = "csv"

    sample = {
        "plan_number": ["PLAN-2026-001"],
        "fiscal_year": ["2026/2027"],
        "title": ["Annual Procurement Plan"],
        "currency_code": ["NGN"],
        "line_number": [1],
        "description": ["Office supplies procurement"],
        "budget_line_code": ["BL-001"],
        "budget_id": [""],
        "estimated_value": [2500000],
        "procurement_method": ["OPEN_COMPETITIVE"],
        "planned_quarter": [1],
        "category": ["Office Supplies"],
    }

    if fmt == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(IMPORT_ALL_COLUMNS)
        writer.writerow([sample[col][0] for col in IMPORT_ALL_COLUMNS])
        content = output.getvalue()
        return Response(
            content,
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=procurement_plans_template.csv"
            },
        )

    output = BytesIO()
    try:
        import openpyxl
    except ImportError:
        return Response("XLSX support requires openpyxl", status_code=500)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(IMPORT_ALL_COLUMNS)
    sheet.append([sample[col][0] for col in IMPORT_ALL_COLUMNS])
    workbook.save(output)
    return Response(
        output.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": "attachment; filename=procurement_plans_template.xlsx"
        },
    )


@router.get("/plans/export")
def plan_export(
    format: str = Query("csv"),
    status: Optional[str] = None,
    fiscal_year: Optional[str] = None,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Export procurement plans to CSV/XLSX."""
    fmt = (format or "csv").lower()
    if fmt not in {"csv", "xlsx"}:
        fmt = "csv"
    if fmt == "xlsx" and not _xlsx_available():
        fmt = "csv"

    query = (
        select(ProcurementPlan)
        .where(ProcurementPlan.organization_id == auth.organization_id)
        .options(selectinload(ProcurementPlan.items))
        .order_by(ProcurementPlan.created_at.desc())
    )
    if status:
        try:
            query = query.where(ProcurementPlan.status == ProcurementPlanStatus(status))
        except ValueError:
            pass
    if fiscal_year:
        query = query.where(ProcurementPlan.fiscal_year == fiscal_year)

    plans = list(db.scalars(query).all())
    rows: List[List[object]] = []
    for plan in plans:
        if plan.items:
            for item in plan.items:
                rows.append(
                    [
                        plan.plan_number,
                        plan.fiscal_year,
                        plan.title,
                        plan.currency_code,
                        item.line_number,
                        item.description,
                        item.budget_line_code or "",
                        str(item.budget_id) if item.budget_id else "",
                        item.estimated_value,
                        item.procurement_method.value
                        if item.procurement_method
                        else "",
                        item.planned_quarter,
                        item.category or "",
                    ]
                )
        else:
            rows.append(
                [
                    plan.plan_number,
                    plan.fiscal_year,
                    plan.title,
                    plan.currency_code,
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )

    filename = "procurement_plans_export"
    if fmt == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(IMPORT_ALL_COLUMNS)
        writer.writerows(rows)
        return Response(
            output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )

    output = BytesIO()
    try:
        import openpyxl
    except ImportError:
        return Response("XLSX support requires openpyxl", status_code=500)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(IMPORT_ALL_COLUMNS)
    for row in rows:
        sheet.append(row)
    workbook.save(output)
    return Response(
        output.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


@router.get("/plans/{plan_id}", response_class=HTMLResponse)
def plan_detail(
    request: Request,
    plan_id: UUID,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Plan detail view."""
    context = base_context(request, auth, "Plan Details", "procurement", db=db)
    web_service = ProcurementWebService(db)
    try:
        context.update(web_service.plan_detail_context(auth.organization_id, plan_id))
        return templates.TemplateResponse(
            request, "procurement/plans/detail.html", context
        )
    except NotFoundError:
        return RedirectResponse(
            url="/procurement/plans?error=not_found", status_code=303
        )


@router.post("/plans/import")
async def plan_import(
    request: Request,
    file: UploadFile = File(...),
    format: Optional[str] = Form(None),
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Import procurement plans from CSV/XLSX."""
    if not auth.user_id:
        return RedirectResponse(
            url="/procurement/plans?error=Missing+user+context", status_code=303
        )

    if not file or not file.filename:
        return RedirectResponse(
            url="/procurement/plans?error=No+file+provided", status_code=303
        )

    fmt = (format or "").lower().strip()
    if not fmt:
        fmt = "xlsx" if file.filename.lower().endswith(".xlsx") else "csv"
    if fmt not in {"csv", "xlsx"}:
        return RedirectResponse(
            url="/procurement/plans?error=Unsupported+file+format", status_code=303
        )

    content = await file.read()
    if not content:
        return RedirectResponse(
            url="/procurement/plans?error=Empty+file+uploaded", status_code=303
        )

    try:
        rows, headers = _load_import_rows(content, fmt)
    except RuntimeError:
        return RedirectResponse(
            url="/procurement/plans?error=XLSX+support+requires+openpyxl", status_code=303
        )
    except Exception:
        return RedirectResponse(
            url="/procurement/plans?error=Failed+to+read+file", status_code=303
        )

    if not rows:
        return RedirectResponse(
            url="/procurement/plans?error=No+rows+found+in+file", status_code=303
        )

    normalized_headers = [_normalize_column(col) for col in headers]
    missing = set(IMPORT_REQUIRED_COLUMNS) - set(normalized_headers)
    if missing:
        msg = quote(f"Missing required columns: {', '.join(sorted(missing))}")
        return RedirectResponse(url=f"/procurement/plans?error={msg}", status_code=303)

    rows_normalized: List[Dict[str, object]] = []
    header_map = {orig: norm for orig, norm in zip(headers, normalized_headers)}
    for row in rows:
        normalized_row: Dict[str, object] = {}
        for key, value in row.items():
            norm_key = header_map.get(key, _normalize_column(str(key)))
            normalized_row[norm_key] = value
        for col in IMPORT_OPTIONAL_COLUMNS:
            normalized_row.setdefault(col, None)
        rows_normalized.append(normalized_row)

    errors: List[str] = []
    plans: Dict[str, Dict[str, object]] = {}

    for idx, row in enumerate(rows_normalized):
        row_num = idx + 2
        if all(_is_empty(row.get(col)) for col in IMPORT_ALL_COLUMNS):
            continue

        plan_number = str(row.get("plan_number", "")).strip()
        fiscal_year = str(row.get("fiscal_year", "")).strip()
        title = str(row.get("title", "")).strip()

        if not plan_number:
            errors.append(f"Row {row_num}: plan_number is required")
        if not fiscal_year:
            errors.append(f"Row {row_num}: fiscal_year is required")
        if not title:
            errors.append(f"Row {row_num}: title is required")

        currency_code_raw = row.get("currency_code", "")
        currency_code = (
            "NGN" if _is_empty(currency_code_raw) else str(currency_code_raw).strip()
        )
        currency_code = currency_code.upper() or "NGN"

        try:
            line_number = _parse_int(row.get("line_number"))
            if line_number < 1:
                raise InvalidOperation("line_number must be >= 1")
        except InvalidOperation:
            errors.append(f"Row {row_num}: line_number must be a whole number >= 1")
            line_number = 1

        description = str(row.get("description", "")).strip()
        if not description:
            errors.append(f"Row {row_num}: description is required")

        try:
            estimated_value = _parse_decimal(row.get("estimated_value"))
            if estimated_value < 0:
                raise InvalidOperation("estimated_value must be >= 0")
        except InvalidOperation:
            errors.append(f"Row {row_num}: estimated_value must be a number >= 0")
            estimated_value = Decimal("0")

        try:
            planned_quarter = _parse_int(row.get("planned_quarter"))
            if planned_quarter not in {1, 2, 3, 4}:
                raise InvalidOperation("planned_quarter must be 1-4")
        except InvalidOperation:
            errors.append(f"Row {row_num}: planned_quarter must be 1-4")
            planned_quarter = 1

        method_raw = row.get("procurement_method")
        method_value = (
            ProcurementMethod.OPEN_COMPETITIVE
            if _is_empty(method_raw)
            else str(method_raw).strip().upper().replace(" ", "_")
        )
        if isinstance(method_value, str):
            if method_value not in ProcurementMethod.__members__:
                errors.append(
                    f"Row {row_num}: procurement_method must be one of {', '.join(ProcurementMethod.__members__.keys())}"
                )
                method_value = ProcurementMethod.OPEN_COMPETITIVE
            else:
                method_value = ProcurementMethod[method_value]

        budget_line_code = (
            None if _is_empty(row.get("budget_line_code")) else str(row.get("budget_line_code")).strip()
        )
        category = None if _is_empty(row.get("category")) else str(row.get("category")).strip()

        budget_id_value = None
        if not _is_empty(row.get("budget_id")):
            try:
                budget_id_value = UUID(str(row.get("budget_id")).strip())
            except (ValueError, TypeError):
                errors.append(f"Row {row_num}: budget_id must be a valid UUID")

        if plan_number not in plans:
            plans[plan_number] = {
                "plan_number": plan_number,
                "fiscal_year": fiscal_year,
                "title": title,
                "currency_code": currency_code,
                "items": [],
                "line_numbers": set(),
            }
        else:
            plan = plans[plan_number]
            if plan["fiscal_year"] != fiscal_year:
                errors.append(
                    f"Row {row_num}: fiscal_year mismatch for plan_number {plan_number}"
                )
            if plan["title"] != title:
                errors.append(
                    f"Row {row_num}: title mismatch for plan_number {plan_number}"
                )
            if plan["currency_code"] != currency_code:
                errors.append(
                    f"Row {row_num}: currency_code mismatch for plan_number {plan_number}"
                )

        plan = plans.get(plan_number)
        if plan is not None:
            if line_number in plan["line_numbers"]:
                errors.append(
                    f"Row {row_num}: duplicate line_number {line_number} for plan_number {plan_number}"
                )
            plan["line_numbers"].add(line_number)
            plan["items"].append(
                PlanItemCreate(
                    line_number=line_number,
                    description=description,
                    budget_line_code=budget_line_code,
                    budget_id=budget_id_value,
                    estimated_value=estimated_value,
                    procurement_method=method_value,
                    planned_quarter=planned_quarter,
                    category=category,
                )
            )

    if errors:
        msg = quote("; ".join(errors[:8]))
        if len(errors) > 8:
            msg = quote("; ".join(errors[:8]) + f"; and {len(errors) - 8} more")
        return RedirectResponse(url=f"/procurement/plans?error={msg}", status_code=303)

    existing = set(
        db.scalars(
            select(ProcurementPlan.plan_number).where(
                ProcurementPlan.organization_id == auth.organization_id
            )
        ).all()
    )
    duplicates = [num for num in plans.keys() if num in existing]
    if duplicates:
        msg = quote(
            "Plan number(s) already exist: " + ", ".join(sorted(duplicates))
        )
        return RedirectResponse(url=f"/procurement/plans?error={msg}", status_code=303)

    service = ProcurementPlanService(db)
    created_count = 0
    try:
        for plan_data in plans.values():
            data = ProcurementPlanCreate(
                plan_number=plan_data["plan_number"],
                fiscal_year=plan_data["fiscal_year"],
                title=plan_data["title"],
                currency_code=plan_data["currency_code"],
                items=plan_data["items"],
            )
            service.create(auth.organization_id, data, auth.user_id)
            created_count += 1
        db.commit()
    except (ValidationError, ValueError) as exc:
        db.rollback()
        msg = quote(f"Import failed: {str(exc)}")
        return RedirectResponse(url=f"/procurement/plans?error={msg}", status_code=303)

    return RedirectResponse(
        url=f"/procurement/plans?success=Imported+{created_count}+plans",
        status_code=303,
    )


# =============================================================================
# Requisitions
# =============================================================================


@router.get("/requisitions", response_class=HTMLResponse)
def requisition_list(
    request: Request,
    status: Optional[str] = None,
    urgency: Optional[str] = None,
    offset: int = Query(0, ge=0),
    limit: int = Query(25, ge=1, le=100),
    success: Optional[str] = None,
    error: Optional[str] = None,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """List purchase requisitions."""
    context = base_context(request, auth, "Requisitions", "procurement", db=db)
    web_service = ProcurementWebService(db)
    context.update(
        web_service.requisition_list_context(
            auth.organization_id,
            status=status,
            urgency=urgency,
            offset=offset,
            limit=limit,
        )
    )
    context["xlsx_available"] = _xlsx_available()
    context["success"] = success
    context["error"] = error
    return templates.TemplateResponse(
        request, "procurement/requisitions/list.html", context
    )


@router.get("/requisitions/template")
def requisition_import_template(
    format: str = Query("csv"),
    auth: WebAuthContext = Depends(require_procurement_access),
):
    """Download a requisition import template."""
    fmt = (format or "csv").lower()
    if fmt not in {"csv", "xlsx"}:
        fmt = "csv"
    if fmt == "xlsx" and not _xlsx_available():
        fmt = "csv"

    sample = {
        "requisition_number": ["PR-2026-001"],
        "requisition_date": ["2026-02-01"],
        "requester_id": ["00000000-0000-0000-0000-000000000001"],
        "department_id": [""],
        "urgency": ["NORMAL"],
        "justification": ["Replace aging laptops"],
        "currency_code": ["NGN"],
        "material_request_id": [""],
        "plan_item_id": [""],
        "line_number": [1],
        "item_id": [""],
        "description": ["Laptop procurement"],
        "quantity": [5],
        "uom": ["pcs"],
        "estimated_unit_price": [1200],
        "estimated_amount": [6000],
        "expense_account_id": [""],
        "cost_center_id": [""],
        "project_id": [""],
        "delivery_date": ["2026-03-01"],
    }

    if fmt == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(REQUISITION_ALL_COLUMNS)
        writer.writerow([sample[col][0] for col in REQUISITION_ALL_COLUMNS])
        return Response(
            output.getvalue(),
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=requisitions_template.csv"
            },
        )

    output = BytesIO()
    try:
        import openpyxl
    except ImportError:
        return Response("XLSX support requires openpyxl", status_code=500)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(REQUISITION_ALL_COLUMNS)
    sheet.append([sample[col][0] for col in REQUISITION_ALL_COLUMNS])
    workbook.save(output)
    return Response(
        output.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": "attachment; filename=requisitions_template.xlsx"
        },
    )


@router.get("/requisitions/export")
def requisition_export(
    format: str = Query("csv"),
    status: Optional[str] = None,
    urgency: Optional[str] = None,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Export requisitions to CSV/XLSX."""
    fmt = (format or "csv").lower()
    if fmt not in {"csv", "xlsx"}:
        fmt = "csv"
    if fmt == "xlsx" and not _xlsx_available():
        fmt = "csv"

    query = (
        select(PurchaseRequisition)
        .where(PurchaseRequisition.organization_id == auth.organization_id)
        .options(selectinload(PurchaseRequisition.lines))
        .order_by(PurchaseRequisition.created_at.desc())
    )
    if status:
        try:
            query = query.where(PurchaseRequisition.status == RequisitionStatus(status))
        except ValueError:
            pass
    if urgency:
        try:
            query = query.where(PurchaseRequisition.urgency == UrgencyLevel(urgency))
        except ValueError:
            pass

    requisitions = list(db.scalars(query).all())
    rows: List[List[object]] = []
    for req in requisitions:
        header_values = [
            req.requisition_number,
            req.requisition_date.isoformat() if req.requisition_date else "",
            str(req.requester_id),
            str(req.department_id) if req.department_id else "",
            req.urgency.value if req.urgency else "",
            req.justification or "",
            req.currency_code,
            str(req.material_request_id) if req.material_request_id else "",
            str(req.plan_item_id) if req.plan_item_id else "",
        ]
        if req.lines:
            for line in req.lines:
                rows.append(
                    header_values
                    + [
                        line.line_number,
                        str(line.item_id) if line.item_id else "",
                        line.description,
                        line.quantity,
                        line.uom or "",
                        line.estimated_unit_price,
                        line.estimated_amount,
                        str(line.expense_account_id)
                        if line.expense_account_id
                        else "",
                        str(line.cost_center_id) if line.cost_center_id else "",
                        str(line.project_id) if line.project_id else "",
                        line.delivery_date.isoformat()
                        if line.delivery_date
                        else "",
                    ]
                )
        else:
            rows.append(
                header_values
                + ["", "", "", "", "", "", "", "", "", "", ""]
            )

    filename = "requisitions_export"
    if fmt == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(REQUISITION_ALL_COLUMNS)
        writer.writerows(rows)
        return Response(
            output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )

    output = BytesIO()
    try:
        import openpyxl
    except ImportError:
        return Response("XLSX support requires openpyxl", status_code=500)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(REQUISITION_ALL_COLUMNS)
    for row in rows:
        sheet.append(row)
    workbook.save(output)
    return Response(
        output.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


@router.post("/requisitions/import")
async def requisition_import(
    request: Request,
    file: UploadFile = File(...),
    format: Optional[str] = Form(None),
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Import requisitions from CSV/XLSX."""
    if not auth.user_id:
        return RedirectResponse(
            url="/procurement/requisitions?error=Missing+user+context", status_code=303
        )

    if not file or not file.filename:
        return RedirectResponse(
            url="/procurement/requisitions?error=No+file+provided", status_code=303
        )

    fmt = (format or "").lower().strip()
    if not fmt:
        fmt = "xlsx" if file.filename.lower().endswith(".xlsx") else "csv"
    if fmt not in {"csv", "xlsx"}:
        return RedirectResponse(
            url="/procurement/requisitions?error=Unsupported+file+format",
            status_code=303,
        )

    content = await file.read()
    if not content:
        return RedirectResponse(
            url="/procurement/requisitions?error=Empty+file+uploaded",
            status_code=303,
        )

    try:
        rows, headers = _load_import_rows(content, fmt)
    except RuntimeError:
        return RedirectResponse(
            url="/procurement/requisitions?error=XLSX+support+requires+openpyxl",
            status_code=303,
        )
    except Exception:
        return RedirectResponse(
            url="/procurement/requisitions?error=Failed+to+read+file",
            status_code=303,
        )

    if not rows:
        return RedirectResponse(
            url="/procurement/requisitions?error=No+rows+found+in+file",
            status_code=303,
        )

    normalized_headers = [_normalize_column(col) for col in headers]
    missing = set(REQUISITION_REQUIRED_COLUMNS) - set(normalized_headers)
    if missing:
        msg = quote(f"Missing required columns: {', '.join(sorted(missing))}")
        return RedirectResponse(
            url=f"/procurement/requisitions?error={msg}", status_code=303
        )

    rows_normalized: List[Dict[str, object]] = []
    header_map = {orig: norm for orig, norm in zip(headers, normalized_headers)}
    for row in rows:
        normalized_row: Dict[str, object] = {}
        for key, value in row.items():
            norm_key = header_map.get(key, _normalize_column(str(key)))
            normalized_row[norm_key] = value
        for col in REQUISITION_OPTIONAL_COLUMNS:
            normalized_row.setdefault(col, None)
        rows_normalized.append(normalized_row)

    errors: List[str] = []
    requisitions: Dict[str, Dict[str, object]] = {}

    for idx, row in enumerate(rows_normalized):
        row_num = idx + 2
        if all(_is_empty(row.get(col)) for col in REQUISITION_ALL_COLUMNS):
            continue

        requisition_number = str(row.get("requisition_number", "")).strip()
        if not requisition_number:
            errors.append(f"Row {row_num}: requisition_number is required")

        try:
            requisition_date = _parse_date(row.get("requisition_date"))
        except InvalidOperation:
            errors.append(
                f"Row {row_num}: requisition_date must be YYYY-MM-DD"
            )
            requisition_date = date.today()

        try:
            requester_id = _parse_uuid(row.get("requester_id"))
            if requester_id is None:
                raise InvalidOperation("Missing requester_id")
        except InvalidOperation:
            errors.append(f"Row {row_num}: requester_id must be a valid UUID")
            requester_id = None

        department_id = None
        try:
            department_id = _parse_uuid(row.get("department_id"))
        except InvalidOperation:
            errors.append(f"Row {row_num}: department_id must be a valid UUID")

        urgency_raw = row.get("urgency")
        urgency_value = (
            UrgencyLevel.NORMAL
            if _is_empty(urgency_raw)
            else str(urgency_raw).strip().upper().replace(" ", "_")
        )
        if isinstance(urgency_value, str):
            if urgency_value not in UrgencyLevel.__members__:
                errors.append(
                    f"Row {row_num}: urgency must be one of {', '.join(UrgencyLevel.__members__.keys())}"
                )
                urgency_value = UrgencyLevel.NORMAL
            else:
                urgency_value = UrgencyLevel[urgency_value]

        justification = (
            None
            if _is_empty(row.get("justification"))
            else str(row.get("justification")).strip()
        )

        currency_code_raw = row.get("currency_code", "")
        currency_code = (
            "NGN" if _is_empty(currency_code_raw) else str(currency_code_raw).strip()
        )
        currency_code = currency_code.upper() or "NGN"

        material_request_id = None
        try:
            material_request_id = _parse_uuid(row.get("material_request_id"))
        except InvalidOperation:
            errors.append(f"Row {row_num}: material_request_id must be a valid UUID")

        plan_item_id = None
        try:
            plan_item_id = _parse_uuid(row.get("plan_item_id"))
        except InvalidOperation:
            errors.append(f"Row {row_num}: plan_item_id must be a valid UUID")

        try:
            line_number = _parse_int(row.get("line_number"))
            if line_number < 1:
                raise InvalidOperation("line_number must be >= 1")
        except InvalidOperation:
            errors.append(f"Row {row_num}: line_number must be a whole number >= 1")
            line_number = 1

        description = str(row.get("description", "")).strip()
        if not description:
            errors.append(f"Row {row_num}: description is required")

        try:
            quantity = _parse_decimal(row.get("quantity"))
            if quantity <= 0:
                raise InvalidOperation("quantity must be > 0")
        except InvalidOperation:
            errors.append(f"Row {row_num}: quantity must be a number > 0")
            quantity = Decimal("1")

        try:
            estimated_unit_price = _parse_decimal(row.get("estimated_unit_price"))
            if estimated_unit_price < 0:
                raise InvalidOperation("estimated_unit_price must be >= 0")
        except InvalidOperation:
            errors.append(
                f"Row {row_num}: estimated_unit_price must be a number >= 0"
            )
            estimated_unit_price = Decimal("0")

        try:
            estimated_amount = _parse_decimal(row.get("estimated_amount"))
            if estimated_amount < 0:
                raise InvalidOperation("estimated_amount must be >= 0")
        except InvalidOperation:
            errors.append(f"Row {row_num}: estimated_amount must be a number >= 0")
            estimated_amount = Decimal("0")

        item_id = None
        try:
            item_id = _parse_uuid(row.get("item_id"))
        except InvalidOperation:
            errors.append(f"Row {row_num}: item_id must be a valid UUID")

        uom = None if _is_empty(row.get("uom")) else str(row.get("uom")).strip()

        expense_account_id = None
        try:
            expense_account_id = _parse_uuid(row.get("expense_account_id"))
        except InvalidOperation:
            errors.append(f"Row {row_num}: expense_account_id must be a valid UUID")

        cost_center_id = None
        try:
            cost_center_id = _parse_uuid(row.get("cost_center_id"))
        except InvalidOperation:
            errors.append(f"Row {row_num}: cost_center_id must be a valid UUID")

        project_id = None
        try:
            project_id = _parse_uuid(row.get("project_id"))
        except InvalidOperation:
            errors.append(f"Row {row_num}: project_id must be a valid UUID")

        delivery_date = None
        if not _is_empty(row.get("delivery_date")):
            try:
                delivery_date = _parse_date(row.get("delivery_date"))
            except InvalidOperation:
                errors.append(f"Row {row_num}: delivery_date must be YYYY-MM-DD")

        if requisition_number and requisition_number not in requisitions:
            requisitions[requisition_number] = {
                "requisition_number": requisition_number,
                "requisition_date": requisition_date,
                "requester_id": requester_id,
                "department_id": department_id,
                "urgency": urgency_value,
                "justification": justification,
                "currency_code": currency_code,
                "material_request_id": material_request_id,
                "plan_item_id": plan_item_id,
                "lines": [],
                "line_numbers": set(),
            }
        elif requisition_number:
            req = requisitions[requisition_number]
            if req["requisition_date"] != requisition_date:
                errors.append(
                    f"Row {row_num}: requisition_date mismatch for requisition_number {requisition_number}"
                )
            if req["requester_id"] != requester_id:
                errors.append(
                    f"Row {row_num}: requester_id mismatch for requisition_number {requisition_number}"
                )
            if req["department_id"] != department_id:
                errors.append(
                    f"Row {row_num}: department_id mismatch for requisition_number {requisition_number}"
                )
            if req["urgency"] != urgency_value:
                errors.append(
                    f"Row {row_num}: urgency mismatch for requisition_number {requisition_number}"
                )
            if req["currency_code"] != currency_code:
                errors.append(
                    f"Row {row_num}: currency_code mismatch for requisition_number {requisition_number}"
                )
            if req["material_request_id"] != material_request_id:
                errors.append(
                    f"Row {row_num}: material_request_id mismatch for requisition_number {requisition_number}"
                )
            if req["plan_item_id"] != plan_item_id:
                errors.append(
                    f"Row {row_num}: plan_item_id mismatch for requisition_number {requisition_number}"
                )

        req = requisitions.get(requisition_number)
        if req is not None:
            if line_number in req["line_numbers"]:
                errors.append(
                    f"Row {row_num}: duplicate line_number {line_number} for requisition_number {requisition_number}"
                )
            req["line_numbers"].add(line_number)
            req["lines"].append(
                RequisitionLineCreate(
                    line_number=line_number,
                    item_id=item_id,
                    description=description,
                    quantity=quantity,
                    uom=uom,
                    estimated_unit_price=estimated_unit_price,
                    estimated_amount=estimated_amount,
                    expense_account_id=expense_account_id,
                    cost_center_id=cost_center_id,
                    project_id=project_id,
                    delivery_date=delivery_date,
                )
            )

    if errors:
        msg = quote("; ".join(errors[:8]))
        if len(errors) > 8:
            msg = quote("; ".join(errors[:8]) + f"; and {len(errors) - 8} more")
        return RedirectResponse(
            url=f"/procurement/requisitions?error={msg}", status_code=303
        )

    existing = set(
        db.scalars(
            select(PurchaseRequisition.requisition_number).where(
                PurchaseRequisition.organization_id == auth.organization_id
            )
        ).all()
    )
    duplicates = [num for num in requisitions.keys() if num in existing]
    if duplicates:
        msg = quote(
            "Requisition number(s) already exist: " + ", ".join(sorted(duplicates))
        )
        return RedirectResponse(
            url=f"/procurement/requisitions?error={msg}", status_code=303
        )

    service = RequisitionService(db)
    created_count = 0
    try:
        for req_data in requisitions.values():
            if req_data["requester_id"] is None:
                raise ValidationError("Missing requester_id")
            data = RequisitionCreate(
                requisition_number=req_data["requisition_number"],
                requisition_date=req_data["requisition_date"],
                requester_id=req_data["requester_id"],
                department_id=req_data["department_id"],
                urgency=req_data["urgency"],
                justification=req_data["justification"],
                currency_code=req_data["currency_code"],
                material_request_id=req_data["material_request_id"],
                plan_item_id=req_data["plan_item_id"],
                lines=req_data["lines"],
            )
            service.create(auth.organization_id, data, auth.user_id)
            created_count += 1
        db.commit()
    except (ValidationError, ValueError) as exc:
        db.rollback()
        msg = quote(f"Import failed: {str(exc)}")
        return RedirectResponse(
            url=f"/procurement/requisitions?error={msg}", status_code=303
        )

    return RedirectResponse(
        url=f"/procurement/requisitions?success=Imported+{created_count}+requisitions",
        status_code=303,
    )


@router.get("/requisitions/new", response_class=HTMLResponse)
def requisition_new(
    request: Request,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """New requisition form."""
    context = base_context(request, auth, "New Requisition", "procurement", db=db)
    web_service = ProcurementWebService(db)
    context.update(web_service.requisition_form_context(auth.organization_id))
    return templates.TemplateResponse(
        request, "procurement/requisitions/form.html", context
    )


@router.get("/requisitions/{requisition_id}", response_class=HTMLResponse)
def requisition_detail(
    request: Request,
    requisition_id: UUID,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Requisition detail view."""
    context = base_context(request, auth, "Requisition Details", "procurement", db=db)
    web_service = ProcurementWebService(db)
    try:
        context.update(
            web_service.requisition_detail_context(auth.organization_id, requisition_id)
        )
        return templates.TemplateResponse(
            request,
            "procurement/requisitions/detail.html",
            context,
        )
    except NotFoundError:
        return RedirectResponse(
            url="/procurement/requisitions?error=not_found",
            status_code=303,
        )


# =============================================================================
# RFQs
# =============================================================================


@router.get("/rfqs", response_class=HTMLResponse)
def rfq_list(
    request: Request,
    status: Optional[str] = None,
    method: Optional[str] = None,
    offset: int = Query(0, ge=0),
    limit: int = Query(25, ge=1, le=100),
    success: Optional[str] = None,
    error: Optional[str] = None,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """List RFQs."""
    context = base_context(request, auth, "RFQs", "procurement", db=db)
    web_service = ProcurementWebService(db)
    context.update(
        web_service.rfq_list_context(
            auth.organization_id,
            status=status,
            procurement_method=method,
            offset=offset,
            limit=limit,
        )
    )
    context["xlsx_available"] = _xlsx_available()
    context["success"] = success
    context["error"] = error
    return templates.TemplateResponse(request, "procurement/rfqs/list.html", context)


@router.get("/rfqs/template")
def rfq_import_template(
    format: str = Query("csv"),
    auth: WebAuthContext = Depends(require_procurement_access),
):
    """Download an RFQ import template."""
    fmt = (format or "csv").lower()
    if fmt not in {"csv", "xlsx"}:
        fmt = "csv"
    if fmt == "xlsx" and not _xlsx_available():
        fmt = "csv"

    sample = {
        "rfq_number": ["RFQ-2026-001"],
        "title": ["Office IT Equipment"],
        "rfq_date": ["2026-02-01"],
        "closing_date": ["2026-02-20"],
        "procurement_method": ["OPEN_COMPETITIVE"],
        "requisition_id": [""],
        "plan_item_id": [""],
        "evaluation_criteria": ['[{"name":"Price","weight":60},{"name":"Quality","weight":40}]'],
        "terms_and_conditions": ["Net 30 payment terms."],
        "estimated_value": [50000],
        "currency_code": ["NGN"],
    }

    if fmt == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(RFQ_ALL_COLUMNS)
        writer.writerow([sample[col][0] for col in RFQ_ALL_COLUMNS])
        return Response(
            output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=rfqs_template.csv"},
        )

    output = BytesIO()
    try:
        import openpyxl
    except ImportError:
        return Response("XLSX support requires openpyxl", status_code=500)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(RFQ_ALL_COLUMNS)
    sheet.append([sample[col][0] for col in RFQ_ALL_COLUMNS])
    workbook.save(output)
    return Response(
        output.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": "attachment; filename=rfqs_template.xlsx"},
    )


@router.get("/rfqs/export")
def rfq_export(
    format: str = Query("csv"),
    status: Optional[str] = None,
    method: Optional[str] = None,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Export RFQs to CSV/XLSX."""
    fmt = (format or "csv").lower()
    if fmt not in {"csv", "xlsx"}:
        fmt = "csv"
    if fmt == "xlsx" and not _xlsx_available():
        fmt = "csv"

    query = (
        select(RequestForQuotation)
        .where(RequestForQuotation.organization_id == auth.organization_id)
        .order_by(RequestForQuotation.created_at.desc())
    )
    if status:
        try:
            query = query.where(RequestForQuotation.status == RFQStatus(status))
        except ValueError:
            pass
    if method:
        try:
            method_enum = ProcurementMethod(method)
        except ValueError:
            method_enum = None
        if method_enum:
            query = query.where(RequestForQuotation.procurement_method == method_enum)

    rfqs = list(db.scalars(query).all())
    rows: List[List[object]] = []
    for rfq in rfqs:
        criteria = rfq.evaluation_criteria
        criteria_value = ""
        if criteria is not None:
            try:
                criteria_value = json.dumps(criteria)
            except (TypeError, ValueError):
                criteria_value = ""
        rows.append(
            [
                rfq.rfq_number,
                rfq.title,
                rfq.rfq_date.isoformat() if rfq.rfq_date else "",
                rfq.closing_date.isoformat() if rfq.closing_date else "",
                rfq.procurement_method.value if rfq.procurement_method else "",
                str(rfq.requisition_id) if rfq.requisition_id else "",
                str(rfq.plan_item_id) if rfq.plan_item_id else "",
                criteria_value,
                rfq.terms_and_conditions or "",
                rfq.estimated_value if rfq.estimated_value is not None else "",
                rfq.currency_code,
            ]
        )

    filename = "rfqs_export"
    if fmt == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(RFQ_ALL_COLUMNS)
        writer.writerows(rows)
        return Response(
            output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )

    output = BytesIO()
    try:
        import openpyxl
    except ImportError:
        return Response("XLSX support requires openpyxl", status_code=500)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(RFQ_ALL_COLUMNS)
    for row in rows:
        sheet.append(row)
    workbook.save(output)
    return Response(
        output.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


@router.post("/rfqs/import")
async def rfq_import(
    request: Request,
    file: UploadFile = File(...),
    format: Optional[str] = Form(None),
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Import RFQs from CSV/XLSX."""
    if not auth.user_id:
        return RedirectResponse(
            url="/procurement/rfqs?error=Missing+user+context", status_code=303
        )

    if not file or not file.filename:
        return RedirectResponse(
            url="/procurement/rfqs?error=No+file+provided", status_code=303
        )

    fmt = (format or "").lower().strip()
    if not fmt:
        fmt = "xlsx" if file.filename.lower().endswith(".xlsx") else "csv"
    if fmt not in {"csv", "xlsx"}:
        return RedirectResponse(
            url="/procurement/rfqs?error=Unsupported+file+format", status_code=303
        )

    content = await file.read()
    if not content:
        return RedirectResponse(
            url="/procurement/rfqs?error=Empty+file+uploaded", status_code=303
        )

    try:
        rows, headers = _load_import_rows(content, fmt)
    except RuntimeError:
        return RedirectResponse(
            url="/procurement/rfqs?error=XLSX+support+requires+openpyxl",
            status_code=303,
        )
    except Exception:
        return RedirectResponse(
            url="/procurement/rfqs?error=Failed+to+read+file", status_code=303
        )

    if not rows:
        return RedirectResponse(
            url="/procurement/rfqs?error=No+rows+found+in+file", status_code=303
        )

    normalized_headers = [_normalize_column(col) for col in headers]
    missing = set(RFQ_REQUIRED_COLUMNS) - set(normalized_headers)
    if missing:
        msg = quote(f"Missing required columns: {', '.join(sorted(missing))}")
        return RedirectResponse(url=f"/procurement/rfqs?error={msg}", status_code=303)

    rows_normalized: List[Dict[str, object]] = []
    header_map = {orig: norm for orig, norm in zip(headers, normalized_headers)}
    for row in rows:
        normalized_row: Dict[str, object] = {}
        for key, value in row.items():
            norm_key = header_map.get(key, _normalize_column(str(key)))
            normalized_row[norm_key] = value
        for col in RFQ_OPTIONAL_COLUMNS:
            normalized_row.setdefault(col, None)
        rows_normalized.append(normalized_row)

    errors: List[str] = []
    rfqs: Dict[str, Dict[str, object]] = {}

    for idx, row in enumerate(rows_normalized):
        row_num = idx + 2
        if all(_is_empty(row.get(col)) for col in RFQ_ALL_COLUMNS):
            continue

        rfq_number = str(row.get("rfq_number", "")).strip()
        title = str(row.get("title", "")).strip()
        if not rfq_number:
            errors.append(f"Row {row_num}: rfq_number is required")
        if not title:
            errors.append(f"Row {row_num}: title is required")

        try:
            rfq_date = _parse_date(row.get("rfq_date"))
        except InvalidOperation:
            errors.append(f"Row {row_num}: rfq_date must be YYYY-MM-DD")
            rfq_date = date.today()

        try:
            closing_date = _parse_date(row.get("closing_date"))
        except InvalidOperation:
            errors.append(f"Row {row_num}: closing_date must be YYYY-MM-DD")
            closing_date = date.today()

        method_raw = row.get("procurement_method")
        method_value = (
            ProcurementMethod.OPEN_COMPETITIVE
            if _is_empty(method_raw)
            else str(method_raw).strip().upper().replace(" ", "_")
        )
        if isinstance(method_value, str):
            if method_value not in ProcurementMethod.__members__:
                errors.append(
                    f"Row {row_num}: procurement_method must be one of {', '.join(ProcurementMethod.__members__.keys())}"
                )
                method_value = ProcurementMethod.OPEN_COMPETITIVE
            else:
                method_value = ProcurementMethod[method_value]

        requisition_id = None
        try:
            requisition_id = _parse_uuid(row.get("requisition_id"))
        except InvalidOperation:
            errors.append(f"Row {row_num}: requisition_id must be a valid UUID")

        plan_item_id = None
        try:
            plan_item_id = _parse_uuid(row.get("plan_item_id"))
        except InvalidOperation:
            errors.append(f"Row {row_num}: plan_item_id must be a valid UUID")

        criteria_value = None
        if not _is_empty(row.get("evaluation_criteria")):
            try:
                criteria_value = json.loads(str(row.get("evaluation_criteria")).strip())
                if not isinstance(criteria_value, list):
                    raise ValueError("evaluation_criteria must be a JSON list")
            except (json.JSONDecodeError, ValueError):
                errors.append(
                    f"Row {row_num}: evaluation_criteria must be valid JSON list"
                )
                criteria_value = None

        terms_and_conditions = (
            None
            if _is_empty(row.get("terms_and_conditions"))
            else str(row.get("terms_and_conditions")).strip()
        )

        estimated_value = None
        if not _is_empty(row.get("estimated_value")):
            try:
                estimated_value = _parse_decimal(row.get("estimated_value"))
                if estimated_value < 0:
                    raise InvalidOperation("estimated_value must be >= 0")
            except InvalidOperation:
                errors.append(
                    f"Row {row_num}: estimated_value must be a number >= 0"
                )
                estimated_value = None

        currency_code_raw = row.get("currency_code", "")
        currency_code = (
            "NGN" if _is_empty(currency_code_raw) else str(currency_code_raw).strip()
        )
        currency_code = currency_code.upper() or "NGN"

        if rfq_number:
            if rfq_number in rfqs:
                errors.append(f"Row {row_num}: duplicate rfq_number {rfq_number}")
            else:
                rfqs[rfq_number] = {
                    "rfq_number": rfq_number,
                    "title": title,
                    "rfq_date": rfq_date,
                    "closing_date": closing_date,
                    "procurement_method": method_value,
                    "requisition_id": requisition_id,
                    "plan_item_id": plan_item_id,
                    "evaluation_criteria": criteria_value,
                    "terms_and_conditions": terms_and_conditions,
                    "estimated_value": estimated_value,
                    "currency_code": currency_code,
                }

    if errors:
        msg = quote("; ".join(errors[:8]))
        if len(errors) > 8:
            msg = quote("; ".join(errors[:8]) + f"; and {len(errors) - 8} more")
        return RedirectResponse(url=f"/procurement/rfqs?error={msg}", status_code=303)

    existing = set(
        db.scalars(
            select(RequestForQuotation.rfq_number).where(
                RequestForQuotation.organization_id == auth.organization_id
            )
        ).all()
    )
    duplicates = [num for num in rfqs.keys() if num in existing]
    if duplicates:
        msg = quote("RFQ number(s) already exist: " + ", ".join(sorted(duplicates)))
        return RedirectResponse(url=f"/procurement/rfqs?error={msg}", status_code=303)

    service = RFQService(db)
    created_count = 0
    try:
        for rfq_data in rfqs.values():
            data = RFQCreate(
                rfq_number=rfq_data["rfq_number"],
                title=rfq_data["title"],
                rfq_date=rfq_data["rfq_date"],
                closing_date=rfq_data["closing_date"],
                procurement_method=rfq_data["procurement_method"],
                requisition_id=rfq_data["requisition_id"],
                plan_item_id=rfq_data["plan_item_id"],
                evaluation_criteria=rfq_data["evaluation_criteria"],
                terms_and_conditions=rfq_data["terms_and_conditions"],
                estimated_value=rfq_data["estimated_value"],
                currency_code=rfq_data["currency_code"],
            )
            service.create(auth.organization_id, data, auth.user_id)
            created_count += 1
        db.commit()
    except (ValidationError, ValueError) as exc:
        db.rollback()
        msg = quote(f"Import failed: {str(exc)}")
        return RedirectResponse(url=f"/procurement/rfqs?error={msg}", status_code=303)

    return RedirectResponse(
        url=f"/procurement/rfqs?success=Imported+{created_count}+rfqs",
        status_code=303,
    )


@router.get("/rfqs/new", response_class=HTMLResponse)
def rfq_new(
    request: Request,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """New RFQ form."""
    context = base_context(request, auth, "New RFQ", "procurement", db=db)
    web_service = ProcurementWebService(db)
    context.update(web_service.rfq_form_context(auth.organization_id))
    return templates.TemplateResponse(request, "procurement/rfqs/form.html", context)


@router.get("/rfqs/{rfq_id}", response_class=HTMLResponse)
def rfq_detail(
    request: Request,
    rfq_id: UUID,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """RFQ detail view."""
    context = base_context(request, auth, "RFQ Details", "procurement", db=db)
    web_service = ProcurementWebService(db)
    try:
        context.update(web_service.rfq_detail_context(auth.organization_id, rfq_id))
        return templates.TemplateResponse(
            request, "procurement/rfqs/detail.html", context
        )
    except NotFoundError:
        return RedirectResponse(
            url="/procurement/rfqs?error=not_found", status_code=303
        )


@router.get("/evaluations", response_class=HTMLResponse)
def evaluation_list(
    request: Request,
    status: Optional[str] = None,
    offset: int = Query(0, ge=0),
    limit: int = Query(25, ge=1, le=100),
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """List bid evaluations."""
    context = base_context(request, auth, "Bid Evaluations", "proc_evaluations", db=db)
    web_service = ProcurementWebService(db)
    context.update(
        web_service.evaluation_list_context(
            auth.organization_id,
            status=status,
            offset=offset,
            limit=limit,
        )
    )
    return templates.TemplateResponse(
        request, "procurement/evaluations/list.html", context
    )


@router.get("/rfqs/{rfq_id}/evaluate", response_class=HTMLResponse)
def evaluation_matrix(
    request: Request,
    rfq_id: UUID,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Evaluation matrix view."""
    context = base_context(request, auth, "Bid Evaluation", "proc_evaluations", db=db)
    web_service = ProcurementWebService(db)
    try:
        context.update(
            web_service.evaluation_matrix_context(auth.organization_id, rfq_id)
        )
        return templates.TemplateResponse(
            request,
            "procurement/evaluations/matrix.html",
            context,
        )
    except NotFoundError:
        return RedirectResponse(
            url="/procurement/rfqs?error=not_found", status_code=303
        )


# =============================================================================
# Contracts
# =============================================================================


@router.get("/contracts", response_class=HTMLResponse)
def contract_list(
    request: Request,
    status: Optional[str] = None,
    offset: int = Query(0, ge=0),
    limit: int = Query(25, ge=1, le=100),
    success: Optional[str] = None,
    error: Optional[str] = None,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """List contracts."""
    context = base_context(request, auth, "Contracts", "procurement", db=db)
    web_service = ProcurementWebService(db)
    context.update(
        web_service.contract_list_context(
            auth.organization_id,
            status=status,
            offset=offset,
            limit=limit,
        )
    )
    context["xlsx_available"] = _xlsx_available()
    context["success"] = success
    context["error"] = error
    return templates.TemplateResponse(
        request, "procurement/contracts/list.html", context
    )


@router.get("/contracts/template")
def contract_import_template(
    format: str = Query("csv"),
    auth: WebAuthContext = Depends(require_procurement_access),
):
    """Download a contract import template."""
    fmt = (format or "csv").lower()
    if fmt not in {"csv", "xlsx"}:
        fmt = "csv"
    if fmt == "xlsx" and not _xlsx_available():
        fmt = "csv"

    sample = {
        "contract_number": ["CTR-2026-001"],
        "title": ["Office Renovation"],
        "supplier_id": ["00000000-0000-0000-0000-000000000001"],
        "contract_date": ["2026-02-01"],
        "start_date": ["2026-02-15"],
        "end_date": ["2026-06-30"],
        "contract_value": [250000],
        "rfq_id": [""],
        "evaluation_id": [""],
        "currency_code": ["NGN"],
        "bpp_clearance_number": [""],
        "bpp_clearance_date": [""],
        "payment_terms": ["Milestone-based"],
        "terms_and_conditions": ["Standard procurement terms apply."],
        "performance_bond_required": ["false"],
        "performance_bond_amount": [""],
        "retention_percentage": [""],
    }

    if fmt == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(CONTRACT_ALL_COLUMNS)
        writer.writerow([sample[col][0] for col in CONTRACT_ALL_COLUMNS])
        return Response(
            output.getvalue(),
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=contracts_template.csv"
            },
        )

    output = BytesIO()
    try:
        import openpyxl
    except ImportError:
        return Response("XLSX support requires openpyxl", status_code=500)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(CONTRACT_ALL_COLUMNS)
    sheet.append([sample[col][0] for col in CONTRACT_ALL_COLUMNS])
    workbook.save(output)
    return Response(
        output.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": "attachment; filename=contracts_template.xlsx"},
    )


@router.get("/contracts/export")
def contract_export(
    format: str = Query("csv"),
    status: Optional[str] = None,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Export contracts to CSV/XLSX."""
    fmt = (format or "csv").lower()
    if fmt not in {"csv", "xlsx"}:
        fmt = "csv"
    if fmt == "xlsx" and not _xlsx_available():
        fmt = "csv"

    query = (
        select(ProcurementContract)
        .where(ProcurementContract.organization_id == auth.organization_id)
        .order_by(ProcurementContract.created_at.desc())
    )
    if status:
        try:
            query = query.where(ProcurementContract.status == ContractStatus(status))
        except ValueError:
            pass

    contracts = list(db.scalars(query).all())
    rows: List[List[object]] = []
    for contract in contracts:
        rows.append(
            [
                contract.contract_number,
                contract.title,
                str(contract.supplier_id),
                contract.contract_date.isoformat()
                if contract.contract_date
                else "",
                contract.start_date.isoformat() if contract.start_date else "",
                contract.end_date.isoformat() if contract.end_date else "",
                contract.contract_value,
                str(contract.rfq_id) if contract.rfq_id else "",
                str(contract.evaluation_id) if contract.evaluation_id else "",
                contract.currency_code,
                contract.bpp_clearance_number or "",
                contract.bpp_clearance_date.isoformat()
                if contract.bpp_clearance_date
                else "",
                contract.payment_terms or "",
                contract.terms_and_conditions or "",
                "true" if contract.performance_bond_required else "false",
                contract.performance_bond_amount
                if contract.performance_bond_amount is not None
                else "",
                contract.retention_percentage
                if contract.retention_percentage is not None
                else "",
            ]
        )

    filename = "contracts_export"
    if fmt == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(CONTRACT_ALL_COLUMNS)
        writer.writerows(rows)
        return Response(
            output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )

    output = BytesIO()
    try:
        import openpyxl
    except ImportError:
        return Response("XLSX support requires openpyxl", status_code=500)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(CONTRACT_ALL_COLUMNS)
    for row in rows:
        sheet.append(row)
    workbook.save(output)
    return Response(
        output.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


@router.post("/contracts/import")
async def contract_import(
    request: Request,
    file: UploadFile = File(...),
    format: Optional[str] = Form(None),
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Import contracts from CSV/XLSX."""
    if not auth.user_id:
        return RedirectResponse(
            url="/procurement/contracts?error=Missing+user+context", status_code=303
        )

    if not file or not file.filename:
        return RedirectResponse(
            url="/procurement/contracts?error=No+file+provided", status_code=303
        )

    fmt = (format or "").lower().strip()
    if not fmt:
        fmt = "xlsx" if file.filename.lower().endswith(".xlsx") else "csv"
    if fmt not in {"csv", "xlsx"}:
        return RedirectResponse(
            url="/procurement/contracts?error=Unsupported+file+format",
            status_code=303,
        )

    content = await file.read()
    if not content:
        return RedirectResponse(
            url="/procurement/contracts?error=Empty+file+uploaded",
            status_code=303,
        )

    try:
        rows, headers = _load_import_rows(content, fmt)
    except RuntimeError:
        return RedirectResponse(
            url="/procurement/contracts?error=XLSX+support+requires+openpyxl",
            status_code=303,
        )
    except Exception:
        return RedirectResponse(
            url="/procurement/contracts?error=Failed+to+read+file",
            status_code=303,
        )

    if not rows:
        return RedirectResponse(
            url="/procurement/contracts?error=No+rows+found+in+file",
            status_code=303,
        )

    normalized_headers = [_normalize_column(col) for col in headers]
    missing = set(CONTRACT_REQUIRED_COLUMNS) - set(normalized_headers)
    if missing:
        msg = quote(f"Missing required columns: {', '.join(sorted(missing))}")
        return RedirectResponse(
            url=f"/procurement/contracts?error={msg}", status_code=303
        )

    rows_normalized: List[Dict[str, object]] = []
    header_map = {orig: norm for orig, norm in zip(headers, normalized_headers)}
    for row in rows:
        normalized_row: Dict[str, object] = {}
        for key, value in row.items():
            norm_key = header_map.get(key, _normalize_column(str(key)))
            normalized_row[norm_key] = value
        for col in CONTRACT_OPTIONAL_COLUMNS:
            normalized_row.setdefault(col, None)
        rows_normalized.append(normalized_row)

    errors: List[str] = []
    contracts: Dict[str, Dict[str, object]] = {}

    for idx, row in enumerate(rows_normalized):
        row_num = idx + 2
        if all(_is_empty(row.get(col)) for col in CONTRACT_ALL_COLUMNS):
            continue

        contract_number = str(row.get("contract_number", "")).strip()
        title = str(row.get("title", "")).strip()
        if not contract_number:
            errors.append(f"Row {row_num}: contract_number is required")
        if not title:
            errors.append(f"Row {row_num}: title is required")

        try:
            supplier_id = _parse_uuid(row.get("supplier_id"))
            if supplier_id is None:
                raise InvalidOperation("Missing supplier_id")
        except InvalidOperation:
            errors.append(f"Row {row_num}: supplier_id must be a valid UUID")
            supplier_id = None

        try:
            contract_date = _parse_date(row.get("contract_date"))
        except InvalidOperation:
            errors.append(f"Row {row_num}: contract_date must be YYYY-MM-DD")
            contract_date = date.today()

        try:
            start_date = _parse_date(row.get("start_date"))
        except InvalidOperation:
            errors.append(f"Row {row_num}: start_date must be YYYY-MM-DD")
            start_date = date.today()

        try:
            end_date = _parse_date(row.get("end_date"))
        except InvalidOperation:
            errors.append(f"Row {row_num}: end_date must be YYYY-MM-DD")
            end_date = date.today()

        try:
            contract_value = _parse_decimal(row.get("contract_value"))
            if contract_value <= 0:
                raise InvalidOperation("contract_value must be > 0")
        except InvalidOperation:
            errors.append(f"Row {row_num}: contract_value must be > 0")
            contract_value = Decimal("0")

        rfq_id = None
        try:
            rfq_id = _parse_uuid(row.get("rfq_id"))
        except InvalidOperation:
            errors.append(f"Row {row_num}: rfq_id must be a valid UUID")

        evaluation_id = None
        try:
            evaluation_id = _parse_uuid(row.get("evaluation_id"))
        except InvalidOperation:
            errors.append(f"Row {row_num}: evaluation_id must be a valid UUID")

        currency_code_raw = row.get("currency_code", "")
        currency_code = (
            "NGN" if _is_empty(currency_code_raw) else str(currency_code_raw).strip()
        )
        currency_code = currency_code.upper() or "NGN"

        bpp_clearance_number = (
            None
            if _is_empty(row.get("bpp_clearance_number"))
            else str(row.get("bpp_clearance_number")).strip()
        )

        bpp_clearance_date = None
        if not _is_empty(row.get("bpp_clearance_date")):
            try:
                bpp_clearance_date = _parse_date(row.get("bpp_clearance_date"))
            except InvalidOperation:
                errors.append(
                    f"Row {row_num}: bpp_clearance_date must be YYYY-MM-DD"
                )

        payment_terms = (
            None
            if _is_empty(row.get("payment_terms"))
            else str(row.get("payment_terms")).strip()
        )

        terms_and_conditions = (
            None
            if _is_empty(row.get("terms_and_conditions"))
            else str(row.get("terms_and_conditions")).strip()
        )

        try:
            performance_bond_required = _parse_bool(
                row.get("performance_bond_required")
            )
        except InvalidOperation:
            errors.append(
                f"Row {row_num}: performance_bond_required must be true/false"
            )
            performance_bond_required = False

        performance_bond_amount = None
        if not _is_empty(row.get("performance_bond_amount")):
            try:
                performance_bond_amount = _parse_decimal(
                    row.get("performance_bond_amount")
                )
                if performance_bond_amount < 0:
                    raise InvalidOperation("performance_bond_amount must be >= 0")
            except InvalidOperation:
                errors.append(
                    f"Row {row_num}: performance_bond_amount must be >= 0"
                )
                performance_bond_amount = None

        retention_percentage = None
        if not _is_empty(row.get("retention_percentage")):
            try:
                retention_percentage = _parse_decimal(
                    row.get("retention_percentage")
                )
                if retention_percentage < 0:
                    raise InvalidOperation("retention_percentage must be >= 0")
            except InvalidOperation:
                errors.append(f"Row {row_num}: retention_percentage must be >= 0")
                retention_percentage = None

        if contract_number:
            if contract_number in contracts:
                errors.append(
                    f"Row {row_num}: duplicate contract_number {contract_number}"
                )
            else:
                contracts[contract_number] = {
                    "contract_number": contract_number,
                    "title": title,
                    "supplier_id": supplier_id,
                    "rfq_id": rfq_id,
                    "evaluation_id": evaluation_id,
                    "contract_date": contract_date,
                    "start_date": start_date,
                    "end_date": end_date,
                    "contract_value": contract_value,
                    "currency_code": currency_code,
                    "bpp_clearance_number": bpp_clearance_number,
                    "bpp_clearance_date": bpp_clearance_date,
                    "payment_terms": payment_terms,
                    "terms_and_conditions": terms_and_conditions,
                    "performance_bond_required": performance_bond_required,
                    "performance_bond_amount": performance_bond_amount,
                    "retention_percentage": retention_percentage,
                }

    if errors:
        msg = quote("; ".join(errors[:8]))
        if len(errors) > 8:
            msg = quote("; ".join(errors[:8]) + f"; and {len(errors) - 8} more")
        return RedirectResponse(
            url=f"/procurement/contracts?error={msg}", status_code=303
        )

    existing = set(
        db.scalars(
            select(ProcurementContract.contract_number).where(
                ProcurementContract.organization_id == auth.organization_id
            )
        ).all()
    )
    duplicates = [num for num in contracts.keys() if num in existing]
    if duplicates:
        msg = quote("Contract number(s) already exist: " + ", ".join(sorted(duplicates)))
        return RedirectResponse(
            url=f"/procurement/contracts?error={msg}", status_code=303
        )

    service = ContractService(db)
    created_count = 0
    try:
        for contract_data in contracts.values():
            if contract_data["supplier_id"] is None:
                raise ValidationError("Missing supplier_id")
            data = ContractCreate(
                contract_number=contract_data["contract_number"],
                title=contract_data["title"],
                supplier_id=contract_data["supplier_id"],
                rfq_id=contract_data["rfq_id"],
                evaluation_id=contract_data["evaluation_id"],
                contract_date=contract_data["contract_date"],
                start_date=contract_data["start_date"],
                end_date=contract_data["end_date"],
                contract_value=contract_data["contract_value"],
                currency_code=contract_data["currency_code"],
                bpp_clearance_number=contract_data["bpp_clearance_number"],
                bpp_clearance_date=contract_data["bpp_clearance_date"],
                payment_terms=contract_data["payment_terms"],
                terms_and_conditions=contract_data["terms_and_conditions"],
                performance_bond_required=contract_data["performance_bond_required"],
                performance_bond_amount=contract_data["performance_bond_amount"],
                retention_percentage=contract_data["retention_percentage"],
            )
            service.create(auth.organization_id, data, auth.user_id)
            created_count += 1
        db.commit()
    except (ValidationError, ValueError) as exc:
        db.rollback()
        msg = quote(f"Import failed: {str(exc)}")
        return RedirectResponse(
            url=f"/procurement/contracts?error={msg}", status_code=303
        )

    return RedirectResponse(
        url=f"/procurement/contracts?success=Imported+{created_count}+contracts",
        status_code=303,
    )


@router.get("/contracts/new", response_class=HTMLResponse)
def contract_new(
    request: Request,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """New contract form."""
    context = base_context(request, auth, "New Contract", "procurement", db=db)
    web_service = ProcurementWebService(db)
    context.update(web_service.contract_form_context(auth.organization_id))
    return templates.TemplateResponse(
        request, "procurement/contracts/form.html", context
    )


@router.post("/contracts/new")
async def contract_create(
    request: Request,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Handle contract form submission."""
    form = getattr(request.state, "csrf_form", None)
    if form is None:
        form = await request.form()

    try:
        contract_number = str(form.get("contract_number") or "").strip()
        title = str(form.get("title") or "").strip()
        if not contract_number:
            raise ValidationError("Contract number is required")
        if not title:
            raise ValidationError("Title is required")

        supplier_id = _parse_uuid(form.get("supplier_id"))
        if not supplier_id:
            raise ValidationError("Supplier ID is required")

        contract_date = _parse_date(form.get("contract_date"))
        start_date = _parse_date(form.get("start_date"))
        end_date = _parse_date(form.get("end_date"))
        contract_value = _parse_decimal(form.get("contract_value"))
        currency_code = (form.get("currency_code") or "").strip() or "NGN"

        bpp_clearance_date = (
            _parse_date(form.get("bpp_clearance_date"))
            if not _is_empty(form.get("bpp_clearance_date"))
            else None
        )

        data = ContractCreate(
            contract_number=contract_number,
            title=title,
            supplier_id=supplier_id,
            rfq_id=_parse_uuid(form.get("rfq_id")),
            evaluation_id=_parse_uuid(form.get("evaluation_id")),
            contract_date=contract_date,
            start_date=start_date,
            end_date=end_date,
            contract_value=contract_value,
            currency_code=currency_code[:3],
            bpp_clearance_number=(form.get("bpp_clearance_number") or "").strip() or None,
            bpp_clearance_date=bpp_clearance_date,
            payment_terms=(form.get("payment_terms") or "").strip() or None,
            terms_and_conditions=(form.get("terms_and_conditions") or "").strip() or None,
            performance_bond_required=_parse_bool(
                form.get("performance_bond_required")
            ),
            performance_bond_amount=(
                _parse_decimal(form.get("performance_bond_amount"))
                if not _is_empty(form.get("performance_bond_amount"))
                else None
            ),
            retention_percentage=(
                _parse_decimal(form.get("retention_percentage"))
                if not _is_empty(form.get("retention_percentage"))
                else None
            ),
        )

        if not auth.user_id:
            raise ValidationError("Missing user identity")

        service = ContractService(db)
        contract = service.create(auth.organization_id, data, auth.user_id)
        db.commit()
        return RedirectResponse(
            url=f"/procurement/contracts/{contract.contract_id}?created=Contract+created",
            status_code=303,
        )
    except (ValidationError, InvalidOperation, ValueError, PydanticValidationError) as exc:
        db.rollback()
        return RedirectResponse(
            url=f"/procurement/contracts/new?error={quote(str(exc))}",
            status_code=303,
        )
    except IntegrityError as exc:
        db.rollback()
        message = str(getattr(exc, "orig", exc))
        if "uq_proc_contract_org_number" in message:
            msg = "Contract number already exists. Please choose a different number."
        else:
            msg = "Contract could not be saved due to a data conflict."
        return RedirectResponse(
            url=f"/procurement/contracts/new?error={quote(msg)}",
            status_code=303,
        )
    except DataError:
        db.rollback()
        return RedirectResponse(
            url="/procurement/contracts/new?error=Some+fields+have+invalid+values",
            status_code=303,
        )
    except Exception:
        db.rollback()
        return RedirectResponse(
            url="/procurement/contracts/new?error=Unable+to+save+contract",
            status_code=303,
        )


@router.get("/contracts/{contract_id}", response_class=HTMLResponse)
def contract_detail(
    request: Request,
    contract_id: UUID,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Contract detail view."""
    context = base_context(request, auth, "Contract Details", "procurement", db=db)
    web_service = ProcurementWebService(db)
    try:
        context.update(
            web_service.contract_detail_context(auth.organization_id, contract_id)
        )
        return templates.TemplateResponse(
            request,
            "procurement/contracts/detail.html",
            context,
        )
    except NotFoundError:
        return RedirectResponse(
            url="/procurement/contracts?error=not_found",
            status_code=303,
        )


# =============================================================================
# Vendors
# =============================================================================


@router.get("/vendors", response_class=HTMLResponse)
def vendor_list(
    request: Request,
    status: Optional[str] = None,
    q: Optional[str] = None,
    offset: int = Query(0, ge=0),
    limit: int = Query(25, ge=1, le=100),
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Vendor registry."""
    context = base_context(request, auth, "Vendor Registry", "procurement", db=db)
    web_service = ProcurementWebService(db)
    context.update(
        web_service.vendor_list_context(
            auth.organization_id,
            status=status,
            q=q,
            offset=offset,
            limit=limit,
        )
    )
    return templates.TemplateResponse(request, "procurement/vendors/list.html", context)


@router.get("/vendors/prequalify")
def prequalification_shortcut():
    """Redirect legacy prequalify URL to prequalification form."""
    return RedirectResponse(
        url="/procurement/vendors/prequalification/new", status_code=302
    )


@router.get("/vendors/prequalification", response_class=HTMLResponse)
def prequalification_list(
    request: Request,
    status: Optional[str] = None,
    q: Optional[str] = None,
    offset: int = Query(0, ge=0),
    limit: int = Query(25, ge=1, le=100),
    success: Optional[str] = None,
    error: Optional[str] = None,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Vendor prequalification list."""
    context = base_context(
        request, auth, "Vendor Prequalification", "proc_prequalification", db=db
    )
    web_service = ProcurementWebService(db)
    context.update(
        web_service.vendor_list_context(
            auth.organization_id,
            status=status,
            q=q,
            offset=offset,
            limit=limit,
        )
    )
    context["success"] = success
    context["error"] = error
    return templates.TemplateResponse(request, "procurement/vendors/list.html", context)


@router.get("/vendors/prequalification/new", response_class=HTMLResponse)
def prequalification_new(
    request: Request,
    error: Optional[str] = None,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """New vendor prequalification form."""
    context = base_context(
        request, auth, "Prequalify Vendor", "proc_prequalification", db=db
    )
    web_service = ProcurementWebService(db)
    context.update(web_service.prequalification_form_context(auth.organization_id))
    context["error"] = error
    return templates.TemplateResponse(
        request, "procurement/vendors/prequalification_form.html", context
    )


@router.post("/vendors/prequalification")
def prequalification_create(
    request: Request,
    supplier_id: str = Form(...),
    application_date: str = Form(...),
    categories: Optional[List[str]] = Form(None),
    categories_json: Optional[str] = Form(None),
    documents_verified: Optional[str] = Form(None),
    tax_clearance_valid: Optional[str] = Form(None),
    pension_compliance: Optional[str] = Form(None),
    itf_compliance: Optional[str] = Form(None),
    nsitf_compliance: Optional[str] = Form(None),
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Create a new vendor prequalification record."""
    if not auth.user_id:
        return RedirectResponse(
            url="/procurement/vendors/prequalification?error=Missing+user+context",
            status_code=303,
        )

    errors: List[str] = []
    try:
        supplier_uuid = _parse_uuid(supplier_id)
        if supplier_uuid is None:
            raise InvalidOperation("Missing supplier_id")
    except InvalidOperation:
        errors.append("supplier_id must be a valid UUID")
        supplier_uuid = None

    try:
        app_date = _parse_date(application_date)
    except InvalidOperation:
        errors.append("application_date must be YYYY-MM-DD")
        app_date = date.today()

    categories_value = None
    if categories:
        categories_value = [{"name": c} for c in categories if c and c.strip()]
    elif categories_json and categories_json.strip():
        try:
            categories_value = json.loads(categories_json)
            if not isinstance(categories_value, list):
                raise ValueError("categories must be a JSON list")
        except (json.JSONDecodeError, ValueError):
            errors.append("categories must be a valid JSON list")

    try:
        documents_verified_val = _parse_bool(documents_verified)
    except InvalidOperation:
        errors.append("documents_verified must be true/false")
        documents_verified_val = False

    try:
        tax_clearance_valid_val = _parse_bool(tax_clearance_valid)
    except InvalidOperation:
        errors.append("tax_clearance_valid must be true/false")
        tax_clearance_valid_val = False

    try:
        pension_compliance_val = _parse_bool(pension_compliance)
    except InvalidOperation:
        errors.append("pension_compliance must be true/false")
        pension_compliance_val = False

    try:
        itf_compliance_val = _parse_bool(itf_compliance)
    except InvalidOperation:
        errors.append("itf_compliance must be true/false")
        itf_compliance_val = False

    try:
        nsitf_compliance_val = _parse_bool(nsitf_compliance)
    except InvalidOperation:
        errors.append("nsitf_compliance must be true/false")
        nsitf_compliance_val = False

    if errors:
        msg = quote("; ".join(errors))
        return RedirectResponse(
            url=f"/procurement/vendors/prequalification/new?error={msg}",
            status_code=303,
        )

    service = VendorPrequalificationService(db)
    try:
        data = PrequalificationCreate(
            supplier_id=supplier_uuid,  # type: ignore[arg-type]
            application_date=app_date,
            categories=categories_value,
            documents_verified=documents_verified_val,
            tax_clearance_valid=tax_clearance_valid_val,
            pension_compliance=pension_compliance_val,
            itf_compliance=itf_compliance_val,
            nsitf_compliance=nsitf_compliance_val,
        )
        preq = service.create(auth.organization_id, data)
        db.commit()
    except ValidationError as exc:
        db.rollback()
        msg = quote(str(exc))
        return RedirectResponse(
            url=f"/procurement/vendors/prequalification/new?error={msg}",
            status_code=303,
        )

    return RedirectResponse(
        url="/procurement/vendors/prequalification?success=Prequalification+created",
        status_code=303,
    )


@router.get("/vendors/{prequalification_id}", response_class=HTMLResponse)
def prequalification_redirect(
    prequalification_id: UUID,
):
    """Backward-compatible redirect to prequalification detail."""
    return RedirectResponse(
        url=f"/procurement/vendors/{prequalification_id}/prequalification",
        status_code=302,
    )


@router.get(
    "/vendors/{prequalification_id}/prequalification", response_class=HTMLResponse
)
def prequalification_detail(
    request: Request,
    prequalification_id: UUID,
    auth: WebAuthContext = Depends(require_procurement_access),
    db: Session = Depends(get_db),
):
    """Prequalification detail view."""
    context = base_context(
        request, auth, "Prequalification Details", "procurement", db=db
    )
    web_service = ProcurementWebService(db)
    try:
        context.update(
            web_service.prequalification_detail_context(
                auth.organization_id,
                prequalification_id,
            )
        )
        return templates.TemplateResponse(
            request,
            "procurement/vendors/prequalification.html",
            context,
        )
    except NotFoundError:
        return RedirectResponse(
            url="/procurement/vendors?error=not_found",
            status_code=303,
        )
