"""
Base Importer Module.

Provides the foundation for all entity-specific importers with common
functionality for CSV parsing, validation, duplicate detection, and error handling.

Supports multiple CSV formats:
- Zoho Books
- QuickBooks Online/Desktop
- Sage 50/Intacct
- Xero
- Wave Accounting
- FreshBooks
- Generic CSV
"""

import csv
import logging
import re
from abc import ABC, abstractmethod
from collections.abc import Callable, Sequence
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from enum import Enum
from pathlib import Path
from typing import (
    Any,
    Generic,
    TypeVar,
    Union,
)
from uuid import UUID

from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════════════════════
# COLUMN ALIASES FOR MULTI-FORMAT SUPPORT
# ═══════════════════════════════════════════════════════════════════════════════

# Maps standard field names to various source column aliases
COLUMN_ALIASES: dict[str, list[str]] = {
    # Account fields
    "account_name": [
        "Account Name",
        "AccountName",
        "Name",
        "account_name",
        "Account",
        "GL Account",
        "GL Account Name",
        "Ledger Name",
        # QuickBooks
        "Account:Name",
        "FullyQualifiedName",
        # Sage
        "Nominal Name",
        "Nominal Account Name",
        # Xero
        "Account name",
        "*Account Name",
    ],
    "account_code": [
        "Account Code",
        "AccountCode",
        "Code",
        "account_code",
        "GL Code",
        "Account Number",
        "Account No",
        "Acct No",
        # QuickBooks
        "Account:Code",
        "AcctNum",
        # Sage
        "Nominal Code",
        "N/C",
        # Xero
        "Account Code",
        "*Code",
    ],
    "account_type": [
        "Account Type",
        "AccountType",
        "Type",
        "account_type",
        "Category",
        "Account Category",
        # QuickBooks
        "Account:Type",
        "AccountType",
        "Classification",
        # Sage
        "Account Type",
        "Type",
        # Xero
        "Account Type",
        "*Type",
    ],
    # Contact fields
    "display_name": [
        "Display Name",
        "DisplayName",
        "Name",
        "display_name",
        "Contact Name",
        "Full Name",
        "Company/Name",
        # QuickBooks
        "DisplayName",
        "PrintOnCheckName",
        "CompanyName",
        # Sage
        "Account Name",
        "Name",
        # Xero
        "Contact Name",
        "*ContactName",
        "Name",
    ],
    "company_name": [
        "Company Name",
        "CompanyName",
        "Company",
        "company_name",
        "Business Name",
        "Organization",
        "Org Name",
        # QuickBooks
        "CompanyName",
        "Company",
        # Sage
        "Company",
        "Organisation",
        # Xero
        "Company Name",
    ],
    "email": [
        "Email",
        "email",
        "Email Address",
        "E-mail",
        "EmailAddress",
        "Primary Email",
        "Contact Email",
        # QuickBooks
        "PrimaryEmailAddr:Address",
        "Email",
        # Sage
        "E-mail",
        "Email Address",
        # Xero
        "Email",
        "EmailAddress",
    ],
    "phone": [
        "Phone",
        "phone",
        "Phone Number",
        "Telephone",
        "Tel",
        "Mobile",
        "Primary Phone",
        "Work Phone",
        # QuickBooks
        "PrimaryPhone:FreeFormNumber",
        "Phone",
        # Sage
        "Telephone",
        "Tel No",
        # Xero
        "Phone",
        "PhoneNumber",
    ],
    "billing_address": [
        "Billing Address",
        "Address",
        "billing_address",
        "Street Address",
        "Address Line 1",
        # QuickBooks
        "BillAddr:Line1",
        "BillingAddress",
        # Sage
        "Address 1",
        "Invoice Address",
        # Xero
        "AddressLine1",
        "Billing Address",
    ],
    # Item fields
    "item_name": [
        "Item Name",
        "ItemName",
        "Name",
        "item_name",
        "Product Name",
        "Product",
        "Service Name",
        # QuickBooks
        "Name",
        "FullyQualifiedName",
        "Item",
        # Sage
        "Stock Item",
        "Product Name",
        # Xero
        "Item Name",
        "*Name",
        "ItemCode",
    ],
    "item_code": [
        "Item Code",
        "ItemCode",
        "Code",
        "SKU",
        "Product Code",
        "Part Number",
        "Part No",
        "item_code",
        "sku",
        # QuickBooks
        "SKU",
        "Sku",
        # Sage
        "Stock Code",
        "Product Code",
        # Xero
        "Item Code",
        "Code",
    ],
    "unit_price": [
        "Unit Price",
        "Price",
        "Rate",
        "Selling Price",
        "Sales Price",
        "unit_price",
        "price",
        # QuickBooks
        "UnitPrice",
        "Rate",
        # Sage
        "Sales Price",
        "Unit Price",
        # Xero
        "Unit Price",
        "SalesUnitPrice",
    ],
    "cost_price": [
        "Cost Price",
        "Cost",
        "Purchase Price",
        "Buy Price",
        "cost_price",
        "cost",
        # QuickBooks
        "PurchaseCost",
        "Cost",
        # Sage
        "Cost Price",
        "Purchase Price",
        # Xero
        "Cost Price",
        "PurchaseUnitPrice",
    ],
    # Financial fields
    "amount": [
        "Amount",
        "Total",
        "Value",
        "amount",
        "total",
        "Net Amount",
        "Gross Amount",
        "Line Amount",
        # QuickBooks
        "Amount",
        "TotalAmt",
        # Sage
        "Net",
        "Amount",
        # Xero
        "Amount",
        "LineAmount",
    ],
    "tax_amount": [
        "Tax Amount",
        "Tax",
        "VAT",
        "GST",
        "Sales Tax",
        "tax_amount",
        "tax",
        "vat",
        # QuickBooks
        "TaxAmt",
        "Tax",
        # Sage
        "VAT",
        "Tax Amount",
        # Xero
        "TaxAmount",
        "Tax",
    ],
    "currency": [
        "Currency",
        "Currency Code",
        "currency",
        "currency_code",
        "Curr",
        "CurrencyCode",
        # QuickBooks
        "CurrencyRef:value",
        "Currency",
        # Sage
        "Currency",
        "Currency Code",
        # Xero
        "Currency",
        "CurrencyCode",
    ],
    # Date fields
    "date": [
        "Date",
        "date",
        "Transaction Date",
        "Trans Date",
        "Entry Date",
        "Doc Date",
        # QuickBooks
        "TxnDate",
        "Date",
        # Sage
        "Date",
        "Trans Date",
        # Xero
        "Date",
        "DateString",
    ],
    "due_date": [
        "Due Date",
        "DueDate",
        "due_date",
        "Payment Due",
        "Due",
        "Terms Date",
        # QuickBooks
        "DueDate",
        "Due Date",
        # Sage
        "Due Date",
        # Xero
        "DueDate",
        "Due Date",
    ],
    # Invoice fields
    "invoice_number": [
        "Invoice Number",
        "Invoice No",
        "InvoiceNumber",
        "Invoice #",
        "Inv No",
        "Doc No",
        "Reference",
        "invoice_number",
        # QuickBooks
        "DocNumber",
        "RefNumber",
        "Invoice No.",
        # Sage
        "Invoice No",
        "Inv No",
        # Xero
        "InvoiceNumber",
        "Invoice Number",
    ],
    "customer_name": [
        "Customer Name",
        "Customer",
        "CustomerName",
        "customer_name",
        "Client Name",
        "Client",
        "Bill To",
        # QuickBooks
        "CustomerRef:name",
        "Customer:DisplayName",
        # Sage
        "Customer",
        "Account Name",
        # Xero
        "Contact",
        "ContactName",
    ],
    "vendor_name": [
        "Vendor Name",
        "Vendor",
        "VendorName",
        "Supplier Name",
        "Supplier",
        "vendor_name",
        "supplier_name",
        # QuickBooks
        "VendorRef:name",
        "Vendor:DisplayName",
        # Sage
        "Supplier",
        "Account Name",
        # Xero
        "Contact",
        "ContactName",
    ],
    # Fleet fields
    "vehicle_code": [
        "Vehicle Code",
        "Fleet Code",
        "Vehicle ID",
        "Fleet ID",
        "Vehicle Number",
        "Vehicle No",
        "Code",
    ],
    "registration_number": [
        "Registration Number",
        "Reg Number",
        "Registration No",
        "Plate Number",
        "License Plate",
        "License Plate Number",
    ],
    "vehicle_make": [
        "Make",
        "Manufacturer",
        "Brand",
    ],
    "vehicle_model": [
        "Model",
        "Model Name",
    ],
    "vehicle_year": [
        "Year",
        "Model Year",
        "Manufacture Year",
        "Manufacturing Year",
    ],
    "vehicle_type": [
        "Vehicle Type",
        "Body Type",
        "Type",
    ],
    "fuel_type": [
        "Fuel Type",
        "Fuel",
        "Fuel Type Code",
    ],
    "ownership_type": [
        "Ownership Type",
        "Ownership",
        "Ownership Model",
    ],
    "assignment_type": [
        "Assignment Type",
        "Assigned To Type",
    ],
    "current_odometer": [
        "Current Odometer",
        "Odometer",
        "Current Mileage",
        "Mileage",
    ],
    "purchase_date": [
        "Purchase Date",
        "Acquisition Date",
        "Date Purchased",
    ],
    "purchase_price": [
        "Purchase Price",
        "Acquisition Cost",
        "Purchase Cost",
        "Cost",
    ],
    "lease_start_date": [
        "Lease Start Date",
        "Lease Start",
    ],
    "lease_end_date": [
        "Lease End Date",
        "Lease End",
    ],
    "lease_monthly_cost": [
        "Lease Monthly Cost",
        "Lease Cost",
        "Monthly Lease Cost",
    ],
    "seating_capacity": [
        "Seating Capacity",
        "Seats",
    ],
    "fuel_tank_capacity_liters": [
        "Fuel Tank Capacity",
        "Fuel Tank (Liters)",
        "Tank Capacity",
    ],
    "expected_fuel_efficiency": [
        "Expected Fuel Efficiency",
        "Fuel Efficiency",
        "KM/L",
        "Miles per Gallon",
    ],
    "engine_capacity_cc": [
        "Engine Capacity",
        "Engine Capacity (CC)",
        "Engine CC",
    ],
    "vehicle_status": [
        "Vehicle Status",
        "Status",
    ],
    # Project fields
    "project_code": [
        "Project Code",
        "Code",
        "Project ID",
    ],
    "project_name": [
        "Project Name",
        "Name",
        "Project",
    ],
    "project_status": [
        "Project Status",
        "Status",
    ],
    "project_priority": [
        "Project Priority",
        "Priority",
    ],
    "project_type": [
        "Project Type",
        "Type",
    ],
    "project_manager": [
        "Project Manager",
        "Project Manager Email",
        "Manager Email",
    ],
    "business_unit_code": [
        "Business Unit Code",
        "Business Unit",
        "Unit Code",
    ],
    "segment_code": [
        "Segment Code",
        "Segment",
    ],
    "cost_center_code": [
        "Cost Center Code",
        "Cost Center",
    ],
    "start_date": [
        "Start Date",
        "Begin Date",
    ],
    "end_date": [
        "End Date",
        "Finish Date",
    ],
    "budget_amount": [
        "Budget Amount",
        "Budget",
    ],
    "budget_currency_code": [
        "Budget Currency",
        "Budget Currency Code",
        "Currency",
    ],
    "is_capitalizable": [
        "Capitalizable",
        "Is Capitalizable",
    ],
    # HR fields
    "department_code": [
        "Department Code",
        "Department ID",
        "Dept Code",
    ],
    "department_name": [
        "Department Name",
        "Department",
        "Dept Name",
    ],
    "designation_code": [
        "Designation Code",
        "Title Code",
    ],
    "designation_name": [
        "Designation Name",
        "Designation",
        "Job Title",
        "Title",
    ],
    "employment_type_code": [
        "Employment Type Code",
        "Employment Type ID",
    ],
    "employment_type_name": [
        "Employment Type",
        "Employment Type Name",
    ],
    "employee_code": [
        "Employee Code",
        "Employee ID",
        "Staff ID",
        "Staff Code",
    ],
    "first_name": [
        "First Name",
        "First",
        "Given Name",
    ],
    "last_name": [
        "Last Name",
        "Last",
        "Surname",
        "Family Name",
    ],
    "work_email": [
        "Work Email",
        "Email",
        "Company Email",
        "Official Email",
    ],
    "date_of_joining": [
        "Date of Joining",
        "Join Date",
        "Employment Date",
        "Start Date",
    ],
    "employee_status": [
        "Employee Status",
        "Status",
    ],
    # ── Bank statement fields ───────────────────────────────────────────────
    "transaction_date": [
        "Transaction Date",
        "Date",
        "Txn Date",
        "Posting Date",
        "Trans Date",
        "Tran Date",
        "Book Date",
    ],
    "debit": [
        "Debit",
        "Dr",
        "Withdrawal",
        "Withdrawals",
        "Debit Amount",
    ],
    "credit": [
        "Credit",
        "Cr",
        "Deposit",
        "Deposits",
        "Credit Amount",
    ],
    "transaction_type": [
        "Transaction Type",
        "Txn Type",
        "Trans Type",
        "Tran Type",
        "Type",
    ],
    "description": [
        "Description",
        "Narration",
        "Narrative",
        "Particulars",
        "Details",
        "Memo",
        "Remark",
        "Remarks",
    ],
    "reference": [
        "Reference",
        "Ref",
        "Ref No",
        "Reference Number",
    ],
    "payee_payer": [
        "Payee Payer",
        "Payee",
        "Payer",
        "Beneficiary",
        "Party",
        "Counterparty",
    ],
    "check_number": [
        "Check Number",
        "Cheque",
        "Check",
        "Chq No",
        "Cheque Number",
        "Cheque No",
    ],
    "running_balance": [
        "Running Balance",
        "Balance",
        "Closing Balance",
    ],
    "value_date": [
        "Value Date",
        "Val Date",
    ],
    "transaction_id": [
        "Transaction ID",
        "Txn ID",
        "Trans ID",
    ],
    "bank_reference": [
        "Bank Reference",
        "Bank Ref",
    ],
    "bank_category": [
        "Bank Category",
        "Category",
    ],
    "bank_code": [
        "Bank Code",
    ],
}


def build_alias_map(
    field_types: list[str] | None = None,
) -> dict[str, str]:
    """Flatten ``COLUMN_ALIASES`` into ``{normalized_alias: canonical_field}``.

    If *field_types* is provided, only aliases for those field types are
    included.  Otherwise **all** aliases are included.

    The alias keys are lower-cased with spaces replaced by underscores to
    match the normalisation used by the bank-statement importer.
    """
    result: dict[str, str] = {}
    entries = (
        COLUMN_ALIASES
        if field_types is None
        else {k: v for k, v in COLUMN_ALIASES.items() if k in field_types}
    )
    for canonical, aliases in entries.items():
        # Also add the canonical name itself as an identity mapping
        norm_canonical = canonical.strip().lower().replace(" ", "_")
        if norm_canonical not in result:
            result[norm_canonical] = canonical
        for alias in aliases:
            norm = alias.strip().lower().replace(" ", "_")
            if norm not in result:
                result[norm] = canonical
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# VALIDATION RULES & PATTERNS
# ═══════════════════════════════════════════════════════════════════════════════

# Email validation pattern
EMAIL_PATTERN = re.compile(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$")

# Phone number pattern (flexible to handle various international formats)
PHONE_PATTERN = re.compile(
    r"^[\+]?[(]?[0-9]{1,3}[)]?[-\s\.]?[(]?[0-9]{1,4}[)]?[-\s\.]?[0-9]{1,4}[-\s\.]?[0-9]{1,9}$"
)

# Currency code pattern (ISO 4217)
CURRENCY_PATTERN = re.compile(r"^[A-Z]{3}$")

# Valid ISO currency codes (subset of common ones)
VALID_CURRENCY_CODES = {
    "NGN",
    "USD",
    "EUR",
    "GBP",
    "CAD",
    "AUD",
    "JPY",
    "CNY",
    "INR",
    "ZAR",
    "AED",
    "CHF",
    "SEK",
    "NOK",
    "DKK",
    "NZD",
    "SGD",
    "HKD",
    "KRW",
    "MXN",
    "BRL",
    "RUB",
    "TRY",
    "PLN",
    "THB",
    "MYR",
    "IDR",
    "PHP",
    "VND",
    "EGP",
    "KES",
    "GHS",
    "TZS",
    "UGX",
    "XOF",
    "XAF",
}

# Account type mappings for validation
VALID_ACCOUNT_TYPES = {
    # Zoho Books types
    "other_asset",
    "other_current_asset",
    "cash",
    "bank",
    "fixed_asset",
    "other_current_liability",
    "credit_card",
    "long_term_liability",
    "other_liability",
    "equity",
    "income",
    "other_income",
    "expense",
    "cost_of_goods_sold",
    "other_expense",
    "accounts_receivable",
    "accounts_payable",
    "stock",
    "payment_clearing",
    # QuickBooks types
    "accounts receivable",
    "other current asset",
    "fixed asset",
    "other asset",
    "accounts payable",
    "credit card",
    "other current liability",
    "long term liability",
    "cost of goods sold",
    "other income",
    "other expense",
    # Xero types
    "current",
    "fixed",
    "inventory",
    "non-current",
    "prepayment",
    "receivable",
    "liability",
    "current liability",
    "non-current liability",
    "direct costs",
    "overhead",
    "revenue",
    "sales",
    # Sage types
    "current assets",
    "fixed assets",
    "current liabilities",
    "long term liabilities",
    "capital & reserves",
    "purchases",
    "direct expenses",
    "overheads",
}


@dataclass
class ValidationRule:
    """Defines a validation rule for a field."""

    field_name: str
    rule_type: str  # required, pattern, min_length, max_length, min_value, max_value, choices, custom
    value: Any = None
    message: str | None = None

    def validate(self, field_value: Any) -> tuple[bool, str | None]:
        """
        Validate a value against this rule.
        Returns (is_valid, error_message).
        """
        if field_value is None or str(field_value).strip() == "":
            if self.rule_type == "required":
                return False, self.message or f"'{self.field_name}' is required"
            return True, None  # Empty values pass non-required validation

        str_value = str(field_value).strip()

        if self.rule_type == "pattern":
            if not self.value.match(str_value):
                return False, self.message or f"'{self.field_name}' has invalid format"

        elif self.rule_type == "min_length":
            if len(str_value) < self.value:
                return (
                    False,
                    self.message
                    or f"'{self.field_name}' must be at least {self.value} characters",
                )

        elif self.rule_type == "max_length":
            if len(str_value) > self.value:
                return (
                    False,
                    self.message
                    or f"'{self.field_name}' must not exceed {self.value} characters",
                )

        elif self.rule_type == "min_value":
            try:
                num = Decimal(re.sub(r"[^\d.\-]", "", str_value))
                if num < self.value:
                    return (
                        False,
                        self.message
                        or f"'{self.field_name}' must be at least {self.value}",
                    )
            except (ValueError, TypeError, ArithmeticError):
                return False, f"'{self.field_name}' must be a valid number"

        elif self.rule_type == "max_value":
            try:
                num = Decimal(re.sub(r"[^\d.\-]", "", str_value))
                if num > self.value:
                    return (
                        False,
                        self.message
                        or f"'{self.field_name}' must not exceed {self.value}",
                    )
            except (ValueError, TypeError, ArithmeticError):
                return False, f"'{self.field_name}' must be a valid number"

        elif self.rule_type == "choices":
            normalized = str_value.lower().replace(" ", "_").replace("-", "_")
            valid_choices = {
                str(c).lower().replace(" ", "_").replace("-", "_") for c in self.value
            }
            if normalized not in valid_choices:
                return (
                    False,
                    self.message
                    or f"'{self.field_name}' must be one of: {', '.join(sorted(self.value)[:10])}",
                )

        elif self.rule_type == "email":
            if not EMAIL_PATTERN.match(str_value):
                return (
                    False,
                    self.message
                    or f"'{self.field_name}' must be a valid email address",
                )

        elif self.rule_type == "phone":
            # Clean phone number before validation
            cleaned = re.sub(r"[\s\-\.\(\)]", "", str_value)
            if not (
                cleaned.isdigit() or (cleaned.startswith("+") and cleaned[1:].isdigit())
            ):
                return (
                    False,
                    self.message or f"'{self.field_name}' must be a valid phone number",
                )
            if len(cleaned) < 7 or len(cleaned) > 20:
                return (
                    False,
                    self.message or f"'{self.field_name}' must be a valid phone number",
                )

        elif self.rule_type == "currency":
            upper_value = str_value.upper()
            if upper_value not in VALID_CURRENCY_CODES:
                return (
                    False,
                    self.message
                    or f"'{self.field_name}' must be a valid ISO currency code (e.g., USD, NGN)",
                )

        elif self.rule_type == "positive":
            try:
                num = Decimal(re.sub(r"[^\d.\-]", "", str_value))
                if num < 0:
                    return (
                        False,
                        self.message or f"'{self.field_name}' must be positive",
                    )
            except (ValueError, TypeError, ArithmeticError):
                return False, f"'{self.field_name}' must be a valid number"

        elif self.rule_type == "date":
            # Try to parse the date
            date_formats = [
                "%Y-%m-%d",
                "%d/%m/%Y",
                "%m/%d/%Y",
                "%d-%m-%Y",
                "%Y/%m/%d",
                "%d %b %Y",
                "%d %B %Y",
                "%b %d, %Y",
            ]
            parsed = False
            for fmt in date_formats:
                try:
                    datetime.strptime(str_value, fmt)
                    parsed = True
                    break
                except ValueError:
                    continue
            if not parsed:
                return (
                    False,
                    self.message or f"'{self.field_name}' must be a valid date",
                )

        elif self.rule_type == "custom":
            # Custom validation function
            if callable(self.value):
                result = self.value(field_value)
                if (
                    isinstance(result, tuple)
                    and len(result) == 2
                    and isinstance(result[0], bool)
                    and (result[1] is None or isinstance(result[1], str))
                ):
                    return result[0], result[1]
                if isinstance(result, bool):
                    return result, None if result else (self.message or "Invalid value")
                return False, self.message or "Invalid value"

        return True, None


class ImportStatus(str, Enum):
    """Status of an import operation."""

    PENDING = "pending"
    IN_PROGRESS = "in_progress"
    COMPLETED = "completed"
    COMPLETED_WITH_ERRORS = "completed_with_errors"
    FAILED = "failed"


@dataclass
class ImportError:
    """Represents an import error."""

    row_number: int
    field: str | None
    value: str | None
    message: str

    def __str__(self) -> str:
        if self.field:
            return f"Row {self.row_number}, field '{self.field}': {self.message}"
        return f"Row {self.row_number}: {self.message}"


@dataclass
class ImportWarning:
    """Represents an import warning (non-fatal issue)."""

    row_number: int
    field: str | None
    message: str

    def __str__(self) -> str:
        if self.field:
            return f"Row {self.row_number}, field '{self.field}': {self.message}"
        return f"Row {self.row_number}: {self.message}"


@dataclass
class ImportResult:
    """Result of an import operation."""

    entity_type: str
    status: ImportStatus = ImportStatus.PENDING
    total_rows: int = 0
    imported_count: int = 0
    skipped_count: int = 0
    duplicate_count: int = 0
    error_count: int = 0
    errors: list[ImportError] = field(default_factory=list)
    warnings: list[ImportWarning] = field(default_factory=list)
    imported_ids: list[UUID] = field(default_factory=list)
    duration_seconds: float = 0.0

    @property
    def success_rate(self) -> float:
        """Calculate the success rate as a percentage."""
        if self.total_rows == 0:
            return 0.0
        return (self.imported_count / self.total_rows) * 100

    def add_error(
        self,
        row: int,
        message: str,
        field: str | None = None,
        value: str | None = None,
    ):
        """Add an error to the result."""
        self.errors.append(ImportError(row, field, value, message))
        self.error_count += 1

    def add_warning(self, row: int, message: str, field: str | None = None):
        """Add a warning to the result."""
        self.warnings.append(ImportWarning(row, field, message))

    def to_dict(self) -> dict[str, Any]:
        """Convert result to dictionary for API responses."""
        return {
            "entity_type": self.entity_type,
            "status": self.status.value,
            "total_rows": self.total_rows,
            "imported_count": self.imported_count,
            "skipped_count": self.skipped_count,
            "duplicate_count": self.duplicate_count,
            "error_count": self.error_count,
            "success_rate": f"{self.success_rate:.1f}%",
            "duration_seconds": round(self.duration_seconds, 2),
            "errors": [str(e) for e in self.errors[:50]],  # Limit errors shown
            "warnings": [str(w) for w in self.warnings[:50]],
        }


@dataclass
class FieldMapping:
    """Defines mapping from source CSV field to target model field."""

    source_field: str  # CSV column name
    target_field: str  # Model attribute name
    required: bool = False
    transformer: Callable[[Any], Any] | None = None
    default: Any = None

    def transform(self, value: Any) -> Any:
        """Transform the value using the configured transformer."""
        if value is None or value == "":
            return self.default
        if self.transformer:
            return self.transformer(value)
        return value


@dataclass
class ImportConfig:
    """Configuration for an import operation."""

    organization_id: UUID
    user_id: UUID
    skip_duplicates: bool = True
    dry_run: bool = False
    batch_size: int = 100
    stop_on_error: bool = False
    date_format: str = "%Y-%m-%d"
    decimal_separator: str = "."
    thousands_separator: str = ","
    encoding: str = "utf-8"
    # Column mapping overrides (source_column -> target_field)
    column_mapping: dict[str, str] | None = None


@dataclass
class ColumnMapping:
    """Represents a detected or suggested column mapping."""

    source_column: str  # Column name in the CSV
    target_field: str  # Standard field name
    confidence: float  # 0.0 to 1.0
    sample_values: list[str] = field(default_factory=list)


@dataclass
class PreviewResult:
    """Result of a file preview operation with visual data."""

    entity_type: str
    total_rows: int
    detected_columns: list[str]
    required_columns: list[str]
    optional_columns: list[str]
    missing_required: list[str]
    column_mappings: list[ColumnMapping]
    sample_data: list[dict[str, Any]]  # First N rows for preview
    validation_errors: list[str]
    detected_format: str  # "zoho", "quickbooks", "xero", "sage", "generic"
    is_valid: bool

    def to_dict(self) -> dict[str, Any]:
        """Convert to dictionary for API response."""
        return {
            "entity_type": self.entity_type,
            "total_rows": self.total_rows,
            "detected_columns": self.detected_columns,
            "required_columns": self.required_columns,
            "optional_columns": self.optional_columns,
            "missing_required": self.missing_required,
            "column_mappings": [
                {
                    "source": m.source_column,
                    "target": m.target_field,
                    "confidence": round(m.confidence, 2),
                    "samples": m.sample_values[:3],
                }
                for m in self.column_mappings
            ],
            "sample_data": self.sample_data[:10],
            "validation_errors": self.validation_errors[:20],
            "detected_format": self.detected_format,
            "is_valid": self.is_valid,
        }


# Type variable for the model being imported
T = TypeVar("T")


def resolve_column_alias(column_name: str, field_type: str) -> str | None:
    """
    Check if a column name matches any known alias for a field type.
    Returns the standardized field name if matched, None otherwise.
    """
    aliases = COLUMN_ALIASES.get(field_type, [])
    normalized = column_name.strip().lower().replace("_", " ").replace("-", " ")

    for alias in aliases:
        if alias.lower().replace("_", " ").replace("-", " ") == normalized:
            return field_type
    return None


def detect_csv_format(columns: Sequence[str]) -> str:
    """
    Detect the likely source format of the CSV based on column names.
    Returns: "zoho", "quickbooks", "xero", "sage", "wave", "freshbooks", or "generic"
    """
    column_set = {c.lower() for c in columns}

    # Zoho Books indicators
    zoho_indicators = {
        "account name",
        "account type",
        "zoho",
        "display name as",
        "currency code",
    }
    if len(column_set & zoho_indicators) >= 2:
        return "zoho"

    # QuickBooks indicators
    qb_indicators = {
        "fullyqualifiedname",
        "acctnum",
        "txndate",
        "docnumber",
        "customerref",
    }
    if any(c in str(columns) for c in ["CustomerRef:", "VendorRef:", "AccountRef:"]):
        return "quickbooks"
    if len(column_set & qb_indicators) >= 1:
        return "quickbooks"

    # Xero indicators
    xero_indicators = {
        "contactname",
        "*name",
        "invoicenumber",
        "duedate",
        "emailaddress",
    }
    if any(c.startswith("*") for c in columns):
        return "xero"
    if len(column_set & xero_indicators) >= 2:
        return "xero"

    # Sage indicators
    sage_indicators = {"nominal code", "n/c", "t/c", "nominal name", "tax code"}
    if len(column_set & sage_indicators) >= 2:
        return "sage"

    # Wave indicators
    wave_indicators = {"transaction type", "transaction id", "transaction date"}
    if len(column_set & wave_indicators) >= 2:
        return "wave"

    # FreshBooks indicators
    fb_indicators = {"invoice #", "client name", "p.o. number"}
    if len(column_set & fb_indicators) >= 2:
        return "freshbooks"

    return "generic"


def find_account_by_subledger_type(
    db: Session, organization_id: UUID, subledger_type: str
) -> UUID | None:
    """Find account by subledger type.

    Args:
        db: Database session
        organization_id: Organization UUID
        subledger_type: The subledger type to search for (e.g., 'ar', 'ap')

    Returns:
        Account UUID if found, None otherwise
    """
    from sqlalchemy import select

    from app.models.finance.gl.account import Account

    result = db.execute(
        select(Account).where(
            Account.organization_id == organization_id,
            Account.subledger_type == subledger_type,
        )
    ).scalar_one_or_none()
    return result.account_id if result else None


def find_account_by_name_pattern(
    db: Session, organization_id: UUID, pattern: str
) -> UUID | None:
    """Find account by name pattern (case-insensitive).

    Args:
        db: Database session
        organization_id: Organization UUID
        pattern: The pattern to search for in account name

    Returns:
        Account UUID if found, None otherwise
    """
    from sqlalchemy import select

    from app.models.finance.gl.account import Account

    result = db.execute(
        select(Account).where(
            Account.organization_id == organization_id,
            Account.account_name.ilike(f"%{pattern}%"),
        )
    ).first()
    return result[0].account_id if result else None


class BaseImporter(ABC, Generic[T]):
    """
    Base class for all entity importers.

    Provides common functionality for:
    - CSV file parsing
    - Field mapping and transformation
    - Data validation
    - Duplicate detection
    - Error handling and reporting
    - Batch processing
    """

    # Subclasses must define these
    entity_name: str = "Entity"
    model_class: type[T] | None = None

    def __init__(self, db: Session, config: ImportConfig):
        self.db = db
        self.config = config
        self.result = ImportResult(entity_type=self.entity_name)
        self._field_mappings: list[FieldMapping] = []
        self._id_cache: dict[str, UUID] = {}  # Cache for lookups

    @abstractmethod
    def get_field_mappings(self) -> list[FieldMapping]:
        """Return the field mappings for this entity type."""
        pass

    @abstractmethod
    def get_unique_key(self, row: dict[str, Any]) -> str:
        """Return a unique key for duplicate detection."""
        pass

    @abstractmethod
    def check_duplicate(self, row: dict[str, Any]) -> T | None:
        """Check if the entity already exists. Return existing entity or None."""
        pass

    @abstractmethod
    def create_entity(self, row: dict[str, Any]) -> T:
        """Create a new entity from the row data."""
        pass

    def validate_row(self, row: dict[str, Any], row_num: int) -> bool:
        """
        Validate a row of data. Return True if valid.
        Override in subclasses for entity-specific validation.
        """
        mappings = self.get_field_mappings()
        is_valid = True

        for mapping in mappings:
            if mapping.required:
                value = row.get(mapping.source_field)
                if value is None or str(value).strip() == "":
                    self.result.add_error(
                        row_num,
                        f"Required field '{mapping.source_field}' is missing or empty",
                        mapping.source_field,
                    )
                    is_valid = False

        return is_valid

    def transform_row(self, row: dict[str, Any], row_num: int) -> dict[str, Any]:
        """Transform CSV row data to model field values."""
        transformed = {}
        mappings = self.get_field_mappings()

        for mapping in mappings:
            source_value = row.get(mapping.source_field)
            try:
                transformed[mapping.target_field] = mapping.transform(source_value)
            except Exception as e:
                if mapping.required:
                    raise ValueError(
                        f"Failed to transform required field '{mapping.source_field}': {e}"
                    ) from e
                self.result.add_warning(
                    row_num,
                    f"Failed to transform value '{source_value}': {str(e)}",
                    mapping.source_field,
                )
                transformed[mapping.target_field] = mapping.default

        return transformed

    def import_file(self, file_path: Union[str, Path]) -> ImportResult:
        """
        Import data from a CSV file.

        Args:
            file_path: Path to the CSV file

        Returns:
            ImportResult with statistics and any errors
        """
        import time

        start_time = time.time()

        file_path = Path(file_path)
        if not file_path.exists():
            self.result.status = ImportStatus.FAILED
            self.result.add_error(0, f"File not found: {file_path}", None)
            return self.result

        self.result.status = ImportStatus.IN_PROGRESS

        try:
            with open(file_path, encoding=self.config.encoding) as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                self.result.total_rows = len(rows)

                self._import_rows(rows)

        except UnicodeDecodeError as e:
            self.result.status = ImportStatus.FAILED
            self.result.add_error(0, f"File encoding error: {str(e)}", None)
        except csv.Error as e:
            self.result.status = ImportStatus.FAILED
            self.result.add_error(0, f"CSV parsing error: {str(e)}", None)
        except Exception as e:
            self.result.status = ImportStatus.FAILED
            self.result.add_error(0, f"Unexpected error: {str(e)}", None)

        self.result.duration_seconds = time.time() - start_time

        if self.result.status == ImportStatus.IN_PROGRESS:
            if self.result.error_count > 0:
                self.result.status = ImportStatus.COMPLETED_WITH_ERRORS
            else:
                self.result.status = ImportStatus.COMPLETED

        return self.result

    def import_rows(self, rows: list[dict[str, Any]]) -> ImportResult:
        """
        Import data from a list of dictionaries.

        Args:
            rows: List of row dictionaries

        Returns:
            ImportResult with statistics and any errors
        """
        import time

        start_time = time.time()

        self.result.status = ImportStatus.IN_PROGRESS
        self.result.total_rows = len(rows)

        try:
            self._import_rows(rows)
        except Exception as e:
            self.result.status = ImportStatus.FAILED
            self.result.add_error(0, f"Unexpected error: {str(e)}", None)

        self.result.duration_seconds = time.time() - start_time

        if self.result.status == ImportStatus.IN_PROGRESS:
            if self.result.error_count > 0:
                self.result.status = ImportStatus.COMPLETED_WITH_ERRORS
            else:
                self.result.status = ImportStatus.COMPLETED

        return self.result

    # ── XLSX support ────────────────────────────────────────────────────

    @staticmethod
    def parse_xlsx_file(file_path: Union[str, Path]) -> list[dict[str, str]]:
        """Parse an XLSX/XLSM file into a list of dicts (all values as strings).

        Uses openpyxl in read-only/data-only mode for performance.  Dates
        are converted via ``.isoformat()``, numbers via ``str()``, and
        everything else via ``str()``.
        """
        from openpyxl import load_workbook

        wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
        try:
            sheet = wb.active
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                raw_headers = next(rows_iter)
            except StopIteration:
                return []

            headers = [
                str(h).strip() if h is not None else f"Column_{i}"
                for i, h in enumerate(raw_headers)
            ]

            result: list[dict[str, str]] = []
            for values in rows_iter:
                if not values or not any(v is not None for v in values):
                    continue
                row: dict[str, str] = {}
                for i, h in enumerate(headers):
                    v = values[i] if i < len(values) else None
                    if v is None:
                        row[h] = ""
                    elif isinstance(v, datetime):
                        row[h] = v.date().isoformat()
                    elif isinstance(v, date):
                        row[h] = v.isoformat()
                    else:
                        row[h] = str(v)
                result.append(row)
            return result
        finally:
            wb.close()

    def import_xlsx_file(self, file_path: Union[str, Path]) -> ImportResult:
        """Import data from an XLSX/XLSM file."""
        import time

        start_time = time.time()
        file_path = Path(file_path)

        if not file_path.exists():
            self.result.status = ImportStatus.FAILED
            self.result.add_error(0, f"File not found: {file_path}", None)
            return self.result

        self.result.status = ImportStatus.IN_PROGRESS

        try:
            rows = self.parse_xlsx_file(file_path)
            self.result.total_rows = len(rows)
            self._import_rows(rows)
        except Exception as e:
            self.result.status = ImportStatus.FAILED
            self.result.add_error(0, f"XLSX parsing error: {str(e)}", None)

        self.result.duration_seconds = time.time() - start_time

        if self.result.status == ImportStatus.IN_PROGRESS:
            if self.result.error_count > 0:
                self.result.status = ImportStatus.COMPLETED_WITH_ERRORS
            else:
                self.result.status = ImportStatus.COMPLETED

        return self.result

    def preview_xlsx_file(
        self, file_path: Union[str, Path], max_rows: int = 10
    ) -> PreviewResult:
        """Preview an XLSX/XLSM file — same output as ``preview_file``."""
        file_path = Path(file_path)

        if not file_path.exists():
            return PreviewResult(
                entity_type=self.entity_name,
                total_rows=0,
                detected_columns=[],
                required_columns=self.get_required_fields(),
                optional_columns=self.get_optional_fields(),
                missing_required=self.get_required_fields(),
                column_mappings=[],
                sample_data=[],
                validation_errors=[f"File not found: {file_path}"],
                detected_format="unknown",
                is_valid=False,
            )

        try:
            rows = self.parse_xlsx_file(file_path)
        except Exception as e:
            return PreviewResult(
                entity_type=self.entity_name,
                total_rows=0,
                detected_columns=[],
                required_columns=self.get_required_fields(),
                optional_columns=self.get_optional_fields(),
                missing_required=self.get_required_fields(),
                column_mappings=[],
                sample_data=[],
                validation_errors=[f"Failed to read XLSX: {str(e)}"],
                detected_format="unknown",
                is_valid=False,
            )

        if not rows:
            return PreviewResult(
                entity_type=self.entity_name,
                total_rows=0,
                detected_columns=[],
                required_columns=self.get_required_fields(),
                optional_columns=self.get_optional_fields(),
                missing_required=self.get_required_fields(),
                column_mappings=[],
                sample_data=[],
                validation_errors=["File is empty or has no data rows"],
                detected_format="unknown",
                is_valid=False,
            )

        columns = list(rows[0].keys())
        return self._build_preview_result(columns, rows, max_rows, "xlsx")

    # ── Dispatcher methods ──────────────────────────────────────────────

    def preview_any_file(
        self, file_path: Union[str, Path], max_rows: int = 10
    ) -> PreviewResult:
        """Preview a CSV or XLSX file based on extension."""
        ext = Path(file_path).suffix.lower()
        if ext in (".xlsx", ".xlsm"):
            return self.preview_xlsx_file(file_path, max_rows)
        return self.preview_file(file_path, max_rows)

    def import_any_file(self, file_path: Union[str, Path]) -> ImportResult:
        """Import a CSV or XLSX file based on extension."""
        ext = Path(file_path).suffix.lower()
        if ext in (".xlsx", ".xlsm"):
            return self.import_xlsx_file(file_path)
        return self.import_file(file_path)

    def _remap_row(self, row: dict[str, Any]) -> dict[str, Any]:
        """Apply user-provided column mapping overrides to a row.

        When ``self.config.column_mapping`` is set, remaps row keys from CSV
        header names to canonical target field names.  Unmapped columns are
        passed through unchanged so that subclass validators can still access
        extra data.
        """
        mapping = self.config.column_mapping
        if not mapping:
            return row
        remapped: dict[str, Any] = {}
        for key, value in row.items():
            target = mapping.get(key, key)
            remapped[target] = value
        return remapped

    def _import_rows(self, rows: list[dict[str, Any]]) -> None:
        """Internal method to process rows."""
        batch = []

        for idx, row in enumerate(rows, start=1):
            if self.config.stop_on_error and self.result.error_count > 0:
                break

            try:
                # Apply column mapping overrides
                row = self._remap_row(row)

                # Validate
                if not self.validate_with_rules(row, idx):
                    self.result.skipped_count += 1
                    continue

                # Check duplicate
                if self.config.skip_duplicates:
                    existing = self.check_duplicate(row)
                    if existing:
                        self.result.duplicate_count += 1
                        self.result.skipped_count += 1
                        self.result.add_warning(
                            idx,
                            f"Duplicate entry skipped (key: {self.get_unique_key(row)})",
                        )
                        continue

                # Transform
                transformed = self.transform_row(row, idx)

                # Create entity (dry run check)
                if not self.config.dry_run:
                    entity = self.create_entity(transformed)
                    batch.append(entity)

                    # Batch commit
                    if len(batch) >= self.config.batch_size:
                        self._commit_batch(batch)
                        batch = []

                self.result.imported_count += 1

            except Exception as e:
                self.result.add_error(idx, str(e), None)
                self.result.skipped_count += 1

        # Commit remaining batch
        if batch and not self.config.dry_run:
            self._commit_batch(batch)

    def _commit_batch(self, batch: list[T]) -> None:
        """Commit a batch of entities to the database."""
        try:
            for entity in batch:
                self.db.add(entity)
                if hasattr(entity, "account_id"):
                    self.result.imported_ids.append(entity.account_id)
                elif hasattr(entity, "customer_id"):
                    self.result.imported_ids.append(entity.customer_id)
                elif hasattr(entity, "supplier_id"):
                    self.result.imported_ids.append(entity.supplier_id)
                elif hasattr(entity, "item_id"):
                    self.result.imported_ids.append(entity.item_id)
            self.db.flush()
        except Exception as e:
            self.db.rollback()
            raise e

    # === Utility Methods for Field Transformation ===

    @staticmethod
    def parse_date(value: Any, format: str = "%Y-%m-%d") -> date | None:
        """Parse a date string to a date object."""
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value

        # Try multiple formats
        formats = [format, "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]
        for fmt in formats:
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except ValueError:
                continue

        raise ValueError(f"Cannot parse date: {value}")

    @staticmethod
    def parse_decimal(
        value: Any, thousands_sep: str = ",", decimal_sep: str = "."
    ) -> Decimal | None:
        """Parse a string to Decimal, handling various formats."""
        if value is None or value == "":
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))

        # Clean the string
        s = str(value).strip()
        s = s.replace(thousands_sep, "")
        if decimal_sep != ".":
            s = s.replace(decimal_sep, ".")

        # Remove currency symbols
        s = re.sub(r"[^\d.\-]", "", s)

        try:
            return Decimal(s)
        except InvalidOperation:
            raise ValueError(f"Cannot parse decimal: {value}")

    @staticmethod
    def parse_boolean(value: Any) -> bool | None:
        """Parse a value to boolean."""
        if value is None or value == "":
            return None
        if isinstance(value, bool):
            return value

        s = str(value).lower().strip()
        if s in ("true", "yes", "1", "y", "t"):
            return True
        if s in ("false", "no", "0", "n", "f"):
            return False

        raise ValueError(f"Cannot parse boolean: {value}")

    @staticmethod
    def clean_string(value: Any, max_length: int | None = None) -> str | None:
        """Clean and optionally truncate a string value."""
        if value is None:
            return None
        s = str(value).strip()
        if s == "":
            return None
        if max_length and len(s) > max_length:
            s = s[:max_length]
        return s

    @staticmethod
    def parse_enum(
        value: Any, enum_class: type[Enum], default: Enum | None = None
    ) -> Enum | None:
        """Parse a string to an enum value."""
        if value is None or value == "":
            return default

        s = str(value).upper().strip().replace(" ", "_")

        try:
            return enum_class(s)
        except ValueError:
            # Try by name
            try:
                return enum_class[s]
            except KeyError:
                if default:
                    return default
                raise ValueError(
                    f"Invalid enum value: {value} for {enum_class.__name__}"
                )

    # === Preview and Validation Methods ===

    def get_validation_rules(self) -> list[ValidationRule]:
        """
        Return validation rules for this entity type.
        Override in subclasses to add entity-specific validation.
        """
        return []

    def get_required_fields(self) -> list[str]:
        """Return list of required field names from mappings."""
        return [m.source_field for m in self.get_field_mappings() if m.required]

    def get_optional_fields(self) -> list[str]:
        """Return list of optional field names from mappings."""
        return [m.source_field for m in self.get_field_mappings() if not m.required]

    def resolve_column(self, columns: Sequence[str], field_type: str) -> str | None:
        """
        Find a matching column for a field type using column aliases.
        Returns the actual column name from the CSV if found.
        """
        aliases = COLUMN_ALIASES.get(field_type, [])

        for col in columns:
            normalized_col = col.strip().lower().replace("_", " ").replace("-", " ")
            for alias in aliases:
                if alias.lower().replace("_", " ").replace("-", " ") == normalized_col:
                    return col

        # Try direct field type match
        for col in columns:
            if col.strip().lower().replace("_", " ") == field_type.replace("_", " "):
                return col

        return None

    def auto_map_columns(self, columns: Sequence[str]) -> dict[str, ColumnMapping]:
        """
        Automatically map CSV columns to expected fields.
        Returns dict of target_field -> ColumnMapping.
        """
        mappings = {}
        used_columns: set[str] = set()

        # First pass: try exact matches and aliases
        for mapping in self.get_field_mappings():
            target = mapping.target_field
            source = mapping.source_field

            # Check if source column exists directly
            if source in columns:
                mappings[target] = ColumnMapping(
                    source_column=source,
                    target_field=target,
                    confidence=1.0,
                )
                used_columns.add(source)
                continue

            # Try to find via aliases
            for field_type, aliases in COLUMN_ALIASES.items():
                if source.lower() in [a.lower() for a in aliases]:
                    found_col = self.resolve_column(columns, field_type)
                    if found_col and found_col not in used_columns:
                        mappings[target] = ColumnMapping(
                            source_column=found_col,
                            target_field=target,
                            confidence=0.9,
                        )
                        used_columns.add(found_col)
                        break

        # Second pass: fuzzy matching for unmapped required fields
        for mapping in self.get_field_mappings():
            if mapping.target_field in mappings:
                continue

            best_match = None
            best_score = 0.0

            for col in columns:
                if col in used_columns:
                    continue

                # Calculate similarity score
                col_norm = col.lower().replace("_", " ").replace("-", " ")
                source_norm = (
                    mapping.source_field.lower().replace("_", " ").replace("-", " ")
                )

                # Check for substring match
                if source_norm in col_norm or col_norm in source_norm:
                    score = 0.7
                else:
                    # Word overlap
                    col_words = set(col_norm.split())
                    source_words = set(source_norm.split())
                    overlap = len(col_words & source_words)
                    if overlap > 0:
                        score = 0.5 * (overlap / max(len(col_words), len(source_words)))
                    else:
                        continue

                if score > best_score:
                    best_score = score
                    best_match = col

            if best_match and best_score >= 0.5:
                mappings[mapping.target_field] = ColumnMapping(
                    source_column=best_match,
                    target_field=mapping.target_field,
                    confidence=best_score,
                )
                used_columns.add(best_match)

        return mappings

    def preview_file(
        self, file_path: Union[str, Path], max_rows: int = 10
    ) -> PreviewResult:
        """Preview a CSV file with column mapping suggestions and validation."""
        file_path = Path(file_path)

        if not file_path.exists():
            return PreviewResult(
                entity_type=self.entity_name,
                total_rows=0,
                detected_columns=[],
                required_columns=self.get_required_fields(),
                optional_columns=self.get_optional_fields(),
                missing_required=self.get_required_fields(),
                column_mappings=[],
                sample_data=[],
                validation_errors=[f"File not found: {file_path}"],
                detected_format="unknown",
                is_valid=False,
            )

        try:
            with open(file_path, encoding=self.config.encoding) as f:
                reader = csv.DictReader(f)
                columns = list(reader.fieldnames or [])
                rows: list[dict[str, Any]] = []
                for i, row in enumerate(reader):
                    rows.append(row)
                    if i >= 99:  # Read up to 100 rows for validation
                        break

        except Exception as e:
            return PreviewResult(
                entity_type=self.entity_name,
                total_rows=0,
                detected_columns=[],
                required_columns=self.get_required_fields(),
                optional_columns=self.get_optional_fields(),
                missing_required=self.get_required_fields(),
                column_mappings=[],
                sample_data=[],
                validation_errors=[f"Failed to read file: {str(e)}"],
                detected_format="unknown",
                is_valid=False,
            )

        # For CSV we can get an accurate total row count cheaply
        try:
            with open(file_path, encoding=self.config.encoding) as f:
                total_override = sum(1 for _ in f) - 1  # Subtract header
        except (OSError, UnicodeDecodeError):
            total_override = None

        return self._build_preview_result(
            columns, rows, max_rows, "csv", total_override=total_override
        )

    def _build_preview_result(
        self,
        columns: list[str],
        rows: list[dict[str, Any]],
        max_rows: int,
        file_format: str,
        *,
        total_override: int | None = None,
    ) -> PreviewResult:
        """Shared preview logic used by both CSV and XLSX preview methods."""
        errors: list[str] = []

        detected_format = detect_csv_format(columns)

        # Auto-map columns
        column_mappings_dict = self.auto_map_columns(columns)

        # Add sample values to mappings
        for _target_field, mapping in column_mappings_dict.items():
            samples = []
            for row in rows[:5]:
                val = row.get(mapping.source_column, "")
                if val and str(val).strip():
                    samples.append(str(val).strip()[:50])
            mapping.sample_values = samples

        # Find missing required fields
        required_fields = self.get_required_fields()
        missing_required = []

        for req_field in required_fields:
            is_mapped = False
            for mapping in column_mappings_dict.values():
                if mapping.source_column == req_field:
                    is_mapped = True
                    break
            if not is_mapped:
                for column_name in columns:
                    if column_name.lower() == req_field.lower():
                        is_mapped = True
                        break
            if not is_mapped:
                missing_required.append(req_field)

        # Validate sample rows
        validation_rules = self.get_validation_rules()
        for idx, row in enumerate(rows[:20], start=1):
            for required_field in required_fields:
                col: str | None = None
                for m in column_mappings_dict.values():
                    if (
                        m.target_field == required_field
                        or m.source_column == required_field
                    ):
                        col = m.source_column
                        break
                if col:
                    value = str(row.get(col, "") or "").strip()
                    if not value:
                        errors.append(
                            f"Row {idx}: Required field '{required_field}' is empty"
                        )

            for rule in validation_rules:
                col = None
                for m in column_mappings_dict.values():
                    if m.target_field == rule.field_name:
                        col = m.source_column
                        break
                if col:
                    value = row.get(col, "")
                    is_valid, error_msg = rule.validate(value)
                    if not is_valid:
                        errors.append(f"Row {idx}: {error_msg or 'Invalid value'}")

        # Prepare sample data for preview display
        sample_data = []
        for row in rows[:max_rows]:
            sample_row = {}
            for col in columns[:15]:
                val = row.get(col, "")
                sample_row[col] = str(val)[:100] if val else ""
            sample_data.append(sample_row)

        total_rows = total_override if total_override is not None else len(rows)
        is_valid = len(missing_required) == 0 and len(errors) == 0

        return PreviewResult(
            entity_type=self.entity_name,
            total_rows=total_rows,
            detected_columns=list(columns),
            required_columns=required_fields,
            optional_columns=self.get_optional_fields(),
            missing_required=missing_required,
            column_mappings=list(column_mappings_dict.values()),
            sample_data=sample_data,
            validation_errors=errors[:50],
            detected_format=detected_format,
            is_valid=is_valid,
        )

    def validate_with_rules(self, row: dict[str, Any], row_num: int) -> bool:
        """
        Validate a row using both field mappings and validation rules.
        Enhanced version of validate_row with full rule support.
        """
        is_valid = self.validate_row(row, row_num)

        # Apply additional validation rules
        for rule in self.get_validation_rules():
            value = row.get(rule.field_name)
            valid, error_msg = rule.validate(value)
            if not valid:
                self.result.add_error(
                    row_num, error_msg or "Invalid value", rule.field_name, str(value)
                )
                is_valid = False

        return is_valid
