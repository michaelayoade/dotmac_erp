"""
Banking web view service.

Provides view-focused data for banking web routes.
"""

from __future__ import annotations

import builtins
import csv
import json
import logging
import re
from datetime import date
from datetime import datetime as _datetime
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Any
from uuid import UUID

from fastapi import HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from pydantic import ValidationError
from sqlalchemy import case, func, or_, select
from sqlalchemy.orm import Session
from starlette.datastructures import UploadFile
from starlette.responses import Response

from app.models.finance.banking.bank_account import BankAccount, BankAccountStatus
from app.models.finance.banking.bank_reconciliation import (
    BankReconciliation,
    BankReconciliationLine,
    ReconciliationStatus,
)
from app.models.finance.banking.bank_statement import (
    BankStatement,
    BankStatementLine,
    BankStatementStatus,
)
from app.models.finance.gl.account import Account
from app.models.finance.gl.journal_entry import JournalEntry, JournalStatus
from app.models.finance.gl.journal_entry_line import JournalEntryLine
from app.schemas.finance.banking import BankStatementImport
from app.services.common import coerce_uuid
from app.services.common_filters import build_active_filters
from app.services.finance.banking import (
    bank_statement_service,
)
from app.services.finance.banking.payment_metadata import (
    PaymentMetadata,
    resolve_payment_metadata,
    resolve_payment_metadata_batch,
)
from app.services.finance.common.sorting import apply_sort
from app.services.finance.platform.currency_context import get_currency_context
from app.services.finance.platform.org_context import org_context_service
from app.services.formatters import format_currency as _base_format_currency
from app.services.formatters import format_date as _format_date
from app.services.formatters import parse_date as _parse_date
from app.services.formatters import parse_decimal as _parse_decimal
from app.services.imports.formats import (
    SPREADSHEET_EXTENSIONS,
    spreadsheet_formats_label,
)
from app.templates import templates
from app.web.deps import WebAuthContext, base_context

logger = logging.getLogger(__name__)

# Human-friendly labels for source_document_type values
_SOURCE_TYPE_LABELS: dict[str, str] = {
    "CUSTOMER_PAYMENT": "Receipt",
    "SUPPLIER_PAYMENT": "Payment",
    "AR_INVOICE": "Invoice",
    "CUSTOMER_INVOICE": "Invoice",
    "INVOICE": "Invoice",
    "SUPPLIER_INVOICE": "Bill",
    "AP_INVOICE": "Bill",
    "EXPENSE": "Expense",
    "EXPENSE_CLAIM": "Expense",
    "JOURNAL_ENTRY": "Journal",
}


def _build_match_detail(
    db: Session,
    entry: JournalEntry,
    source_url: str,
    *,
    metadata: PaymentMetadata | None = None,
) -> dict[str, str]:
    """Build a match detail dict for a journal entry.

    If *metadata* is not provided, resolves it from the entry's source document.
    Falls back to the journal description when no payment metadata is available.
    """
    if metadata is None:
        try:
            metadata = resolve_payment_metadata(
                db,
                getattr(entry, "source_document_type", None),
                getattr(entry, "source_document_id", None),
            )
        except Exception:
            logger.debug(
                "Could not resolve payment metadata for entry %s",
                getattr(entry, "entry_id", None),
            )

    src_type = getattr(entry, "source_document_type", None) or ""
    type_label = _SOURCE_TYPE_LABELS.get(src_type, "GL Entry")

    if metadata:
        return {
            "label": metadata.counterparty_name or type_label,
            "sub": metadata.payment_number or "",
            "type": type_label,
            "url": source_url,
        }

    # Fallback: use journal description
    desc = getattr(entry, "description", "") or ""
    return {
        "label": desc[:60] if desc else type_label,
        "sub": "",
        "type": type_label,
        "url": source_url,
    }


def _format_currency(
    amount: Decimal | None,
    currency: str | None = None,
) -> str:
    """Format currency with em-dash for None values."""
    return str(_base_format_currency(amount, currency, none_value="\u2014"))


def _parse_account_status(value: str | None) -> BankAccountStatus | None:
    """Parse bank account status enum value.

    Logs warning on parse failure for debugging.
    """
    if not value:
        return None
    try:
        return BankAccountStatus(value)
    except ValueError:
        logger.warning("Invalid bank account status value: %r", value)
        return None


def _parse_statement_status(value: str | None) -> BankStatementStatus | None:
    """Parse bank statement status enum value.

    Logs warning on parse failure for debugging.
    """
    if not value:
        return None
    status_map = {
        "in_progress": BankStatementStatus.processing,
        "processing": BankStatementStatus.processing,
    }
    if value in status_map:
        return status_map[value]
    try:
        return BankStatementStatus(value)
    except ValueError:
        logger.warning("Invalid bank statement status value: %r", value)
        return None


def _statement_status_label(status: BankStatementStatus) -> str:
    if status == BankStatementStatus.processing:
        return "in_progress"
    if status == BankStatementStatus.closed:
        return "reconciled"
    return str(status.value)


def _parse_reconciliation_status(
    value: str | None,
) -> ReconciliationStatus | None:
    """Parse reconciliation status enum value.

    Logs warning on parse failure for debugging.
    """
    if not value:
        return None
    try:
        return ReconciliationStatus(value)
    except ValueError:
        logger.warning("Invalid reconciliation status value: %r", value)
        return None


def _account_view(account: BankAccount) -> dict:
    currency = account.currency_code
    return {
        "bank_account_id": account.bank_account_id,
        "bank_name": account.bank_name,
        "bank_code": account.bank_code,
        "branch_code": account.branch_code,
        "branch_name": account.branch_name,
        "account_name": account.account_name,
        "account_number": account.account_number,
        "account_type": account.account_type.value if account.account_type else "",
        "iban": account.iban,
        "currency_code": currency,
        "gl_account_id": account.gl_account_id,
        "status": account.status.value if account.status else "",
        "last_statement_balance": _format_currency(
            account.last_statement_balance, currency
        ),
        "last_statement_date": _format_date(account.last_statement_date),
        "last_reconciled_date": _format_date(account.last_reconciled_date),
        "last_reconciled_balance": _format_currency(
            account.last_reconciled_balance, currency
        ),
        "contact_name": account.contact_name,
        "contact_phone": account.contact_phone,
        "contact_email": account.contact_email,
        "notes": account.notes,
        "allow_overdraft": account.allow_overdraft,
        "overdraft_limit": _format_currency(account.overdraft_limit, currency)
        if account.overdraft_limit
        else None,
    }


def _statement_view(statement: BankStatement) -> dict:
    account = statement.bank_account
    currency = statement.currency_code
    return {
        "statement_id": statement.statement_id,
        "statement_number": statement.statement_number,
        "statement_date": _format_date(statement.statement_date),
        "period_start": _format_date(statement.period_start),
        "period_end": _format_date(statement.period_end),
        "period_start_iso": statement.period_start.isoformat()
        if statement.period_start
        else "",
        "period_end_iso": statement.period_end.isoformat()
        if statement.period_end
        else "",
        "opening_balance": _format_currency(statement.opening_balance, currency),
        "closing_balance": _format_currency(statement.closing_balance, currency),
        "opening_balance_raw": statement.opening_balance,
        "closing_balance_raw": statement.closing_balance,
        "matched_lines": statement.matched_lines,
        "unmatched_lines": statement.unmatched_lines,
        "total_lines": statement.total_lines,
        "total_credits": _format_currency(statement.total_credits, currency),
        "total_debits": _format_currency(statement.total_debits, currency),
        "currency_code": currency,
        "bank_account_id": statement.bank_account_id,
        "bank_name": account.bank_name if account else "",
        "account_number": account.account_number if account else "",
        "account_type": account.account_type if account else "",
        "status": _statement_status_label(statement.status),
    }


def _statement_line_view(line: BankStatementLine, currency: str = "") -> dict:
    return {
        "line_id": line.line_id,
        "line_number": line.line_number,
        "transaction_date": _format_date(line.transaction_date),
        "transaction_type": line.transaction_type.value
        if line.transaction_type
        else "",
        "amount": _format_currency(line.amount, currency),
        "raw_amount": float(line.amount) if line.amount is not None else 0.0,
        "description": line.description,
        "reference": line.reference,
        "payee_payer": line.payee_payer,
        "bank_reference": line.bank_reference,
        "running_balance": _format_currency(line.running_balance, currency),
        "is_matched": line.is_matched,
        "matched_journal_line_id": str(line.matched_journal_line_id)
        if line.matched_journal_line_id
        else None,
        # Categorization fields
        "categorization_status": line.categorization_status.value
        if line.categorization_status
        else None,
        "suggested_account_id": str(line.suggested_account_id)
        if line.suggested_account_id
        else None,
        "suggested_rule_id": str(line.suggested_rule_id)
        if line.suggested_rule_id
        else None,
        "suggested_confidence": line.suggested_confidence,
        "suggested_match_reason": line.suggested_match_reason,
    }


def _reconciliation_view(reconciliation: BankReconciliation) -> dict:
    account = reconciliation.bank_account
    return {
        "reconciliation_id": reconciliation.reconciliation_id,
        "bank_account_id": reconciliation.bank_account_id,
        "bank_name": account.bank_name if account else "",
        "account_number": account.account_number if account else "",
        "reconciliation_date": _format_date(reconciliation.reconciliation_date),
        "period_start": _format_date(reconciliation.period_start),
        "period_end": _format_date(reconciliation.period_end),
        "statement_opening_balance": reconciliation.statement_opening_balance,
        "statement_closing_balance": reconciliation.statement_closing_balance,
        "gl_opening_balance": reconciliation.gl_opening_balance,
        "gl_closing_balance": reconciliation.gl_closing_balance,
        "total_matched": reconciliation.total_matched,
        "total_adjustments": reconciliation.total_adjustments,
        "reconciliation_difference": reconciliation.reconciliation_difference,
        "status": reconciliation.status.value if reconciliation.status else "",
        "currency_code": reconciliation.currency_code,
    }


def _reconciliation_line_view(line: BankReconciliationLine) -> dict:
    return {
        "line_id": line.line_id,
        "transaction_date": _format_date(line.transaction_date),
        "description": line.description,
        "reference": line.reference,
        "statement_amount": line.statement_amount,
        "gl_amount": line.gl_amount,
        "match_type": line.match_type.value if line.match_type else "",
        "adjustment_type": line.adjustment_type,
        "is_adjustment": line.is_adjustment,
        "is_outstanding": line.is_outstanding,
        "outstanding_type": line.outstanding_type,
    }


def _gl_line_view(
    line: JournalEntryLine,
    entry: JournalEntry,
    metadata: PaymentMetadata | None = None,
) -> dict:
    view: dict = {
        "line_id": line.line_id,
        "entry_date": _format_date(entry.entry_date),
        "description": line.description or entry.description,
        "reference": entry.reference,
        "debit_amount": line.debit_amount,
        "credit_amount": line.credit_amount,
        "signed_amount": float(
            (line.debit_amount or Decimal("0")) - (line.credit_amount or Decimal("0"))
        ),
        # Payment metadata (None if not from a payment)
        "source_type": None,
        "source_module": getattr(entry, "source_module", None),
        "payment_number": None,
        "counterparty_name": None,
        "counterparty_type": None,
        "invoice_numbers": [],
    }
    if metadata:
        view["source_type"] = metadata.source_type
        view["payment_number"] = metadata.payment_number
        view["counterparty_name"] = metadata.counterparty_name
        view["counterparty_type"] = metadata.counterparty_type
        view["invoice_numbers"] = metadata.invoice_numbers
    return view


def _line_amount(line: BankReconciliationLine) -> Decimal:
    amount = line.statement_amount
    if amount is None:
        amount = line.gl_amount
    if amount is None:
        return Decimal("0")
    return Decimal(str(amount))


def _gl_line_as_transaction(
    line: JournalEntryLine,
    entry: JournalEntry,
    bank_account: BankAccount,
    currency: str,
    metadata: PaymentMetadata | None = None,
) -> dict[str, Any]:
    """Transform a GL journal line into a transaction dict for the statements template.

    For a bank *asset* account:
    - Debit = money flowing IN  → displayed as CR (credit to the bank)
    - Credit = money flowing OUT → displayed as DR (debit from the bank)
    """
    debit = line.debit_amount or Decimal("0")
    credit = line.credit_amount or Decimal("0")

    if debit > 0:
        txn_type = "credit"  # money in
        amount = debit
    else:
        txn_type = "debit"  # money out
        amount = credit

    src_type = getattr(entry, "source_document_type", None) or ""
    source_label = _SOURCE_TYPE_LABELS.get(src_type, "Journal")

    counterparty = ""
    if metadata and metadata.counterparty_name:
        counterparty = metadata.counterparty_name

    return {
        "transaction_date": _format_date(entry.entry_date),
        "description": line.description or entry.description or "",
        "reference": entry.journal_number or "",
        "bank_name": bank_account.bank_name or "",
        "account_number": bank_account.account_number or "",
        "bank_account_id": str(bank_account.bank_account_id),
        "transaction_type": txn_type,
        "amount": _format_currency(amount, currency),
        "raw_amount": float(amount),
        "payee_payer": counterparty,
        "source_label": source_label,
        "journal_entry_id": str(entry.journal_entry_id),
        "is_matched": None,
    }


def _build_active_filters(
    *,
    account_id: str | None,
    accounts: list[dict],
    status: str | None,
    start_date: str | None,
    end_date: str | None,
    status_labels: dict[str, str] | None = None,
) -> list[dict[str, str]]:
    """Build a list of active filter dicts for the compact_filters macro.

    Each dict has ``name``, ``value``, and ``display_value`` keys.
    """
    filters: list[dict[str, str]] = []
    if account_id:
        # Resolve account display name from the accounts list
        display = account_id
        for acc in accounts:
            acc_id = str(acc.get("bank_account_id", ""))
            if acc_id == account_id:
                display = (
                    f"{acc.get('bank_name', '')} - {acc.get('account_number', '')}"
                )
                break
        filters.append(
            {"name": "account_id", "value": account_id, "display_value": display}
        )
    if status:
        label = status
        if status_labels and status in status_labels:
            label = status_labels[status]
        else:
            label = status.replace("_", " ").title()
        filters.append({"name": "status", "value": status, "display_value": label})
    if start_date:
        filters.append(
            {
                "name": "start_date",
                "value": start_date,
                "display_value": f"From {start_date}",
            }
        )
    if end_date:
        filters.append(
            {"name": "end_date", "value": end_date, "display_value": f"To {end_date}"}
        )
    return filters


class BankingWebService:
    """View service for banking web routes."""

    @staticmethod
    def list_accounts_context(
        db: Session,
        organization_id: str,
        search: str | None,
        status: str | None,
        page: int,
        limit: int = 50,
        sort: str | None = None,
        sort_dir: str | None = None,
    ) -> dict:
        org_id = coerce_uuid(organization_id)
        offset = (page - 1) * limit

        status_value = _parse_account_status(status)

        conditions: list[Any] = [BankAccount.organization_id == org_id]
        if status_value:
            conditions.append(BankAccount.status == status_value)
        if search:
            search_pattern = f"%{search}%"
            conditions.append(
                or_(
                    BankAccount.bank_name.ilike(search_pattern),
                    BankAccount.account_name.ilike(search_pattern),
                    BankAccount.account_number.ilike(search_pattern),
                    BankAccount.branch_name.ilike(search_pattern),
                )
            )

        total_count = (
            db.scalar(
                select(func.count(BankAccount.bank_account_id)).where(*conditions)
            )
            or 0
        )
        account_sort_map: dict[str, Any] = {
            "bank_name": BankAccount.bank_name,
            "account_name": BankAccount.account_name,
            "account_number": BankAccount.account_number,
            "status": BankAccount.status,
        }
        list_stmt = apply_sort(
            select(BankAccount).where(*conditions),
            sort,
            sort_dir,
            account_sort_map,
            default=[BankAccount.bank_name.asc(), BankAccount.account_name.asc()],
        )
        accounts = db.scalars(list_stmt.limit(limit).offset(offset)).all()

        active_count = (
            db.scalar(
                select(func.count(BankAccount.bank_account_id)).where(
                    *conditions, BankAccount.status == BankAccountStatus.active
                )
            )
            or 0
        )
        total_balance = db.scalar(
            select(
                func.coalesce(func.sum(BankAccount.last_statement_balance), 0)
            ).where(*conditions)
        ) or Decimal("0")
        pending_recon = (
            db.scalar(
                select(func.count(BankReconciliation.reconciliation_id)).where(
                    BankReconciliation.organization_id == org_id,
                    BankReconciliation.status.in_(
                        [
                            ReconciliationStatus.draft,
                            ReconciliationStatus.pending_review,
                        ]
                    ),
                )
            )
            or 0
        )

        total_pages = max(1, (total_count + limit - 1) // limit)

        return {
            "accounts": [_account_view(account) for account in accounts],
            "search": search,
            "status": status,
            "sort": sort,
            "sort_dir": sort_dir,
            "page": page,
            "limit": limit,
            "offset": offset,
            "total_count": total_count,
            "total_pages": total_pages,
            "active_count": active_count,
            "total_balance": _format_currency(total_balance),
            "pending_recon": pending_recon,
            "statuses": [s.value for s in BankAccountStatus],
        }

    @staticmethod
    def account_form_context(
        db: Session,
        organization_id: str,
        account_id: str | None = None,
    ) -> dict:
        org_id = coerce_uuid(organization_id)
        account = None
        if account_id:
            account = db.get(BankAccount, coerce_uuid(account_id))

        gl_accounts = db.scalars(
            select(Account)
            .where(
                Account.organization_id == org_id,
                Account.is_active.is_(True),
            )
            .order_by(Account.account_code)
        ).all()

        context = {
            "account": _account_view(account) if account else None,
            "gl_accounts": gl_accounts,
        }
        context.update(get_currency_context(db, organization_id))
        return context

    @staticmethod
    def account_detail_context(
        db: Session,
        organization_id: str,
        account_id: str,
    ) -> dict:
        org_id = coerce_uuid(organization_id)
        account = db.get(BankAccount, coerce_uuid(account_id))
        if not account or account.organization_id != org_id:
            account = None
        transactions: list[dict] = []
        if account:
            rows = db.execute(
                select(BankStatementLine, BankStatement)
                .join(
                    BankStatement,
                    BankStatementLine.statement_id == BankStatement.statement_id,
                )
                .where(
                    BankStatement.organization_id == org_id,
                    BankStatement.bank_account_id == account.bank_account_id,
                )
                .order_by(
                    BankStatementLine.transaction_date.desc(),
                    BankStatementLine.line_number.desc(),
                )
                .limit(50)
            ).all()
            for line, statement in rows:
                view = _statement_line_view(line)
                view.update(
                    {
                        "statement_id": statement.statement_id,
                        "statement_number": statement.statement_number,
                        "statement_date": _format_date(statement.statement_date),
                    }
                )
                transactions.append(view)

        return {
            "account": _account_view(account) if account else None,
            "transactions": transactions,
        }

    @staticmethod
    def transaction_detail_context(
        db: Session,
        organization_id: str,
        line_id: str,
    ) -> dict:
        """Build context for the transaction line detail page."""
        from app.services.finance.banking.bank_reconciliation import (
            _build_source_url,
        )

        org_id = coerce_uuid(organization_id)
        line = db.get(BankStatementLine, coerce_uuid(line_id))
        if not line:
            return {"line": None}

        # Verify org ownership via parent statement
        statement = line.statement
        if not statement or statement.organization_id != org_id:
            return {"line": None}

        account = statement.bank_account
        currency = statement.currency_code

        # Base line view
        line_view = _statement_line_view(line, currency)
        # Add extra fields not in the list helper
        line_view["value_date"] = _format_date(line.value_date)
        line_view["check_number"] = line.check_number
        line_view["bank_category"] = line.bank_category
        line_view["bank_code"] = line.bank_code
        line_view["notes"] = line.notes
        line_view["transaction_id"] = line.transaction_id
        line_view["matched_at"] = (
            line.matched_at.strftime("%d %b %Y %H:%M") if line.matched_at else None
        )

        # Parent statement/account info
        statement_view = {
            "statement_id": str(statement.statement_id),
            "statement_number": statement.statement_number,
            "statement_date": _format_date(statement.statement_date),
            "status": _statement_status_label(statement.status),
        }
        account_view = (
            {
                "bank_account_id": str(account.bank_account_id),
                "account_name": account.account_name,
                "bank_name": account.bank_name,
                "account_number": account.account_number,
            }
            if account
            else None
        )

        # Matched GL entries (via multi-match junction table)
        gl_matches: list[dict] = []
        for match in line.matched_gl_lines:
            jl = db.get(JournalEntryLine, match.journal_line_id)
            if not jl:
                continue
            entry = getattr(jl, "journal_entry", None) or getattr(jl, "entry", None)
            if not entry:
                continue
            source_url = _build_source_url(
                getattr(entry, "source_document_type", None),
                getattr(entry, "source_document_id", None),
                getattr(entry, "entry_id", None),
            )
            meta = resolve_payment_metadata(
                db,
                getattr(entry, "source_document_type", None),
                getattr(entry, "source_document_id", None),
            )
            gl_matches.append(
                {
                    "journal_line_id": str(jl.line_id),
                    "entry_id": str(entry.entry_id),
                    "entry_date": _format_date(entry.entry_date),
                    "description": jl.description or entry.description or "",
                    "reference": entry.reference or "",
                    "debit_amount": _format_currency(jl.debit_amount, currency),
                    "credit_amount": _format_currency(jl.credit_amount, currency),
                    "account_name": (
                        f"{jl.account.account_code} - {jl.account.account_name}"
                        if jl.account
                        else ""
                    ),
                    "source_url": source_url,
                    "match_detail": _build_match_detail(
                        db, entry, source_url, metadata=meta
                    ),
                    "match_type": match.match_type or "",
                    "match_score": float(match.match_score)
                    if match.match_score
                    else None,
                    "is_primary": match.is_primary,
                    "matched_at": (
                        match.matched_at.strftime("%d %b %Y %H:%M")
                        if match.matched_at
                        else None
                    ),
                }
            )

        # Also check legacy single-match field if no multi-matches found
        if not gl_matches and line.matched_journal_line_id:
            jl = db.get(JournalEntryLine, line.matched_journal_line_id)
            if jl:
                entry = getattr(jl, "journal_entry", None) or getattr(jl, "entry", None)
                if entry:
                    source_url = _build_source_url(
                        getattr(entry, "source_document_type", None),
                        getattr(entry, "source_document_id", None),
                        getattr(entry, "entry_id", None),
                    )
                    meta = resolve_payment_metadata(
                        db,
                        getattr(entry, "source_document_type", None),
                        getattr(entry, "source_document_id", None),
                    )
                    gl_matches.append(
                        {
                            "journal_line_id": str(jl.line_id),
                            "entry_id": str(entry.entry_id),
                            "entry_date": _format_date(entry.entry_date),
                            "description": jl.description or entry.description or "",
                            "reference": entry.reference or "",
                            "debit_amount": _format_currency(jl.debit_amount, currency),
                            "credit_amount": _format_currency(
                                jl.credit_amount, currency
                            ),
                            "account_name": (
                                f"{jl.account.account_code} - {jl.account.account_name}"
                                if jl.account
                                else ""
                            ),
                            "source_url": source_url,
                            "match_detail": _build_match_detail(
                                db, entry, source_url, metadata=meta
                            ),
                            "match_type": "LEGACY",
                            "match_score": None,
                            "is_primary": True,
                            "matched_at": (
                                line.matched_at.strftime("%d %b %Y %H:%M")
                                if line.matched_at
                                else None
                            ),
                        }
                    )

        # Suggested account name lookup
        suggested_account_name = None
        if line.suggested_account_id:
            acct = db.get(Account, line.suggested_account_id)
            if acct:
                suggested_account_name = f"{acct.account_code} - {acct.account_name}"

        # Suggested rule name lookup
        suggested_rule_name = None
        if line.suggested_rule_id:
            from app.models.finance.banking.transaction_rule import TransactionRule

            rule = db.get(TransactionRule, line.suggested_rule_id)
            if rule:
                suggested_rule_name = rule.rule_name

        return {
            "line": line_view,
            "statement": statement_view,
            "account": account_view,
            "gl_matches": gl_matches,
            "suggested_account_name": suggested_account_name,
            "suggested_rule_name": suggested_rule_name,
            "currency_code": currency,
        }

    @staticmethod
    def list_statements_context(
        db: Session,
        organization_id: str,
        account_id: str | None,
        status: str | None,
        start_date: str | None,
        end_date: str | None,
        page: int,
        limit: int = 50,
        sort: str | None = None,
        sort_dir: str | None = None,
        match_status: str | None = None,
        search: str | None = None,
    ) -> dict:
        """Build context for flat bank statement lines page.

        Shows bank account summary cards and imported BankStatementLine
        rows across all statements, with categorization and match status.
        """
        from app.models.finance.banking.bank_statement import CategorizationStatus

        org_id = coerce_uuid(organization_id)
        offset = (page - 1) * limit

        from_date = _parse_date(start_date)
        to_date = _parse_date(end_date)

        # ── Bank accounts (always shown as summary cards) ──
        accounts = list(
            db.scalars(
                select(BankAccount)
                .where(
                    BankAccount.organization_id == org_id,
                    BankAccount.status == BankAccountStatus.active,
                )
                .order_by(BankAccount.bank_name, BankAccount.account_number)
            ).all()
        )
        account_views = [_account_view(a) for a in accounts]

        # Map bank_account_id → BankAccount for enriching line views
        bank_map: dict[UUID, BankAccount] = {a.bank_account_id: a for a in accounts}

        # ── Base conditions for BankStatementLine query ──
        base_conditions: list[Any] = [
            BankStatement.organization_id == org_id,
        ]
        if account_id:
            base_conditions.append(
                BankStatement.bank_account_id == coerce_uuid(account_id)
            )
        if from_date:
            base_conditions.append(BankStatementLine.transaction_date >= from_date)
        if to_date:
            base_conditions.append(BankStatementLine.transaction_date <= to_date)
        if status:
            try:
                cat_status = CategorizationStatus(status)
                base_conditions.append(
                    BankStatementLine.categorization_status == cat_status
                )
            except ValueError:
                pass  # ignore invalid status filter
        if match_status == "matched":
            base_conditions.append(BankStatementLine.is_matched.is_(True))
        elif match_status == "unmatched":
            base_conditions.append(BankStatementLine.is_matched.is_(False))

        search_term = (search or "").strip()
        if search_term:
            like_pat = f"%{search_term}%"
            base_conditions.append(
                or_(
                    BankStatementLine.description.ilike(like_pat),
                    BankStatementLine.reference.ilike(like_pat),
                    BankStatementLine.payee_payer.ilike(like_pat),
                    BankStatementLine.bank_reference.ilike(like_pat),
                )
            )

        join_clause = BankStatementLine.statement_id == BankStatement.statement_id

        # Count
        count_stmt = (
            select(func.count(BankStatementLine.line_id))
            .join(BankStatement, join_clause)
            .where(*base_conditions)
        )
        total_count = db.scalar(count_stmt) or 0

        # Aggregates for stat cards
        agg_stmt = (
            select(
                func.count(BankStatementLine.line_id).label("total"),
                func.count(
                    case(
                        (
                            BankStatementLine.categorization_status.is_(None),
                            BankStatementLine.line_id,
                        ),
                    )
                ).label("uncategorized"),
                func.count(
                    case(
                        (
                            BankStatementLine.categorization_status.in_(
                                [
                                    "SUGGESTED",
                                    "FLAGGED",
                                ]
                            ),
                            BankStatementLine.line_id,
                        ),
                    )
                ).label("suggested"),
                func.count(
                    case(
                        (
                            BankStatementLine.is_matched.is_(True),
                            BankStatementLine.line_id,
                        ),
                    )
                ).label("matched"),
                func.count(
                    case(
                        (
                            BankStatementLine.is_matched.is_(False),
                            BankStatementLine.line_id,
                        ),
                    )
                ).label("unmatched"),
            )
            .join(BankStatement, join_clause)
            .where(*base_conditions)
        )
        agg_row = db.execute(agg_stmt).one()
        total_lines = agg_row.total or 0
        uncategorized_count = agg_row.uncategorized or 0
        suggested_count = agg_row.suggested or 0
        matched_count = agg_row.matched or 0
        unmatched_count = agg_row.unmatched or 0

        # Fetch paginated statement lines
        txn_sort_map: dict[str, Any] = {
            "transaction_date": BankStatementLine.transaction_date,
            "amount": BankStatementLine.amount,
            "description": BankStatementLine.description,
        }
        txn_stmt = (
            select(BankStatementLine, BankStatement)
            .join(BankStatement, join_clause)
            .where(*base_conditions)
        )
        txn_stmt = apply_sort(
            txn_stmt,
            sort,
            sort_dir,
            txn_sort_map,
            default=BankStatementLine.transaction_date.desc(),
        )
        rows = db.execute(txn_stmt.limit(limit).offset(offset)).all()

        # Build line view dicts enriched with bank account info
        transactions: list[dict[str, Any]] = []
        suggested_account_ids: set[UUID] = set()
        matched_jl_ids: set[UUID] = set()
        for line, stmt in rows:
            bank_acct = bank_map.get(stmt.bank_account_id)
            currency = (
                stmt.currency_code
                or (bank_acct.currency_code if bank_acct else None)
                or org_context_service.get_functional_currency(db, organization_id)
            )
            txn = _statement_line_view(line, currency)
            txn["statement_id"] = str(stmt.statement_id)
            txn["statement_number"] = stmt.statement_number or ""
            txn["bank_name"] = bank_acct.bank_name if bank_acct else ""
            txn["account_number"] = bank_acct.account_number if bank_acct else ""
            txn["bank_account_id"] = str(stmt.bank_account_id)
            transactions.append(txn)
            if line.suggested_account_id:
                suggested_account_ids.add(line.suggested_account_id)
            if line.matched_journal_line_id:
                matched_jl_ids.add(line.matched_journal_line_id)

        # Batch-resolve matched journal entry info
        match_detail_map: dict[str, dict[str, str]] = {}
        if matched_jl_ids:
            jl_rows = db.execute(
                select(JournalEntryLine, JournalEntry)
                .join(
                    JournalEntry,
                    JournalEntryLine.journal_entry_id == JournalEntry.journal_entry_id,
                )
                .where(JournalEntryLine.line_id.in_(list(matched_jl_ids)))
            ).all()

            # Batch-resolve payment metadata for source labels
            md_pairs: list[tuple[str | None, UUID | None]] = [
                (
                    getattr(je, "source_document_type", None),
                    getattr(je, "source_document_id", None),
                )
                for _jl, je in jl_rows
            ]
            md_map = resolve_payment_metadata_batch(db, md_pairs)

            for jl, je in jl_rows:
                doc_id = getattr(je, "source_document_id", None)
                meta = md_map.get(doc_id) if doc_id else None
                src_type = getattr(je, "source_document_type", None) or ""
                source_label = _SOURCE_TYPE_LABELS.get(src_type, "Journal")
                counterparty = ""
                if meta and meta.counterparty_name:
                    counterparty = meta.counterparty_name
                match_detail_map[str(jl.line_id)] = {
                    "journal_entry_id": str(je.journal_entry_id),
                    "journal_number": je.journal_number or "",
                    "source_label": source_label,
                    "counterparty": counterparty,
                }

        # Enrich transactions with match details
        for txn in transactions:
            jl_id = txn.get("matched_journal_line_id")
            detail = match_detail_map.get(jl_id or "") if jl_id else None
            if detail:
                txn["match_journal_entry_id"] = detail["journal_entry_id"]
                txn["match_journal_number"] = detail["journal_number"]
                txn["match_source_label"] = detail["source_label"]
                txn["match_counterparty"] = detail["counterparty"]

        # Build account name map for suggested accounts
        account_map: dict[str, str] = {}
        if suggested_account_ids:
            gl_accounts = db.scalars(
                select(Account).where(
                    Account.organization_id == org_id,
                    Account.account_id.in_(list(suggested_account_ids)),
                )
            ).all()
            account_map = {
                str(a.account_id): f"{a.account_code} - {a.account_name}"
                for a in gl_accounts
            }

        # Always show the Category column so the workflow is discoverable
        has_category_data = True

        total_pages = max(1, (total_count + limit - 1) // limit)

        # Status filter labels
        cat_status_labels: dict[str, str] = {
            "SUGGESTED": "Suggested",
            "ACCEPTED": "Accepted",
            "REJECTED": "Rejected",
            "AUTO_APPLIED": "Auto-applied",
            "FLAGGED": "Flagged",
        }
        match_status_labels: dict[str, str] = {
            "matched": "Matched",
            "unmatched": "Unmatched",
        }

        active_filters = _build_active_filters(
            account_id=account_id,
            accounts=account_views,
            status=status,
            start_date=start_date,
            end_date=end_date,
            status_labels=cat_status_labels,
        )
        if match_status and match_status in match_status_labels:
            active_filters.append(
                {
                    "name": "match_status",
                    "value": match_status,
                    "display_value": match_status_labels[match_status],
                }
            )
        if search_term:
            active_filters.append(
                {
                    "name": "search",
                    "value": search_term,
                    "display_value": f'Search: "{search_term}"',
                }
            )

        return {
            "transactions": transactions,
            "accounts": account_views,
            "account_id": account_id,
            "status": status,
            "match_status": match_status,
            "search": search_term,
            "sort": sort,
            "sort_dir": sort_dir,
            "start_date": start_date,
            "end_date": end_date,
            "page": page,
            "limit": limit,
            "offset": offset,
            "total_count": total_count,
            "total_pages": total_pages,
            "total_lines": total_lines,
            "uncategorized_count": uncategorized_count,
            "suggested_count": suggested_count,
            "matched_count": matched_count,
            "unmatched_count": unmatched_count,
            "has_category_data": has_category_data,
            "account_map": account_map,
            "active_filters": active_filters,
        }

    @staticmethod
    def statement_import_context(
        db: Session,
        organization_id: str,
    ) -> dict:
        org_id = coerce_uuid(organization_id)
        accounts = db.scalars(
            select(BankAccount)
            .where(BankAccount.organization_id == org_id)
            .order_by(BankAccount.bank_name, BankAccount.account_number)
        ).all()
        # Build JSON-safe account list for Alpine.js tojson serialization.
        accounts_json = [
            {
                "bank_account_id": str(a.bank_account_id),
                "bank_name": a.bank_name or "",
                "account_number": a.account_number or "",
            }
            for a in accounts
        ]
        # Build alias map from unified registry for client-side header matching
        from app.services.finance.banking.bank_statement import (
            BankStatementService,
        )
        from app.services.finance.import_export.base import build_alias_map

        alias_map = build_alias_map(BankStatementService._BANK_FIELD_TYPES)
        return {"accounts": accounts_json, "alias_map": alias_map}

    @staticmethod
    def statement_detail_context(
        db: Session,
        organization_id: str,
        statement_id: str,
        page: int = 1,
        limit: int = 50,
    ) -> dict:
        org_id = coerce_uuid(organization_id)
        statement = db.get(BankStatement, coerce_uuid(statement_id))
        if not statement or statement.organization_id != org_id:
            return {"statement": None, "lines": [], "account_map": {}}

        currency = statement.currency_code
        total_count = (
            db.scalar(
                select(func.count(BankStatementLine.line_id)).where(
                    BankStatementLine.statement_id == statement.statement_id
                )
            )
            or 0
        )
        total_pages = max(1, (total_count + limit - 1) // limit)
        offset = (page - 1) * limit
        paged_lines = list(
            db.scalars(
                select(BankStatementLine)
                .where(BankStatementLine.statement_id == statement.statement_id)
                .order_by(BankStatementLine.transaction_date)
                .offset(offset)
                .limit(limit)
            ).all()
        )
        lines = [_statement_line_view(line, currency) for line in paged_lines]

        # Build account name lookup for suggested accounts
        account_ids = [
            line.suggested_account_id
            for line in paged_lines
            if line.suggested_account_id
        ]
        account_map: dict[str, str] = {}
        if account_ids:
            accounts = db.scalars(
                select(Account).where(
                    Account.organization_id == org_id,
                    Account.account_id.in_(account_ids),
                )
            ).all()
            account_map = {
                str(a.account_id): f"{a.account_code} - {a.account_name}"
                for a in accounts
            }

        # Categorization summary counts (SQL aggregation, not in-memory)
        from app.models.finance.banking.bank_statement import CategorizationStatus

        cat_rows = db.execute(
            select(
                BankStatementLine.categorization_status,
                func.count(BankStatementLine.line_id),
            )
            .where(BankStatementLine.statement_id == statement.statement_id)
            .group_by(BankStatementLine.categorization_status)
        ).all()
        cat_summary = {
            "suggested": 0,
            "accepted": 0,
            "rejected": 0,
            "auto_applied": 0,
            "flagged": 0,
        }
        _cat_key_map = {
            CategorizationStatus.SUGGESTED: "suggested",
            CategorizationStatus.ACCEPTED: "accepted",
            CategorizationStatus.REJECTED: "rejected",
            CategorizationStatus.AUTO_APPLIED: "auto_applied",
            CategorizationStatus.FLAGGED: "flagged",
        }
        for cat_status, cnt in cat_rows:
            key = _cat_key_map.get(cat_status)
            if key:
                cat_summary[key] = cnt

        # GL transaction match suggestions for unmatched lines
        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
        )

        recon_svc = BankReconciliationService()
        match_suggestions_raw = recon_svc.get_statement_match_suggestions(
            db, org_id, statement.statement_id
        )

        # Serialize to JSON-safe dict keyed by string line_id
        match_suggestions: dict[str, dict] = {}
        visible_line_ids = {str(line.line_id) for line in paged_lines}
        for line_id, suggestion in match_suggestions_raw.items():
            line_id_str = str(line_id)
            if line_id_str not in visible_line_ids:
                continue
            match_suggestions[line_id_str] = {
                "journal_line_id": str(suggestion.journal_line_id),
                "confidence": suggestion.confidence,
                "counterparty_name": suggestion.counterparty_name,
                "payment_number": suggestion.payment_number,
                "source_url": suggestion.source_url,
                "amount_matched": suggestion.amount_matched,
            }

        # GL candidates are now lazy-loaded per line via the scored
        # candidates API endpoint — no longer sent at page load.

        # Resolve source URLs for matched lines and build line_amounts map
        from app.services.finance.banking.bank_reconciliation import (
            _build_source_url,
        )

        matched_jl_ids = [
            line.matched_journal_line_id
            for line in paged_lines
            if line.is_matched and line.matched_journal_line_id
        ]
        matched_source_urls: dict[str, str] = {}
        # Map journal_line_id → JournalEntry for metadata resolution
        jl_entry_map: dict[str, JournalEntry] = {}
        if matched_jl_ids:
            jl_rows = (
                db.execute(
                    select(JournalEntryLine)
                    .join(JournalEntry)
                    .where(JournalEntryLine.line_id.in_(matched_jl_ids))
                )
                .scalars()
                .all()
            )
            for jl in jl_rows:
                entry = getattr(jl, "journal_entry", None) or getattr(jl, "entry", None)
                if entry:
                    url = _build_source_url(
                        getattr(entry, "source_document_type", None),
                        getattr(entry, "source_document_id", None),
                        getattr(entry, "entry_id", None),
                    )
                    matched_source_urls[str(jl.line_id)] = url
                    jl_entry_map[str(jl.line_id)] = entry

        # Batch-resolve payment metadata for matched lines
        metadata_pairs: list[tuple[str | None, UUID | None]] = [
            (
                getattr(e, "source_document_type", None),
                getattr(e, "source_document_id", None),
            )
            for e in jl_entry_map.values()
        ]
        metadata_by_doc_id = resolve_payment_metadata_batch(db, metadata_pairs)

        # Build match_details keyed by statement line_id
        # Map statement line → matched journal line id for lookup
        stmt_line_to_jl: dict[str, str] = {}
        for line in paged_lines:
            if line.is_matched and line.matched_journal_line_id:
                stmt_line_to_jl[str(line.line_id)] = str(line.matched_journal_line_id)

        match_details: dict[str, dict[str, str]] = {}
        for stmt_lid, jl_id in stmt_line_to_jl.items():
            entry = jl_entry_map.get(jl_id)
            if not entry:
                continue
            src_doc_id = getattr(entry, "source_document_id", None)
            meta = metadata_by_doc_id.get(src_doc_id) if src_doc_id else None
            url = matched_source_urls.get(jl_id, "")
            match_details[stmt_lid] = _build_match_detail(db, entry, url, metadata=meta)

        # Merge matched_source_url into line views
        line_amounts: dict[str, float] = {}
        for lv in lines:
            lid = str(lv["line_id"])
            line_amounts[lid] = lv["raw_amount"]
            jl_id = lv.get("matched_journal_line_id")
            if lv["is_matched"] and jl_id:
                lv["matched_source_url"] = matched_source_urls.get(jl_id, "")
            else:
                lv["matched_source_url"] = ""

        # Build line_details for modal context card (issue #4)
        # Keyed by string line_id with essential info for visual comparison
        line_details: dict[str, dict] = {}
        for lv in lines:
            lid = str(lv["line_id"])
            line_details[lid] = {
                "date": lv["transaction_date"],
                "description": lv["description"] or "",
                "payee": lv["payee_payer"] or "",
                "amount": lv["amount"],
                "raw_amount": lv["raw_amount"],
                "is_credit": lv["transaction_type"] == "credit",
            }

        # Check if any lines have balance/category data (issue #16/#17)
        has_balance_data = any(line.running_balance is not None for line in paged_lines)
        has_category_data = any(
            line.categorization_status is not None for line in paged_lines
        )

        # GL accounts for "Create Journal & Match" feature
        from app.models.finance.gl.account import Account as GLAccount

        gl_accounts_raw = list(
            db.scalars(
                select(GLAccount)
                .where(
                    GLAccount.organization_id == org_id,
                    GLAccount.is_active == True,  # noqa: E712
                )
                .order_by(GLAccount.account_code)
            ).all()
        )
        gl_accounts = [
            {
                "id": str(a.account_id),
                "code": a.account_code,
                "name": a.account_name,
                "label": f"{a.account_code} - {a.account_name}",
            }
            for a in gl_accounts_raw
        ]
        # Bank account's GL account id for filtering
        bank_gl_account_id = (
            str(statement.bank_account.gl_account_id)
            if statement.bank_account
            and getattr(statement.bank_account, "gl_account_id", None)
            else ""
        )

        return {
            "statement": _statement_view(statement),
            "lines": lines,
            "page": page,
            "limit": limit,
            "offset": offset,
            "total_count": total_count,
            "total_pages": total_pages,
            "account_map": account_map,
            "categorization_summary": cat_summary,
            "match_suggestions": match_suggestions,
            "match_details": match_details,
            "line_amounts": line_amounts,
            "line_details": line_details,
            "statement_currency": currency,
            "has_balance_data": has_balance_data,
            "has_category_data": has_category_data,
            "gl_accounts": gl_accounts,
            "bank_gl_account_id": bank_gl_account_id,
        }

    @staticmethod
    def list_reconciliations_context(
        db: Session,
        organization_id: str,
        account_id: str | None,
        status: str | None,
        start_date: str | None,
        end_date: str | None,
        page: int,
        limit: int = 50,
    ) -> dict:
        org_id = coerce_uuid(organization_id)
        offset = (page - 1) * limit

        status_value = _parse_reconciliation_status(status)
        from_date = _parse_date(start_date)
        to_date = _parse_date(end_date)

        conditions: list[Any] = [BankReconciliation.organization_id == org_id]

        if account_id:
            conditions.append(
                BankReconciliation.bank_account_id == coerce_uuid(account_id)
            )
        if status_value:
            conditions.append(BankReconciliation.status == status_value)
        if from_date:
            conditions.append(BankReconciliation.reconciliation_date >= from_date)
        if to_date:
            conditions.append(BankReconciliation.reconciliation_date <= to_date)

        total_count = (
            db.scalar(
                select(func.count(BankReconciliation.reconciliation_id)).where(
                    *conditions
                )
            )
            or 0
        )
        reconciliations = db.scalars(
            select(BankReconciliation)
            .where(*conditions)
            .order_by(BankReconciliation.reconciliation_date.desc())
            .limit(limit)
            .offset(offset)
        ).all()

        accounts = db.scalars(
            select(BankAccount)
            .where(BankAccount.organization_id == org_id)
            .order_by(BankAccount.bank_name, BankAccount.account_number)
        ).all()

        in_progress_count = sum(
            1 for recon in reconciliations if recon.status == ReconciliationStatus.draft
        )
        pending_review_count = sum(
            1
            for recon in reconciliations
            if recon.status == ReconciliationStatus.pending_review
        )
        approved_count = sum(
            1
            for recon in reconciliations
            if recon.status == ReconciliationStatus.approved
        )

        total_pages = max(1, (total_count + limit - 1) // limit)

        account_views = [_account_view(account) for account in accounts]
        active_filters = _build_active_filters(
            account_id=account_id,
            accounts=account_views,
            status=status,
            start_date=start_date,
            end_date=end_date,
        )

        return {
            "reconciliations": [
                _reconciliation_view(recon) for recon in reconciliations
            ],
            "accounts": account_views,
            "account_id": account_id,
            "status": status,
            "start_date": start_date,
            "end_date": end_date,
            "page": page,
            "limit": limit,
            "offset": offset,
            "total_count": total_count,
            "total_pages": total_pages,
            "in_progress_count": in_progress_count,
            "pending_review_count": pending_review_count,
            "approved_count": approved_count,
            "statuses": [s.value for s in ReconciliationStatus],
            "active_filters": active_filters,
        }

    @staticmethod
    def reconciliation_form_context(
        db: Session,
        organization_id: str,
        *,
        account_id: str | None = None,
    ) -> dict:
        org_id = coerce_uuid(organization_id)
        accounts = db.scalars(
            select(BankAccount)
            .where(
                BankAccount.organization_id == org_id,
                BankAccount.status == BankAccountStatus.active,
            )
            .order_by(BankAccount.bank_name, BankAccount.account_number)
        ).all()
        context: dict = {"accounts": [_account_view(account) for account in accounts]}
        if account_id:
            context["selected_account_id"] = account_id
        return context

    @staticmethod
    def reconciliation_detail_context(
        db: Session,
        organization_id: str,
        reconciliation_id: str,
    ) -> dict:
        from app.services.finance.banking.bank_reconciliation import (
            bank_reconciliation_service as recon_svc,
        )

        org_id = coerce_uuid(organization_id)
        reconciliation = db.get(BankReconciliation, coerce_uuid(reconciliation_id))
        if not reconciliation or reconciliation.organization_id != org_id:
            return {
                "reconciliation": None,
                "lines": [],
                "unmatched_statement_lines": [],
                "unmatched_gl_lines": [],
                "match_suggestions": {},
            }

        bank_account = reconciliation.bank_account

        statement_lines = db.scalars(
            select(BankStatementLine)
            .join(
                BankStatement,
                BankStatementLine.statement_id == BankStatement.statement_id,
            )
            .where(
                BankStatement.organization_id == org_id,
                BankStatement.bank_account_id == reconciliation.bank_account_id,
                BankStatementLine.is_matched.is_(False),
                BankStatementLine.transaction_date >= reconciliation.period_start,
                BankStatementLine.transaction_date <= reconciliation.period_end,
            )
            .order_by(BankStatementLine.transaction_date, BankStatementLine.line_number)
        ).all()

        gl_lines: list[Any] = []
        if bank_account:
            gl_lines = db.execute(
                select(JournalEntryLine, JournalEntry)
                .join(
                    JournalEntry,
                    JournalEntryLine.journal_entry_id == JournalEntry.journal_entry_id,
                )
                .where(
                    JournalEntry.organization_id == org_id,
                    JournalEntryLine.account_id == bank_account.gl_account_id,
                    JournalEntry.status == JournalStatus.POSTED,
                    JournalEntry.entry_date >= reconciliation.period_start,
                    JournalEntry.entry_date <= reconciliation.period_end,
                )
                .order_by(JournalEntry.entry_date, JournalEntryLine.line_number)
            ).all()

        # Batch-resolve payment metadata for GL lines
        metadata_pairs: list[tuple[str | None, UUID | None]] = [
            (
                getattr(entry, "source_document_type", None),
                getattr(entry, "source_document_id", None),
            )
            for _line, entry in gl_lines
        ]
        metadata_map = resolve_payment_metadata_batch(db, metadata_pairs)

        # Build GL line views with metadata
        unmatched_statement_lines = [
            _statement_line_view(line) for line in statement_lines
        ]
        statement_line_amounts = {
            str(line.line_id): float(line.signed_amount) for line in statement_lines
        }
        unmatched_gl_lines = []
        gl_line_amounts: dict[str, float] = {}
        for line, entry in gl_lines:
            doc_id = getattr(entry, "source_document_id", None)
            meta = metadata_map.get(doc_id) if doc_id else None
            line_view = _gl_line_view(line, entry, metadata=meta)
            unmatched_gl_lines.append(line_view)
            gl_line_amounts[str(line.line_id)] = float(
                line_view.get("signed_amount", 0)
            )

        # Generate match suggestions for draft/pending reconciliations
        match_suggestions: dict[str, dict] = {}
        if reconciliation.status in (
            ReconciliationStatus.draft,
            ReconciliationStatus.pending_review,
        ):
            try:
                raw_suggestions = recon_svc.get_match_suggestions(
                    db, org_id, reconciliation.reconciliation_id
                )
                for stmt_id, sug in raw_suggestions.items():
                    match_suggestions[str(stmt_id)] = {
                        "journal_line_id": str(sug.journal_line_id),
                        "confidence": round(sug.confidence, 1),
                        "counterparty_name": sug.counterparty_name or "",
                        "payment_number": sug.payment_number or "",
                        "source_url": sug.source_url or "",
                        "amount_matched": sug.amount_matched,
                    }
            except Exception:
                logger.exception("Failed to generate match suggestions")

        return {
            "reconciliation": _reconciliation_view(reconciliation),
            "lines": [_reconciliation_line_view(line) for line in reconciliation.lines],
            "unmatched_statement_lines": unmatched_statement_lines,
            "unmatched_gl_lines": unmatched_gl_lines,
            "match_suggestions": match_suggestions,
            "statement_line_amounts": statement_line_amounts,
            "gl_line_amounts": gl_line_amounts,
        }

    @staticmethod
    def reconciliation_report_context(
        db: Session,
        organization_id: str,
        reconciliation_id: str,
    ) -> dict:
        org_id = coerce_uuid(organization_id)
        reconciliation = db.get(BankReconciliation, coerce_uuid(reconciliation_id))
        if not reconciliation or reconciliation.organization_id != org_id:
            return {"report": None}

        recon_view = _reconciliation_view(reconciliation)

        matched_lines = [
            line
            for line in reconciliation.lines
            if not line.is_outstanding and not line.is_adjustment
        ]
        outstanding_deposits = [
            line
            for line in reconciliation.lines
            if line.is_outstanding and line.outstanding_type == "deposit"
        ]
        outstanding_payments = [
            line
            for line in reconciliation.lines
            if line.is_outstanding and line.outstanding_type == "payment"
        ]
        adjustments = [line for line in reconciliation.lines if line.is_adjustment]

        total_matched = sum(
            (_line_amount(line) for line in matched_lines), Decimal("0")
        )
        total_deposits = sum(
            (_line_amount(line) for line in outstanding_deposits), Decimal("0")
        )
        total_payments = sum(
            (_line_amount(line) for line in outstanding_payments), Decimal("0")
        )
        total_adjustments = sum(
            (_line_amount(line) for line in adjustments), Decimal("0")
        )

        statement_balance = Decimal(str(reconciliation.statement_closing_balance or 0))
        gl_balance = Decimal(str(reconciliation.gl_closing_balance or 0))
        adjusted_statement = statement_balance - total_payments + total_deposits
        adjusted_gl = gl_balance + total_adjustments
        difference = adjusted_statement - adjusted_gl

        report = {
            "reconciliation": recon_view,
            "summary": {
                "statement_balance": _format_currency(
                    statement_balance, reconciliation.currency_code
                ),
                "gl_balance": _format_currency(
                    gl_balance, reconciliation.currency_code
                ),
                "adjusted_book_balance": _format_currency(
                    adjusted_statement, reconciliation.currency_code
                ),
                "difference": _format_currency(
                    difference, reconciliation.currency_code
                ),
                "is_reconciled": difference == Decimal("0"),
            },
            "matched_items": {
                "count": len(matched_lines),
                "total": _format_currency(total_matched, reconciliation.currency_code),
                "items": [_reconciliation_line_view(line) for line in matched_lines],
            },
            "outstanding_deposits": {
                "count": len(outstanding_deposits),
                "total": _format_currency(total_deposits, reconciliation.currency_code),
                "items": [
                    _reconciliation_line_view(line) for line in outstanding_deposits
                ],
            },
            "outstanding_payments": {
                "count": len(outstanding_payments),
                "total": _format_currency(total_payments, reconciliation.currency_code),
                "items": [
                    _reconciliation_line_view(line) for line in outstanding_payments
                ],
            },
            "adjustments": {
                "count": len(adjustments),
                "total": _format_currency(
                    total_adjustments, reconciliation.currency_code
                ),
                "items": [_reconciliation_line_view(line) for line in adjustments],
            },
        }

        return {"report": report}

    # =========================================================================
    # Payee Context Methods
    # =========================================================================

    @staticmethod
    def list_payees_context(
        db: Session,
        organization_id: str,
        search: str | None = None,
        payee_type: str | None = None,
        page: int = 1,
        per_page: int = 25,
    ) -> dict:
        """Context for payees list page."""
        from app.models.finance.banking.payee import Payee, PayeeType

        org_id = coerce_uuid(organization_id)

        conditions: list[Any] = [
            Payee.organization_id == org_id,
            Payee.is_active.is_(True),
        ]

        if search:
            search_pattern = f"%{search}%"
            conditions.append(
                or_(
                    Payee.payee_name.ilike(search_pattern),
                    Payee.name_patterns.ilike(search_pattern),
                )
            )

        if payee_type:
            try:
                pt = PayeeType(payee_type)
                conditions.append(Payee.payee_type == pt)
            except ValueError:
                pass

        total = db.scalar(select(func.count(Payee.payee_id)).where(*conditions)) or 0
        payees = db.scalars(
            select(Payee)
            .where(*conditions)
            .order_by(Payee.payee_name)
            .offset((page - 1) * per_page)
            .limit(per_page)
        ).all()

        # Get GL accounts for display
        account_map = {}
        account_ids = [p.default_account_id for p in payees if p.default_account_id]
        if account_ids:
            accounts = db.scalars(
                select(Account).where(
                    Account.organization_id == org_id,
                    Account.account_id.in_(account_ids),
                )
            ).all()
            account_map = {
                a.account_id: f"{a.account_code} - {a.account_name}" for a in accounts
            }

        payee_list = []
        for p in payees:
            default_account_id = p.default_account_id
            payee_list.append(
                {
                    "payee_id": str(p.payee_id),
                    "payee_name": p.payee_name,
                    "payee_type": p.payee_type.value if p.payee_type else "",
                    "name_patterns": p.name_patterns or "",
                    "default_account": account_map.get(default_account_id, "")
                    if default_account_id
                    else "",
                    "match_count": p.match_count,
                    "last_matched": _format_date(p.last_matched_at)
                    if p.last_matched_at
                    else "Never",
                }
            )

        total_pages = (total + per_page - 1) // per_page
        active_filters = build_active_filters(
            params={"search": search, "payee_type": payee_type},
            labels={"search": "Search", "payee_type": "Type"},
            options={
                "payee_type": {
                    t.value: t.value.replace("_", " ").title() for t in PayeeType
                }
            },
        )
        return {
            "payees": payee_list,
            "payee_types": [
                {"value": t.value, "label": t.value.replace("_", " ").title()}
                for t in PayeeType
            ],
            "search": search or "",
            "payee_type": payee_type or "",
            "selected_type": payee_type or "",
            "active_filters": active_filters,
            "page": page,
            "limit": per_page,
            "total_count": total,
            "total_pages": total_pages,
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total": total,
                "pages": total_pages,
            },
        }

    @staticmethod
    def payee_form_context(
        db: Session,
        organization_id: str,
        payee_id: str | None = None,
    ) -> dict:
        """Context for payee create/edit form."""
        from app.models.finance.banking.payee import Payee, PayeeType

        org_id = coerce_uuid(organization_id)

        # Get GL accounts for dropdown (template uses model objects)
        gl_accounts = list(
            db.scalars(
                select(Account)
                .where(
                    Account.organization_id == org_id,
                    Account.is_active.is_(True),
                )
                .order_by(Account.account_code)
            ).all()
        )

        # Tax codes for dropdown
        from app.models.finance.tax.tax_code import TaxCode

        tax_codes = list(
            db.scalars(
                select(TaxCode)
                .where(
                    TaxCode.organization_id == org_id,
                    TaxCode.is_active.is_(True),
                )
                .order_by(TaxCode.tax_code)
            ).all()
        )

        # Suppliers and customers for optional linking
        from app.models.finance.ap.supplier import Supplier
        from app.models.finance.ar.customer import Customer

        suppliers = list(
            db.scalars(
                select(Supplier)
                .where(Supplier.organization_id == org_id)
                .order_by(Supplier.legal_name)
            ).all()
        )
        customers = list(
            db.scalars(
                select(Customer)
                .where(Customer.organization_id == org_id)
                .order_by(Customer.legal_name)
            ).all()
        )

        payee = None
        if payee_id:
            payee = db.get(Payee, coerce_uuid(payee_id))
            if payee and payee.organization_id != org_id:
                payee = None

        return {
            "payee": payee,
            "is_edit": payee is not None,
            "payee_types": [
                {"value": t.value, "label": t.value.replace("_", " ").title()}
                for t in PayeeType
            ],
            "gl_accounts": gl_accounts,
            "tax_codes": tax_codes,
            "suppliers": suppliers,
            "customers": customers,
        }

    # =========================================================================
    # Transaction Rule Context Methods
    # =========================================================================

    @staticmethod
    def list_rules_context(
        db: Session,
        organization_id: str,
        rule_type: str | None = None,
        page: int = 1,
        per_page: int = 25,
    ) -> dict:
        """Context for transaction rules list page."""
        from app.models.finance.banking.transaction_rule import (
            RuleType,
            TransactionRule,
        )

        org_id = coerce_uuid(organization_id)

        conditions: list[Any] = [TransactionRule.organization_id == org_id]

        if rule_type:
            try:
                rt = RuleType(rule_type)
                conditions.append(TransactionRule.rule_type == rt)
            except ValueError:
                pass

        total = (
            db.scalar(select(func.count(TransactionRule.rule_id)).where(*conditions))
            or 0
        )
        rules = db.scalars(
            select(TransactionRule)
            .where(*conditions)
            .order_by(TransactionRule.sort_order.asc(), TransactionRule.rule_name)
            .offset((page - 1) * per_page)
            .limit(per_page)
        ).all()

        # Get GL accounts for display
        account_map = {}
        account_ids = [r.target_account_id for r in rules if r.target_account_id]
        if account_ids:
            accounts = db.scalars(
                select(Account).where(
                    Account.organization_id == org_id,
                    Account.account_id.in_(account_ids),
                )
            ).all()
            account_map = {
                a.account_id: f"{a.account_code} - {a.account_name}" for a in accounts
            }

        rule_list = []
        for idx, r in enumerate(rules):
            target_account_id = r.target_account_id
            rule_list.append(
                {
                    "rule_id": str(r.rule_id),
                    "rule_name": r.rule_name,
                    "description": r.description or "",
                    "rule_type": r.rule_type.value if r.rule_type else "",
                    "action": r.action.value if r.action else "",
                    "target_account": account_map.get(target_account_id, "")
                    if target_account_id
                    else "",
                    "sort_order": r.sort_order,
                    "position": idx + 1,
                    "auto_apply": r.auto_apply,
                    "is_active": r.is_active,
                    "match_count": r.match_count,
                    "success_count": r.success_count,
                    "success_rate": f"{r.success_rate:.0f}%"
                    if r.success_count + r.reject_count > 0
                    else "N/A",
                }
            )

        # Aggregate stats across ALL rules for the org (not just current page)
        active_count = (
            db.scalar(
                select(func.count(TransactionRule.rule_id)).where(
                    TransactionRule.organization_id == org_id,
                    TransactionRule.is_active.is_(True),
                )
            )
            or 0
        )
        auto_apply_count = (
            db.scalar(
                select(func.count(TransactionRule.rule_id)).where(
                    TransactionRule.organization_id == org_id,
                    TransactionRule.auto_apply.is_(True),
                )
            )
            or 0
        )
        total_matches = (
            db.scalar(
                select(func.coalesce(func.sum(TransactionRule.match_count), 0)).where(
                    TransactionRule.organization_id == org_id
                )
            )
            or 0
        )

        total_pages = (total + per_page - 1) // per_page
        active_filters = build_active_filters(
            params={"rule_type": rule_type},
            labels={"rule_type": "Type"},
            options={
                "rule_type": {
                    t.value: t.value.replace("_", " ").title() for t in RuleType
                }
            },
        )
        return {
            "rules": rule_list,
            "rule_types": [
                {"value": t.value, "label": t.value.replace("_", " ").title()}
                for t in RuleType
            ],
            "rule_type": rule_type or "",
            "selected_type": rule_type or "",
            "active_filters": active_filters,
            "active_count": active_count,
            "auto_apply_count": auto_apply_count,
            "total_matches": total_matches,
            "page": page,
            "limit": per_page,
            "total_count": total,
            "total_pages": total_pages,
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total": total,
                "pages": total_pages,
            },
        }

    @staticmethod
    def _normalize_to_combined(rule_type: str, conditions: dict) -> dict:
        """Normalize a single-type rule's conditions into COMBINED format.

        The visual builder always works with the COMBINED structure:
        ``{"operator": "AND", "rules": [{"type": "...", "conditions": {...}}]}``

        Legacy rules stored as single types get wrapped so the builder can
        display them.
        """
        if rule_type == "COMBINED":
            return conditions

        return {
            "operator": "AND",
            "rules": [{"type": rule_type, "conditions": conditions}],
        }

    @staticmethod
    def rule_form_context(
        db: Session,
        organization_id: str,
        rule_id: str | None = None,
    ) -> dict:
        """Context for transaction rule create/edit form."""
        from app.models.finance.banking.payee import Payee
        from app.models.finance.banking.transaction_rule import (
            RuleAction,
            RuleType,
            TransactionRule,
        )
        from app.models.finance.tax.tax_code import TaxCode

        org_id = coerce_uuid(organization_id)

        # Get GL accounts for dropdown
        accounts = db.scalars(
            select(Account)
            .where(Account.organization_id == org_id, Account.is_active.is_(True))
            .order_by(Account.account_code)
        ).all()

        account_options = [
            {
                "value": str(a.account_id),
                "label": f"{a.account_code} - {a.account_name}",
            }
            for a in accounts
        ]

        # Get bank accounts for dropdown
        bank_accounts = db.scalars(
            select(BankAccount)
            .where(
                BankAccount.organization_id == org_id,
                BankAccount.status == BankAccountStatus.active,
            )
            .order_by(BankAccount.account_name)
        ).all()

        bank_account_options = [
            {
                "value": str(ba.bank_account_id),
                "label": f"{ba.bank_name} - {ba.account_name}",
            }
            for ba in bank_accounts
        ]

        # Get payees for dropdown
        payees = db.scalars(
            select(Payee)
            .where(Payee.organization_id == org_id, Payee.is_active.is_(True))
            .order_by(Payee.payee_name)
        ).all()

        payee_options = [
            {"value": str(p.payee_id), "label": p.payee_name} for p in payees
        ]

        # Get tax codes for dropdown
        tax_codes = db.scalars(
            select(TaxCode)
            .where(TaxCode.organization_id == org_id, TaxCode.is_active.is_(True))
            .order_by(TaxCode.tax_code)
        ).all()

        tax_code_options = [
            {
                "value": str(tc.tax_code_id),
                "label": f"{tc.tax_code} - {tc.tax_name} ({tc.tax_rate}%)",
            }
            for tc in tax_codes
        ]

        rule = None
        if rule_id:
            rule = db.get(TransactionRule, coerce_uuid(rule_id))
            if rule and rule.organization_id != org_id:
                rule = None

        rule_data = None
        if rule:
            raw_type = rule.rule_type.value if rule.rule_type else "COMBINED"
            raw_conditions = rule.conditions or {}
            combined_conditions = BankingWebService._normalize_to_combined(
                raw_type, raw_conditions
            )

            rule_data = {
                "rule_id": str(rule.rule_id),
                "rule_name": rule.rule_name,
                "description": rule.description or "",
                "rule_type": raw_type,
                "conditions": combined_conditions,
                "action": rule.action.value if rule.action else "",
                "target_account_id": str(rule.target_account_id)
                if rule.target_account_id
                else "",
                "tax_code_id": str(rule.tax_code_id) if rule.tax_code_id else "",
                "bank_account_id": str(rule.bank_account_id)
                if rule.bank_account_id
                else "",
                "payee_id": str(rule.payee_id) if rule.payee_id else "",
                "auto_apply": rule.auto_apply,
                "min_confidence": rule.min_confidence,
                "applies_to_credits": rule.applies_to_credits,
                "applies_to_debits": rule.applies_to_debits,
                "is_active": rule.is_active,
                "match_count": rule.match_count,
                "success_count": rule.success_count,
                "reject_count": rule.reject_count,
                "created_at": rule.created_at,
            }

        return {
            "rule": rule_data,
            "is_edit": rule is not None,
            "rule_types": [
                {"value": t.value, "label": t.value.replace("_", " ").title()}
                for t in RuleType
            ],
            "actions": [
                {"value": a.value, "label": a.value.replace("_", " ").title()}
                for a in RuleAction
            ],
            "accounts": account_options,
            "bank_accounts": bank_account_options,
            "payees": payee_options,
            "tax_codes": tax_code_options,
        }

    @staticmethod
    def rule_duplicate_form_context(
        db: Session,
        organization_id: str,
        rule_id: str,
    ) -> dict:
        """Context for duplicate-rule form."""
        from app.models.finance.banking.transaction_rule import TransactionRule

        org_id = coerce_uuid(organization_id)
        source_rule = db.get(TransactionRule, coerce_uuid(rule_id))
        if not source_rule or source_rule.organization_id != org_id:
            raise HTTPException(status_code=404, detail="Rule not found")

        bank_accounts = db.scalars(
            select(BankAccount)
            .where(
                BankAccount.organization_id == org_id,
                BankAccount.status == BankAccountStatus.active,
            )
            .order_by(BankAccount.account_name)
        ).all()

        # Exclude the source rule's bank account — it's already covered
        source_ba_id = source_rule.bank_account_id
        return {
            "source_rule": {
                "rule_id": str(source_rule.rule_id),
                "rule_name": source_rule.rule_name,
                "rule_type": source_rule.rule_type.value
                if source_rule.rule_type
                else "",
                "action": source_rule.action.value if source_rule.action else "",
                "bank_account_id": str(source_ba_id) if source_ba_id else "",
                "bank_account_label": next(
                    (
                        f"{ba.bank_name} - {ba.account_name}"
                        for ba in bank_accounts
                        if ba.bank_account_id == source_ba_id
                    ),
                    "All Accounts",
                ),
            },
            "bank_accounts": [
                {
                    "bank_account_id": str(ba.bank_account_id),
                    "label": f"{ba.bank_name} - {ba.account_name}",
                    "is_default": False,
                }
                for ba in bank_accounts
                if ba.bank_account_id != source_ba_id
            ],
        }

    @staticmethod
    def bulk_rule_duplicate_form_context(
        db: Session,
        organization_id: str,
        rule_ids: list[str],
    ) -> dict:
        """Context for bulk duplicate-rule form."""
        from app.models.finance.banking.transaction_rule import TransactionRule

        org_id = coerce_uuid(organization_id)
        parsed_ids = [coerce_uuid(rid) for rid in rule_ids if rid]
        unique_ids = list(dict.fromkeys(parsed_ids))
        if not unique_ids:
            raise ValueError("Select at least one rule")

        rules = db.scalars(
            select(TransactionRule)
            .where(
                TransactionRule.organization_id == org_id,
                TransactionRule.rule_id.in_(unique_ids),
            )
            .order_by(TransactionRule.rule_name.asc())
        ).all()
        if not rules:
            raise ValueError("No valid rules selected")

        bank_accounts = db.scalars(
            select(BankAccount)
            .where(
                BankAccount.organization_id == org_id,
                BankAccount.status == BankAccountStatus.active,
            )
            .order_by(BankAccount.account_name)
        ).all()

        return {
            "source_rules": [
                {
                    "rule_id": str(rule.rule_id),
                    "rule_name": rule.rule_name,
                    "rule_type": rule.rule_type.value if rule.rule_type else "",
                    "action": rule.action.value if rule.action else "",
                }
                for rule in rules
            ],
            "bank_accounts": [
                {
                    "bank_account_id": str(ba.bank_account_id),
                    "label": f"{ba.bank_name} - {ba.account_name}",
                }
                for ba in bank_accounts
            ],
        }

    @staticmethod
    def build_rule_input(form_data: dict) -> dict:
        """Parse and validate raw form data into rule kwargs.

        Raises ``ValueError`` on invalid input.
        """
        from app.models.finance.banking.transaction_rule import RuleAction, RuleType

        rule_name = (form_data.get("rule_name") or "").strip()
        if not rule_name:
            raise ValueError("Rule name is required")

        # Parse conditions JSON from the visual builder
        conditions_raw = form_data.get("conditions_json", "")
        if not conditions_raw:
            raise ValueError("At least one matching condition is required")

        try:
            conditions = json.loads(conditions_raw)
        except (json.JSONDecodeError, TypeError) as exc:
            raise ValueError(f"Invalid conditions: {exc}") from exc

        # Validate conditions structure
        if not isinstance(conditions, dict):
            raise ValueError("Conditions must be a JSON object")
        sub_rules = conditions.get("rules", [])
        if not sub_rules:
            raise ValueError("At least one matching condition is required")

        valid_sub_types = {rt.value for rt in RuleType if rt != RuleType.COMBINED}
        for sr in sub_rules:
            if sr.get("type") not in valid_sub_types:
                raise ValueError(f"Invalid condition type: {sr.get('type')}")

        # Parse action
        action_str = form_data.get("action", "CATEGORIZE")
        try:
            action = RuleAction(action_str)
        except ValueError:
            raise ValueError(f"Invalid action: {action_str}")

        # Parse optional UUIDs
        target_account_id: UUID | None = None
        if action == RuleAction.CATEGORIZE:
            raw = form_data.get("target_account_id", "")
            if raw:
                target_account_id = UUID(raw)

        tax_code_id: UUID | None = None
        raw_tax = form_data.get("tax_code_id", "")
        if raw_tax:
            tax_code_id = UUID(raw_tax)

        bank_account_id: UUID | None = None
        raw_ba = form_data.get("bank_account_id", "")
        if raw_ba:
            bank_account_id = UUID(raw_ba)

        return {
            "rule_name": rule_name,
            "description": (form_data.get("description") or "").strip() or None,
            "rule_type": RuleType.COMBINED,
            "conditions": conditions,
            "action": action,
            "target_account_id": target_account_id,
            "tax_code_id": tax_code_id,
            "bank_account_id": bank_account_id,
            "auto_apply": form_data.get("auto_apply") == "on",
            "min_confidence": int(form_data.get("min_confidence", 80)),
            "applies_to_credits": form_data.get("applies_to_credits") == "on",
            "applies_to_debits": form_data.get("applies_to_debits") == "on",
            "is_active": form_data.get("is_active") == "on",
        }

    def create_rule_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        form_data: dict,
    ) -> HTMLResponse | RedirectResponse:
        """Handle POST for new rule creation."""
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        try:
            kwargs = self.build_rule_input(form_data)
        except (ValueError, TypeError) as exc:
            context = base_context(
                request, auth, "New Transaction Rule", "banking", db=db
            )
            context.update(self.rule_form_context(db, str(auth.organization_id)))
            context["error"] = str(exc)
            context["form_data"] = form_data
            return templates.TemplateResponse(
                request, "finance/banking/rule_form.html", context
            )

        # is_active isn't a create_rule() param; pop and set after creation
        is_active = kwargs.pop("is_active", True)

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        service = TransactionCategorizationService()
        rule = service.create_rule(
            db,
            organization_id=org_id,
            created_by=auth.person_id,
            **kwargs,
        )
        rule.is_active = is_active
        db.flush()
        return RedirectResponse(
            url="/finance/banking/rules?success=Rule+created",
            status_code=303,
        )

    def update_rule_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        rule_id: str,
        form_data: dict,
    ) -> HTMLResponse | RedirectResponse:
        """Handle POST for rule update."""
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        try:
            kwargs = self.build_rule_input(form_data)
        except (ValueError, TypeError) as exc:
            context = base_context(
                request, auth, "Edit Transaction Rule", "banking", db=db
            )
            context.update(
                self.rule_form_context(db, str(auth.organization_id), rule_id=rule_id)
            )
            context["error"] = str(exc)
            context["form_data"] = form_data
            return templates.TemplateResponse(
                request, "finance/banking/rule_form.html", context
            )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        service = TransactionCategorizationService()
        rule = service.update_rule(
            db,
            organization_id=org_id,
            rule_id=UUID(rule_id),
            **kwargs,
        )
        if not rule:
            context = base_context(
                request, auth, "Edit Transaction Rule", "banking", db=db
            )
            context.update(
                self.rule_form_context(db, str(auth.organization_id), rule_id=rule_id)
            )
            context["error"] = "Rule not found"
            return templates.TemplateResponse(
                request, "finance/banking/rule_form.html", context
            )

        db.flush()
        return RedirectResponse(
            url="/finance/banking/rules?success=Rule+updated",
            status_code=303,
        )

    def list_accounts_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        search: str | None,
        status: str | None,
        page: int,
        limit: int = 50,
        sort: str | None = None,
        sort_dir: str | None = None,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Bank Accounts", "banking", db=db)
        context.update(
            self.list_accounts_context(
                db,
                str(auth.organization_id),
                search=search,
                status=status,
                page=page,
                limit=limit,
                sort=sort,
                sort_dir=sort_dir,
            )
        )
        return templates.TemplateResponse(
            request, "finance/banking/accounts.html", context
        )

    def account_new_form_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
    ) -> HTMLResponse:
        context = base_context(request, auth, "New Bank Account", "banking", db=db)
        context.update(self.account_form_context(db, str(auth.organization_id)))
        return templates.TemplateResponse(
            request, "finance/banking/account_form.html", context
        )

    def account_detail_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        account_id: str,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Bank Account Details", "banking", db=db)
        context.update(
            self.account_detail_context(
                db,
                str(auth.organization_id),
                account_id,
            )
        )
        return templates.TemplateResponse(
            request, "finance/banking/account_detail.html", context
        )

    def transaction_detail_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        line_id: str,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Transaction Detail", "banking", db=db)
        context.update(
            self.transaction_detail_context(
                db,
                str(auth.organization_id),
                line_id,
            )
        )
        return templates.TemplateResponse(
            request, "finance/banking/transaction_detail.html", context
        )

    def account_edit_form_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        account_id: str,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Edit Bank Account", "banking", db=db)
        context.update(
            self.account_form_context(
                db,
                str(auth.organization_id),
                account_id,
            )
        )
        return templates.TemplateResponse(
            request, "finance/banking/account_form.html", context
        )

    def list_statements_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        account_id: str | None,
        status: str | None,
        start_date: str | None,
        end_date: str | None,
        page: int,
        limit: int = 50,
        sort: str | None = None,
        sort_dir: str | None = None,
        match_status: str | None = None,
        search: str | None = None,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Bank Transactions", "banking", db=db)
        context.update(
            self.list_statements_context(
                db,
                str(auth.organization_id),
                account_id=account_id,
                status=status,
                start_date=start_date,
                end_date=end_date,
                page=page,
                limit=limit,
                sort=sort,
                sort_dir=sort_dir,
                match_status=match_status,
                search=search,
            )
        )
        return templates.TemplateResponse(
            request, "finance/banking/statements.html", context
        )

    @staticmethod
    def list_statement_imports_context(
        db: Session,
        organization_id: str,
        account_id: str | None,
        status: str | None,
        start_date: str | None,
        end_date: str | None,
        page: int,
        limit: int = 25,
        sort: str | None = None,
        sort_dir: str | None = None,
        search: str | None = None,
    ) -> dict:
        """Build context for statement imports list (header-level view).

        Shows imported BankStatement headers grouped by bank account,
        each linking to the per-statement detail/matching page.
        """
        org_id = coerce_uuid(organization_id)
        offset = (page - 1) * limit

        from_date = _parse_date(start_date)
        to_date = _parse_date(end_date)

        # ── Bank accounts for filter dropdown ──
        accounts = list(
            db.scalars(
                select(BankAccount)
                .where(
                    BankAccount.organization_id == org_id,
                    BankAccount.status == BankAccountStatus.active,
                )
                .order_by(BankAccount.bank_name, BankAccount.account_number)
            ).all()
        )

        # ── Statement headers query ──
        stmt = (
            select(BankStatement)
            .join(
                BankAccount,
                BankStatement.bank_account_id == BankAccount.bank_account_id,
            )
            .where(BankStatement.organization_id == org_id)
        )

        if account_id:
            stmt = stmt.where(BankStatement.bank_account_id == coerce_uuid(account_id))
        parsed_status = _parse_statement_status(status)
        if parsed_status:
            stmt = stmt.where(BankStatement.status == parsed_status)
        if from_date:
            stmt = stmt.where(BankStatement.statement_date >= from_date)
        if to_date:
            stmt = stmt.where(BankStatement.statement_date <= to_date)
        if search:
            stmt = stmt.where(BankStatement.statement_number.ilike(f"%{search}%"))

        # ── Sorting ──
        sort_col = sort or "statement_date"
        col_map: dict[str, Any] = {
            "statement_date": BankStatement.statement_date,
            "statement_number": BankStatement.statement_number,
            "total_lines": BankStatement.total_lines,
        }
        order_col = col_map.get(sort_col, BankStatement.statement_date)
        if (sort_dir or "desc").lower() == "asc":
            stmt = stmt.order_by(order_col.asc())
        else:
            stmt = stmt.order_by(order_col.desc())

        # ── Count + paginate ──
        count_stmt = select(func.count()).select_from(stmt.subquery())
        total_count: int = db.scalar(count_stmt) or 0
        total_pages = max(1, (total_count + limit - 1) // limit)

        statements = list(db.scalars(stmt.offset(offset).limit(limit)).all())

        # ── Real matched counts from BankStatementLine ──
        # The BankStatement.matched_lines header column is not reliably updated,
        # so we compute actual counts from the line-level is_matched flag.
        stmt_ids = [s.statement_id for s in statements]
        line_counts: dict[UUID, dict[str, int]] = {}
        if stmt_ids:
            count_rows = db.execute(
                select(
                    BankStatementLine.statement_id,
                    func.count().label("total"),
                    func.count()
                    .where(BankStatementLine.is_matched.is_(True))
                    .label("matched"),
                )
                .where(BankStatementLine.statement_id.in_(stmt_ids))
                .group_by(BankStatementLine.statement_id)
            ).all()
            for row in count_rows:
                line_counts[row.statement_id] = {
                    "total": row.total,
                    "matched": row.matched,
                }

        total_lines = sum(lc["total"] for lc in line_counts.values())
        matched_lines = sum(lc["matched"] for lc in line_counts.values())

        # ── Active filters ──
        active_filters: list[dict[str, str]] = []
        base_params = ""
        if account_id:
            acct = next(
                (a for a in accounts if str(a.bank_account_id) == account_id),
                None,
            )
            label = (
                f"{acct.bank_name} - {acct.account_number}"
                if acct
                else "Selected Account"
            )
            active_filters.append({"label": label, "param": "account_id"})
            base_params += f"&account_id={account_id}"
        if status:
            active_filters.append(
                {
                    "label": f"Status: {status.replace('_', ' ').title()}",
                    "param": "status",
                }
            )
        if search:
            active_filters.append({"label": f'Search: "{search}"', "param": "search"})

        # ── Build statement views with real matched counts ──
        statement_views = []
        for s in statements:
            sv = _statement_view(s)
            lc = line_counts.get(s.statement_id, {"total": 0, "matched": 0})
            sv["total_lines"] = lc["total"]
            sv["matched_lines"] = lc["matched"]
            sv["unmatched_lines"] = lc["total"] - lc["matched"]
            statement_views.append(sv)

        return {
            "statements": statement_views,
            "accounts": [_account_view(a) for a in accounts],
            "account_id": account_id or "",
            "status": status or "",
            "search": search or "",
            "start_date": start_date or "",
            "end_date": end_date or "",
            "sort": sort_col,
            "sort_dir": sort_dir or "desc",
            "page": page,
            "limit": limit,
            "offset": offset,
            "total_count": total_count,
            "total_pages": total_pages,
            "total_lines": total_lines,
            "matched_lines": matched_lines,
            "unmatched_lines": total_lines - matched_lines,
            "active_filters": active_filters,
        }

    def list_statement_imports_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        account_id: str | None,
        status: str | None,
        start_date: str | None,
        end_date: str | None,
        page: int,
        limit: int = 25,
        sort: str | None = None,
        sort_dir: str | None = None,
        search: str | None = None,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Imported Statements", "banking", db=db)
        context.update(
            self.list_statement_imports_context(
                db,
                str(auth.organization_id),
                account_id=account_id,
                status=status,
                start_date=start_date,
                end_date=end_date,
                page=page,
                limit=limit,
                sort=sort,
                sort_dir=sort_dir,
                search=search,
            )
        )
        return templates.TemplateResponse(
            request, "finance/banking/statement_imports.html", context
        )

    def statement_import_form_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        form_data: dict | None = None,
        errors: list[str] | None = None,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Import Bank Statement", "banking", db=db)
        context.update(self.statement_import_context(db, str(auth.organization_id)))
        form_payload = form_data or {}
        if not form_payload:
            if request.query_params.get("account_id"):
                form_payload["bank_account_id"] = request.query_params.get("account_id")
            form_payload.setdefault("statement_date", date.today().isoformat())
        context["form_data"] = form_payload
        context["form_errors"] = errors or []
        return templates.TemplateResponse(
            request, "finance/banking/statement_import.html", context
        )

    async def statement_import_submit_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
    ):
        form = getattr(request.state, "csrf_form", None)
        if form is None:
            form = await request.form()
        form_data = dict(form)
        csv_format = form_data.get("csv_format") or None
        errors: list[str] = []
        org_date_fmt = self._resolve_org_date_format(db, auth.organization_id)

        # Parse lines from file or manual entry.
        lines_data: list[dict] = []
        upload = form.get("statement_file")
        upload_file = upload if isinstance(upload, UploadFile) else None
        has_upload = bool(upload_file and upload_file.filename)
        upload_ext = ""
        column_map = self._parse_column_map(form)
        mapped_lines_data, manual_errors = self._parse_manual_lines(form)
        # When a file is uploaded, ignore manual line fields from the same form
        # submission to avoid stale hidden inputs overriding uploaded data.
        if has_upload:
            mapped_lines_data = []
            manual_errors = []
        if mapped_lines_data:
            lines_data = self._normalize_mapped_lines(
                mapped_lines_data, org_date_fmt=org_date_fmt
            )
            errors.extend(manual_errors)
        if has_upload:
            # CSRF middleware parses form data first, which can advance the file pointer.
            try:
                await upload_file.seek(0)
            except Exception:
                try:
                    upload_file.file.seek(0)
                except Exception:
                    logger.exception("Ignored exception")
            filename = upload_file.filename or ""
            lowered = filename.lower()
            upload_ext = (
                ".csv"
                if lowered.endswith(".csv")
                else ".xls"
                if lowered.endswith(".xls")
                else ".xlsx"
                if lowered.endswith(".xlsx")
                else ".xlsm"
                if lowered.endswith(".xlsm")
                else ""
            )
            if not lowered.endswith(SPREADSHEET_EXTENSIONS):
                errors.append(
                    f"Supported statement files: {spreadsheet_formats_label()}."
                )
            else:
                # Limit upload size to avoid memory blowups.
                max_bytes = 10 * 1024 * 1024  # 10 MiB
                content = await upload_file.read(max_bytes + 1)
                if len(content) > max_bytes:
                    errors.append(
                        "Uploaded file is too large (max 10 MB). Please upload a smaller file."
                    )
                    content = b""
                if not content:
                    # Some middleware (e.g., CSRF) may have consumed the file stream.
                    # Fall back to reading from the underlying file handle.
                    try:
                        upload_file.file.seek(0)
                        content = upload_file.file.read(max_bytes + 1)
                        if len(content) > max_bytes:
                            errors.append(
                                "Uploaded file is too large (max 10 MB). Please upload a smaller file."
                            )
                            content = b""
                    except Exception:
                        content = content or b""
                if content:
                    # Avoid logging file contents (may contain PII).
                    logger.info(
                        "Statement import file read: filename=%s bytes=%s content_type=%s",
                        upload_file.filename,
                        len(content),
                        upload_file.content_type,
                    )
                if not content:
                    logger.warning(
                        "Statement import upload empty after read: filename=%s content_type=%s",
                        upload_file.filename,
                        upload_file.content_type,
                    )
                    errors.append(
                        "Uploaded file appears empty. Please re-select the file and try again."
                    )
                else:
                    if column_map:
                        try:
                            _, source_rows, _total_rows = self._preview_upload_content(
                                content, lowered, sample_limit=None
                            )
                        except ValueError as exc:
                            errors.append(str(exc))
                            source_rows = []
                        mapped_rows = self._map_rows_with_column_map(
                            source_rows, column_map
                        )
                        lines_data = self._normalize_mapped_lines(
                            mapped_rows, org_date_fmt=org_date_fmt
                        )
                        if not lines_data:
                            errors.append(
                                "No rows found after applying column mapping. Please review your selected columns."
                            )
                    elif lowered.endswith(".csv"):
                        rows, parse_errors = bank_statement_service.parse_csv_rows(
                            content, csv_format, date_format=org_date_fmt
                        )
                        lines_data = rows
                        errors.extend(parse_errors)
                    elif lowered.endswith(".xls"):
                        rows, parse_errors = bank_statement_service.parse_xls_rows(
                            content, csv_format, date_format=org_date_fmt
                        )
                        lines_data = rows
                        errors.extend(parse_errors)
                    else:
                        rows, parse_errors = bank_statement_service.parse_xlsx_rows(
                            content, csv_format, date_format=org_date_fmt
                        )
                        lines_data = rows
                        errors.extend(parse_errors)
                    if not lines_data and not errors:
                        logger.warning(
                            "Statement import parsed zero rows: filename=%s csv_format=%s",
                            upload_file.filename,
                            csv_format,
                        )
        elif not lines_data:
            lines_data = mapped_lines_data
            errors.extend(manual_errors)

        if not lines_data and not errors:
            errors.append(
                "Please upload a CSV/Excel file or add at least one transaction."
            )

        payload_data = {
            "bank_account_id": form_data.get("bank_account_id"),
            "statement_number": form_data.get("statement_number") or None,
            "statement_date": form_data.get("statement_date") or None,
            "period_start": form_data.get("period_start"),
            "period_end": form_data.get("period_end"),
            "opening_balance": form_data.get("opening_balance") or None,
            "closing_balance": form_data.get("closing_balance") or None,
            "import_source": (
                "csv"
                if upload_ext == ".csv"
                else "excel"
                if upload_ext in {".xls", ".xlsx", ".xlsm"}
                else "manual"
            ),
            "import_filename": upload_file.filename if upload_file else None,
            "lines": lines_data,
        }

        payload = None
        if not errors:
            try:
                payload = BankStatementImport.model_validate(payload_data)
            except ValidationError as exc:
                errors.extend(self._format_validation_errors(exc))

        if errors or payload is None:
            # Preserve select fields for the form.
            preserved = {
                "bank_account_id": form_data.get("bank_account_id"),
                "statement_number": form_data.get("statement_number"),
                "statement_date": form_data.get("statement_date"),
                "period_start": form_data.get("period_start"),
                "period_end": form_data.get("period_end"),
                "opening_balance": form_data.get("opening_balance"),
                "closing_balance": form_data.get("closing_balance"),
                "csv_format": csv_format,
            }
            return self.statement_import_form_response(
                request, auth, db, form_data=preserved, errors=errors
            )

        line_inputs, line_errors = bank_statement_service.build_line_inputs(
            payload.lines
        )
        if line_errors:
            preserved = {
                "bank_account_id": form_data.get("bank_account_id"),
                "statement_number": form_data.get("statement_number"),
                "statement_date": form_data.get("statement_date"),
                "period_start": form_data.get("period_start"),
                "period_end": form_data.get("period_end"),
                "opening_balance": form_data.get("opening_balance"),
                "closing_balance": form_data.get("closing_balance"),
                "csv_format": csv_format,
            }
            return self.statement_import_form_response(
                request, auth, db, form_data=preserved, errors=line_errors
            )

        if auth.organization_id is None:
            raise HTTPException(status_code=400, detail="Organization is required")

        result = bank_statement_service.import_statement(
            db=db,
            organization_id=auth.organization_id,
            bank_account_id=payload.bank_account_id,
            statement_number=payload.statement_number,
            statement_date=payload.statement_date,
            period_start=payload.period_start,
            period_end=payload.period_end,
            opening_balance=payload.opening_balance,
            closing_balance=payload.closing_balance,
            lines=line_inputs,
            import_source=payload.import_source,
            import_filename=payload.import_filename,
            imported_by=auth.user_id,
        )
        db.flush()
        redirect_url = (
            f"/finance/banking/statements/{result.statement.statement_id}"
            f"?success=Statement+imported+successfully"
            f"+({result.lines_imported}+lines)"
        )
        if result.auto_matched > 0:
            redirect_url += f"&auto_matched={result.auto_matched}"
        return RedirectResponse(url=redirect_url, status_code=303)

    async def statement_import_preview_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
    ) -> JSONResponse:
        form = getattr(request.state, "csrf_form", None)
        if form is None:
            form = await request.form()

        upload = form.get("statement_file")
        upload_file = upload if isinstance(upload, UploadFile) else None
        if not upload_file or not upload_file.filename:
            return JSONResponse(
                status_code=400,
                content={"detail": "Please choose a file to preview."},
            )

        filename = upload_file.filename or ""
        lowered = filename.lower()
        if not lowered.endswith(SPREADSHEET_EXTENSIONS):
            return JSONResponse(
                status_code=400,
                content={
                    "detail": (
                        f"Supported statement files: {spreadsheet_formats_label()}."
                    )
                },
            )

        try:
            await upload_file.seek(0)
        except Exception:
            try:
                upload_file.file.seek(0)
            except Exception:
                logger.exception("Ignored exception")

        max_bytes = 10 * 1024 * 1024  # 10 MiB
        content = await upload_file.read(max_bytes + 1)
        if len(content) > max_bytes:
            return JSONResponse(
                status_code=400,
                content={
                    "detail": "Uploaded file is too large (max 10 MB). Please upload a smaller file."
                },
            )
        if not content:
            try:
                upload_file.file.seek(0)
                content = upload_file.file.read(max_bytes + 1)
            except Exception:
                content = b""
        if not content:
            return JSONResponse(
                status_code=400,
                content={"detail": "Uploaded file appears empty."},
            )

        try:
            headers, sample_rows, total_rows = self._preview_upload_content(
                content, lowered
            )
        except ValueError as exc:
            return JSONResponse(status_code=400, content={"detail": str(exc)})

        return JSONResponse(
            {
                "detected_columns": headers,
                "sample_data": sample_rows,
                "total_rows": total_rows,
            }
        )

    @staticmethod
    def _preview_upload_content(
        content: bytes,
        lowered_filename: str,
        sample_limit: int | None = 5,
    ) -> tuple[list[str], list[dict[str, str]], int]:
        if lowered_filename.endswith(".csv"):
            try:
                text = content.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = content.decode("utf-8", errors="replace")

            delimiter = ","
            try:
                header_line = next(
                    (line for line in text.splitlines() if line.strip()), ""
                )
                candidates = [",", "\t", ";", "|"]
                counts = {d: header_line.count(d) for d in candidates}
                best = max(counts, key=counts.__getitem__)
                if counts.get(best, 0) > 0:
                    delimiter = best
                else:
                    sniff = csv.Sniffer().sniff(text[:2048])
                    if sniff and getattr(sniff, "delimiter", None):
                        delimiter = sniff.delimiter
            except Exception:
                delimiter = ","

            reader = csv.DictReader(StringIO(text), delimiter=delimiter)
            if not reader.fieldnames:
                raise ValueError("CSV file must include a header row.")
            headers = [str(h).strip() for h in reader.fieldnames if h is not None]
            # Keep DictReader row keys aligned with the trimmed headers shown in UI.
            reader.fieldnames = headers
            rows: list[dict[str, str]] = []
            total_rows = 0
            for row in reader:
                if not any(str(v).strip() for v in row.values() if v is not None):
                    continue
                total_rows += 1
                sample = {
                    h: ("" if row.get(h) is None else str(row.get(h))) for h in headers
                }
                if sample_limit is None or len(rows) < sample_limit:
                    rows.append(sample)
            return headers, rows, total_rows

        if lowered_filename.endswith(".xlsx") or lowered_filename.endswith(".xlsm"):
            from openpyxl import load_workbook

            workbook = load_workbook(
                filename=BytesIO(content), read_only=True, data_only=True
            )
            try:
                sheet = workbook.active
                rows_iter = sheet.iter_rows(values_only=True)
                try:
                    header_values = next(rows_iter)
                except StopIteration:
                    raise ValueError("Excel file must include a header row.") from None
                if not header_values:
                    raise ValueError("Excel file must include a header row.")
                headers = [
                    str(h).strip() if h is not None else "" for h in header_values
                ]
                headers = [h for h in headers if h]
                if not headers:
                    raise ValueError("Excel file must include a header row.")
                xlsx_rows: list[dict[str, str]] = []
                total_rows = 0
                for values in rows_iter:
                    if not values or not any(
                        value is not None and str(value).strip() for value in values
                    ):
                        continue
                    total_rows += 1
                    xlsx_row: dict[str, str] = {}
                    for i, header in enumerate(headers):
                        value = values[i] if i < len(values) else ""
                        if isinstance(value, _datetime):
                            xlsx_row[header] = value.strftime("%Y-%m-%d")
                        elif isinstance(value, date):
                            xlsx_row[header] = value.isoformat()
                        elif value is None:
                            xlsx_row[header] = ""
                        else:
                            xlsx_row[header] = str(value)
                    if sample_limit is None or len(xlsx_rows) < sample_limit:
                        xlsx_rows.append(xlsx_row)
                return headers, xlsx_rows, total_rows
            finally:
                workbook.close()

        if lowered_filename.endswith(".xls"):
            try:
                import xlrd
            except builtins.ImportError as exc:
                raise ValueError(
                    "XLS preview requires xlrd. Please install xlrd and retry."
                ) from exc

            try:
                workbook = xlrd.open_workbook(file_contents=content)
            except Exception as exc:
                raise ValueError(
                    "Could not parse XLS file. Please upload a valid .xls file."
                ) from exc

            if workbook.nsheets == 0:
                raise ValueError("Excel file must include a header row.")
            sheet = workbook.sheet_by_index(0)
            if sheet.nrows < 1:
                raise ValueError("Excel file must include a header row.")
            header_values = sheet.row_values(0)
            headers = [str(h).strip() if h is not None else "" for h in header_values]
            headers = [h for h in headers if h]
            if not headers:
                raise ValueError("Excel file must include a header row.")
            xls_rows: list[dict[str, str]] = []
            total_rows = 0
            for row_index in range(1, sheet.nrows):
                row_cells = sheet.row(row_index)
                if not any(
                    cell.value is not None and str(cell.value).strip()
                    for cell in row_cells
                ):
                    continue
                total_rows += 1
                xls_row: dict[str, str] = {}
                for i, header in enumerate(headers):
                    cell = row_cells[i] if i < len(row_cells) else None
                    if cell is None or cell.value is None:
                        xls_row[header] = ""
                    elif cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            dt = xlrd.xldate_as_datetime(cell.value, workbook.datemode)
                            xls_row[header] = dt.strftime("%Y-%m-%d")
                        except Exception:
                            xls_row[header] = str(cell.value)
                    else:
                        xls_row[header] = str(cell.value)
                if sample_limit is None or len(xls_rows) < sample_limit:
                    xls_rows.append(xls_row)
            return headers, xls_rows, total_rows

        raise ValueError(f"Supported statement files: {spreadsheet_formats_label()}.")

    @staticmethod
    def _parse_column_map(form) -> dict[str, str]:
        pattern = re.compile(r"^column_map\[(.+)\]$")
        mapping: dict[str, str] = {}
        for key, value in form.items():
            match = pattern.match(key)
            if not match:
                continue
            source_col = match.group(1)
            target = str(value).strip() if value is not None else ""
            if not target:
                continue
            mapping[source_col] = target
        return mapping

    @staticmethod
    def _map_rows_with_column_map(
        source_rows: list[dict[str, str]],
        column_map: dict[str, str],
    ) -> list[dict]:
        mapped_rows: list[dict] = []
        for idx, source in enumerate(source_rows, start=1):
            row: dict[str, Any] = {"line_number": idx}
            for source_col, target_field in column_map.items():
                value = source.get(source_col, "")
                row[target_field] = "" if value is None else str(value)
            if any(
                str(v).strip()
                for k, v in row.items()
                if k != "line_number" and v is not None
            ):
                mapped_rows.append(row)
        return mapped_rows

    @staticmethod
    def _parse_manual_lines(form) -> tuple[list[dict], list[str]]:
        pattern = re.compile(r"^lines\[(\d+)\]\[(.+)\]$")
        lines: dict[int, dict] = {}
        errors: list[str] = []

        for key, value in form.items():
            match = pattern.match(key)
            if not match:
                continue
            line_index = int(match.group(1))
            field = match.group(2)
            lines.setdefault(line_index, {})[field] = value

        if not lines:
            return [], []

        results: list[dict] = []
        for idx in sorted(lines):
            data = lines[idx]
            # Skip rows that are entirely empty.
            if not any(str(v).strip() for v in data.values() if v is not None):
                continue

            line_number = data.get("line_number") or idx
            result = {
                "line_number": int(line_number),
                "transaction_date": data.get("transaction_date"),
                "transaction_type": data.get("transaction_type"),
                "amount": data.get("amount"),
                "debit": data.get("debit"),
                "credit": data.get("credit"),
                "description": data.get("description"),
                "reference": data.get("reference"),
                "payee_payer": data.get("payee_payer"),
                "bank_reference": data.get("bank_reference"),
                "check_number": data.get("check_number"),
                "bank_category": data.get("bank_category"),
                "bank_code": data.get("bank_code"),
                "value_date": data.get("value_date"),
                "running_balance": data.get("running_balance"),
                "transaction_id": data.get("transaction_id"),
            }
            results.append(result)

        if not results:
            errors.append("Please add at least one transaction line.")
        return results, errors

    @staticmethod
    def _normalize_mapped_lines(
        lines: list[dict],
        *,
        org_date_fmt: str | None = None,
    ) -> list[dict]:
        normalized: list[dict] = []
        date_fields = ("transaction_date", "value_date")
        decimal_fields = ("amount", "debit", "credit", "running_balance")
        for line in lines:
            row = dict(line)
            tx_type = row.get("transaction_type")
            if tx_type is not None:
                cleaned = str(tx_type).strip().lower()
                row["transaction_type"] = cleaned or None

            for field in date_fields:
                raw = row.get(field)
                if raw is None:
                    continue
                cleaned = str(raw).strip()
                if not cleaned:
                    row[field] = None
                    continue
                parsed_date = _parse_date(cleaned, format=org_date_fmt)
                row[field] = parsed_date if parsed_date is not None else cleaned

            for field in decimal_fields:
                raw = row.get(field)
                if raw is None:
                    continue
                cleaned = str(raw).strip()
                if not cleaned:
                    row[field] = None
                    continue
                parsed_decimal = _parse_decimal(cleaned)
                row[field] = parsed_decimal if parsed_decimal is not None else cleaned

            normalized.append(row)
        return normalized

    @staticmethod
    def _resolve_org_date_format(
        db: Session, organization_id: UUID | None
    ) -> str | None:
        if not organization_id:
            return None
        from app.models.finance.core_org.organization import Organization
        from app.services.formatting_context import DATE_FORMAT_MAP

        org = db.get(Organization, organization_id)
        if not org or not getattr(org, "date_format", None):
            return None
        date_format_key = org.date_format
        if not isinstance(date_format_key, str):
            return None
        date_format = DATE_FORMAT_MAP.get(date_format_key)
        return str(date_format) if date_format is not None else None

    @staticmethod
    def _format_validation_errors(exc: ValidationError) -> list[str]:
        errors: list[str] = []
        for err in exc.errors():
            loc = " -> ".join(str(item) for item in err.get("loc", []))
            msg = err.get("msg", "Invalid value")
            if loc:
                errors.append(f"{loc}: {msg}")
            else:
                errors.append(msg)
        return errors

    def statement_detail_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
        *,
        page: int = 1,
        limit: int = 50,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Bank Statement", "banking", db=db)
        context.update(
            self.statement_detail_context(
                db,
                str(auth.organization_id),
                statement_id,
                page=page,
                limit=limit,
            )
        )
        return templates.TemplateResponse(
            request, "finance/banking/statement_detail.html", context
        )

    def apply_rules_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
    ) -> RedirectResponse:
        """Handle POST to apply categorization rules to a statement."""
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        service = TransactionCategorizationService()
        result = service.apply_rules_to_statement(db, org_id, coerce_uuid(statement_id))
        db.flush()

        msg = (
            f"{result.categorized_count}+suggested,"
            f"{result.high_confidence_count}+auto-applied,"
            f"{result.no_match_count}+no+match"
        )
        return RedirectResponse(
            url=f"/finance/banking/statements/{statement_id}?success={msg}",
            status_code=303,
        )

    def accept_suggestion_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
        line_id: str,
    ) -> RedirectResponse:
        """Handle POST to accept a categorization suggestion."""
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        service = TransactionCategorizationService()
        try:
            service.accept_suggestion(
                db, org_id, coerce_uuid(line_id), accepted_by=auth.person_id
            )
            db.flush()
        except ValueError as exc:
            logger.warning("Accept suggestion failed: %s", exc)
            return RedirectResponse(
                url=f"/finance/banking/statements/{statement_id}?error={exc}",
                status_code=303,
            )

        return RedirectResponse(
            url=f"/finance/banking/statements/{statement_id}?success=Suggestion+accepted",
            status_code=303,
        )

    def reject_suggestion_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
        line_id: str,
    ) -> RedirectResponse:
        """Handle POST to reject a categorization suggestion."""
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        service = TransactionCategorizationService()
        try:
            service.reject_suggestion(db, org_id, coerce_uuid(line_id))
            db.flush()
        except ValueError as exc:
            logger.warning("Reject suggestion failed: %s", exc)
            return RedirectResponse(
                url=f"/finance/banking/statements/{statement_id}?error={exc}",
                status_code=303,
            )

        return RedirectResponse(
            url=f"/finance/banking/statements/{statement_id}?success=Suggestion+rejected",
            status_code=303,
        )

    # ------------------------------------------------------------------
    # Flat-view response methods (operate on lines across all statements)
    # ------------------------------------------------------------------

    def apply_rules_flat_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        account_id: str,
    ) -> RedirectResponse:
        """Apply categorization rules to all unprocessed lines for a bank account."""
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        service = TransactionCategorizationService()
        result = service.apply_rules_to_account(db, org_id, coerce_uuid(account_id))
        db.flush()

        msg = (
            f"{result.categorized_count}+suggested,"
            f"{result.high_confidence_count}+auto-applied,"
            f"{result.no_match_count}+no+match"
        )
        return RedirectResponse(
            url=f"/finance/banking/statements?account_id={account_id}&success={msg}",
            status_code=303,
        )

    def accept_suggestion_flat_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        line_id: str,
    ) -> RedirectResponse:
        """Accept a categorization suggestion from the flat lines view."""
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        # Resolve the line's bank account for redirect
        line = (
            db.execute(
                select(BankStatementLine)
                .join(BankStatement)
                .where(
                    BankStatementLine.line_id == coerce_uuid(line_id),
                    BankStatement.organization_id == org_id,
                )
            )
            .scalars()
            .first()
        )
        account_id = ""
        if line:
            stmt = db.get(BankStatement, line.statement_id)
            if stmt:
                account_id = str(stmt.bank_account_id)

        service = TransactionCategorizationService()
        try:
            service.accept_suggestion(
                db, org_id, coerce_uuid(line_id), accepted_by=auth.person_id
            )
            db.flush()
        except ValueError as exc:
            logger.warning("Accept suggestion (flat) failed: %s", exc)
            return RedirectResponse(
                url=f"/finance/banking/statements?account_id={account_id}&error={exc}",
                status_code=303,
            )

        return RedirectResponse(
            url=(
                f"/finance/banking/statements?account_id={account_id}"
                "&success=Suggestion+accepted"
            ),
            status_code=303,
        )

    def reject_suggestion_flat_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        line_id: str,
    ) -> RedirectResponse:
        """Reject a categorization suggestion from the flat lines view."""
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        # Resolve the line's bank account for redirect
        line = (
            db.execute(
                select(BankStatementLine)
                .join(BankStatement)
                .where(
                    BankStatementLine.line_id == coerce_uuid(line_id),
                    BankStatement.organization_id == org_id,
                )
            )
            .scalars()
            .first()
        )
        account_id = ""
        if line:
            stmt = db.get(BankStatement, line.statement_id)
            if stmt:
                account_id = str(stmt.bank_account_id)

        service = TransactionCategorizationService()
        try:
            service.reject_suggestion(db, org_id, coerce_uuid(line_id))
            db.flush()
        except ValueError as exc:
            logger.warning("Reject suggestion (flat) failed: %s", exc)
            return RedirectResponse(
                url=f"/finance/banking/statements?account_id={account_id}&error={exc}",
                status_code=303,
            )

        return RedirectResponse(
            url=(
                f"/finance/banking/statements?account_id={account_id}"
                "&success=Suggestion+rejected"
            ),
            status_code=303,
        )

    def delete_statement_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
    ) -> RedirectResponse:
        """Handle POST to delete a bank statement batch."""
        try:
            deleted = bank_statement_service.delete(
                db,
                coerce_uuid(auth.organization_id),
                coerce_uuid(statement_id),
            )
            if not deleted:
                return RedirectResponse(
                    url=f"/finance/banking/statements/{statement_id}?error=Statement+not+found",
                    status_code=303,
                )
            db.flush()
        except ValueError as exc:
            logger.warning("Delete statement failed: %s", exc)
            return RedirectResponse(
                url=f"/finance/banking/statements/{statement_id}?error={exc}",
                status_code=303,
            )

        return RedirectResponse(
            url="/finance/banking/statements?success=Statement+deleted",
            status_code=303,
        )

    def statement_auto_match_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
    ) -> RedirectResponse:
        """Run auto-match on a specific statement's unmatched lines.

        Delegates to ``AutoReconciliationService.auto_match_statement()``
        for deterministic PaymentIntent + Splynx payment matching.
        """
        from app.services.finance.banking.auto_reconciliation import (
            AutoReconciliationService,
        )

        org_id = coerce_uuid(auth.organization_id)
        stmt_id = coerce_uuid(statement_id)

        auto_svc = AutoReconciliationService()
        result = auto_svc.auto_match_statement(db, org_id, stmt_id)

        if result.matched > 0:
            db.flush()
            msg = f"Auto-matched+{result.matched}+lines"
            if result.skipped > 0:
                msg += f"+({result.skipped}+skipped)"
            return RedirectResponse(
                url=f"/finance/banking/statements/{statement_id}?success={msg}",
                status_code=303,
            )

        if result.errors:
            error_msg = "Auto-match+errors:+" + ",+".join(result.errors[:3])
            return RedirectResponse(
                url=f"/finance/banking/statements/{statement_id}?error={error_msg}",
                status_code=303,
            )

        return RedirectResponse(
            url=f"/finance/banking/statements/{statement_id}?info=No+new+matches+found",
            status_code=303,
        )

    async def match_statement_line_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
        line_id: str,
    ) -> Response:
        """Accept a GL transaction match for a statement line (JSON from Alpine.js)."""
        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        body = await request.json()
        journal_line_id = body.get("journal_line_id")
        force_match = bool(body.get("force_match", False))
        if not journal_line_id:
            return JSONResponse(
                content={"detail": "journal_line_id is required"}, status_code=400
            )

        svc = BankReconciliationService()
        user_id = getattr(auth, "user_id", None) or getattr(auth, "person_id", None)

        try:
            stmt_line = svc.match_statement_line(
                db=db,
                organization_id=org_id,
                statement_line_id=UUID(line_id),
                journal_line_id=UUID(str(journal_line_id)),
                matched_by=user_id,
                force_match=force_match,
            )
            db.flush()
        except HTTPException:
            raise
        except (ValueError, RuntimeError) as e:
            logger.warning("Statement line match failed: %s", e)
            return JSONResponse(content={"detail": str(e)}, status_code=400)

        # Return updated counters so frontend can update stats without reload
        statement = stmt_line.statement
        matched = statement.matched_lines or 0
        total = statement.total_lines or (
            (statement.matched_lines or 0) + (statement.unmatched_lines or 0)
        )
        match_pct = round(matched / total * 100) if total else 0

        # Resolve source URL + match detail for the matched GL line
        from app.services.finance.banking.bank_reconciliation import (
            _build_source_url,
        )

        source_url = ""
        match_detail: dict[str, str] | None = None
        gl_line = db.get(JournalEntryLine, UUID(str(journal_line_id)))
        if gl_line:
            entry = getattr(gl_line, "journal_entry", None) or getattr(
                gl_line, "entry", None
            )
            if entry:
                source_url = _build_source_url(
                    getattr(entry, "source_document_type", None),
                    getattr(entry, "source_document_id", None),
                    getattr(entry, "entry_id", None),
                )
                match_detail = _build_match_detail(db, entry, source_url)

        return JSONResponse(
            content={
                "status": "ok",
                "matched_lines": matched,
                "unmatched_lines": statement.unmatched_lines or 0,
                "match_pct": match_pct,
                "total_lines": total,
                "source_url": source_url,
                "match_detail": match_detail,
            },
            status_code=200,
        )

    async def batch_match_statement_lines_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
    ) -> Response:
        """Match multiple statement lines to GL entries in a single request."""
        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
            _build_source_url,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        body = await request.json()
        matches: list[dict[str, str]] = body.get("matches", [])
        if not matches:
            return JSONResponse(
                content={"detail": "matches array is required"}, status_code=400
            )

        svc = BankReconciliationService()
        user_id = getattr(auth, "user_id", None) or getattr(auth, "person_id", None)

        results: list[dict[str, object]] = []
        matched_count = 0
        error_count = 0

        for match in matches:
            line_id = match.get("line_id", "")
            journal_line_id = match.get("journal_line_id", "")
            force = bool(match.get("force_match", False))

            if not line_id or not journal_line_id:
                results.append(
                    {"line_id": line_id, "status": "error", "detail": "missing IDs"}
                )
                error_count += 1
                continue

            try:
                svc.match_statement_line(
                    db=db,
                    organization_id=org_id,
                    statement_line_id=UUID(line_id),
                    journal_line_id=UUID(str(journal_line_id)),
                    matched_by=user_id,
                    force_match=force,
                )

                # Resolve source URL + match detail
                source_url = ""
                batch_match_detail: dict[str, str] | None = None
                gl_line = db.get(JournalEntryLine, UUID(str(journal_line_id)))
                if gl_line:
                    entry = getattr(gl_line, "journal_entry", None) or getattr(
                        gl_line, "entry", None
                    )
                    if entry:
                        source_url = _build_source_url(
                            getattr(entry, "source_document_type", None),
                            getattr(entry, "source_document_id", None),
                            getattr(entry, "entry_id", None),
                        )
                        batch_match_detail = _build_match_detail(db, entry, source_url)

                results.append(
                    {
                        "line_id": line_id,
                        "status": "ok",
                        "source_url": source_url,
                        "match_detail": batch_match_detail,
                    }
                )
                matched_count += 1
            except HTTPException as e:
                results.append(
                    {"line_id": line_id, "status": "error", "detail": e.detail}
                )
                error_count += 1
            except (ValueError, RuntimeError) as e:
                logger.warning("Batch match failed for line %s: %s", line_id, e)
                results.append(
                    {"line_id": line_id, "status": "error", "detail": str(e)}
                )
                error_count += 1

        # Commit all successful matches in one transaction
        if matched_count > 0:
            db.flush()

        # Get final statement counters
        statement = db.get(BankStatement, UUID(statement_id))
        matched = (statement.matched_lines or 0) if statement else 0
        total = (statement.total_lines or 0) if statement else 0
        match_pct = round(matched / total * 100) if total else 0

        return JSONResponse(
            content={
                "status": "ok",
                "matched_count": matched_count,
                "error_count": error_count,
                "results": results,
                "matched_lines": matched,
                "unmatched_lines": (statement.unmatched_lines or 0) if statement else 0,
                "match_pct": match_pct,
                "total_lines": total,
            },
            status_code=200,
        )

    async def unmatch_statement_line_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
        line_id: str,
    ) -> Response:
        """Remove a direct match from a statement line (JSON from Alpine.js)."""
        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        svc = BankReconciliationService()

        try:
            stmt_line = svc.unmatch_statement_line(
                db=db,
                organization_id=org_id,
                statement_line_id=UUID(line_id),
            )
            db.flush()
        except HTTPException:
            raise
        except (ValueError, RuntimeError) as e:
            logger.warning("Statement line unmatch failed: %s", e)
            return JSONResponse(content={"detail": str(e)}, status_code=400)

        # Return updated counters
        statement = stmt_line.statement
        matched = statement.matched_lines or 0
        total = statement.total_lines or (
            (statement.matched_lines or 0) + (statement.unmatched_lines or 0)
        )
        match_pct = round(matched / total * 100) if total else 0

        return JSONResponse(
            content={
                "status": "ok",
                "matched_lines": matched,
                "unmatched_lines": statement.unmatched_lines or 0,
                "match_pct": match_pct,
                "total_lines": total,
            },
            status_code=200,
        )

    async def create_journal_and_match_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
        line_id: str,
    ) -> Response:
        """Create a GL journal and match it to a bank line (JSON from Alpine.js).

        Accepts: {counterparty_account_id: str, description?: str}
        """
        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
            _build_source_url,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        body = await request.json()
        counterparty_account_id = body.get("counterparty_account_id")
        description = body.get("description") or None
        if not counterparty_account_id:
            return JSONResponse(
                content={"detail": "counterparty_account_id is required"},
                status_code=400,
            )

        svc = BankReconciliationService()
        user_id = getattr(auth, "user_id", None) or getattr(auth, "person_id", None)

        try:
            stmt_line = svc.create_journal_and_match(
                db=db,
                organization_id=org_id,
                statement_line_id=UUID(line_id),
                counterparty_account_id=UUID(str(counterparty_account_id)),
                description=description,
                matched_by=user_id,
            )
            db.flush()
        except HTTPException:
            raise
        except (ValueError, RuntimeError) as e:
            logger.warning("Create journal & match failed: %s", e)
            return JSONResponse(content={"detail": str(e)}, status_code=400)

        # Return updated counters
        statement = stmt_line.statement
        matched = statement.matched_lines or 0
        total = statement.total_lines or (
            (statement.matched_lines or 0) + (statement.unmatched_lines or 0)
        )
        match_pct = round(matched / total * 100) if total else 0

        # Resolve source URL for the newly matched GL line
        source_url = ""
        match_detail: dict[str, str] | None = None
        gl_line = db.get(JournalEntryLine, stmt_line.matched_journal_line_id)
        if gl_line:
            entry = getattr(gl_line, "journal_entry", None) or getattr(
                gl_line, "entry", None
            )
            if entry:
                source_url = _build_source_url(
                    getattr(entry, "source_document_type", None),
                    getattr(entry, "source_document_id", None),
                    getattr(entry, "entry_id", None),
                )
                match_detail = _build_match_detail(db, entry, source_url)

        return JSONResponse(
            content={
                "status": "ok",
                "matched_lines": matched,
                "unmatched_lines": statement.unmatched_lines or 0,
                "match_pct": match_pct,
                "total_lines": total,
                "source_url": source_url,
                "match_detail": match_detail,
            },
            status_code=200,
        )

    def scored_candidates_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
        line_id: str,
        *,
        date_from: str | None = None,
        date_to: str | None = None,
        source_type: str | None = None,
        search: str | None = None,
        direction: str | None = None,
        hide_matched: bool = False,
        sort: str = "relevance",
        page: int = 1,
        per_page: int = 25,
    ) -> Response:
        """Return scored GL candidates for a specific statement line (JSON)."""
        from datetime import date as date_type

        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        # Parse date strings to date objects
        parsed_date_from: date_type | None = None
        parsed_date_to: date_type | None = None
        if date_from:
            try:
                parsed_date_from = date_type.fromisoformat(date_from)
            except ValueError:
                pass
        if date_to:
            try:
                parsed_date_to = date_type.fromisoformat(date_to)
            except ValueError:
                pass

        svc = BankReconciliationService()
        result = svc.get_scored_candidates_for_line(
            db=db,
            organization_id=org_id,
            statement_id=UUID(statement_id),
            statement_line_id=UUID(line_id),
            date_from=parsed_date_from,
            date_to=parsed_date_to,
            source_type=source_type or None,
            search=search or None,
            direction=direction or None,
            hide_matched=hide_matched,
            sort=sort,
            page=max(1, page),
            per_page=max(1, min(per_page, 100)),
        )

        return JSONResponse(content=result, status_code=200)

    async def multi_match_statement_line_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        statement_id: str,
        line_id: str,
    ) -> Response:
        """Match one bank line to multiple GL entries (JSON from Alpine.js)."""
        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
            _build_source_url,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        body = await request.json()
        journal_line_ids_raw = body.get("journal_line_ids", [])
        force_match = bool(body.get("force_match", False))

        if not journal_line_ids_raw:
            return JSONResponse(
                content={"detail": "journal_line_ids is required"},
                status_code=400,
            )

        journal_line_ids = [UUID(str(jl_id)) for jl_id in journal_line_ids_raw]

        svc = BankReconciliationService()
        user_id = getattr(auth, "user_id", None) or getattr(auth, "person_id", None)

        try:
            stmt_line = svc.multi_match_statement_line(
                db=db,
                organization_id=org_id,
                statement_line_id=UUID(line_id),
                journal_line_ids=journal_line_ids,
                matched_by=user_id,
                force_match=force_match,
            )
            db.flush()
        except HTTPException:
            raise
        except (ValueError, RuntimeError) as e:
            logger.warning("Multi-match failed: %s", e)
            return JSONResponse(content={"detail": str(e)}, status_code=400)

        # Return updated counters + source URLs for all matched GL lines
        statement = stmt_line.statement
        matched = statement.matched_lines or 0
        total = statement.total_lines or (
            (statement.matched_lines or 0) + (statement.unmatched_lines or 0)
        )
        match_pct = round(matched / total * 100) if total else 0

        # Resolve source URL + match detail for the primary (first) GL line
        source_url = ""
        multi_match_detail: dict[str, str] | None = None
        if journal_line_ids:
            gl_line = db.get(JournalEntryLine, journal_line_ids[0])
            if gl_line:
                entry = getattr(gl_line, "journal_entry", None) or getattr(
                    gl_line, "entry", None
                )
                if entry:
                    source_url = _build_source_url(
                        getattr(entry, "source_document_type", None),
                        getattr(entry, "source_document_id", None),
                        getattr(entry, "entry_id", None),
                    )
                    multi_match_detail = _build_match_detail(db, entry, source_url)

        return JSONResponse(
            content={
                "status": "ok",
                "matched_lines": matched,
                "unmatched_lines": statement.unmatched_lines or 0,
                "match_pct": match_pct,
                "total_lines": total,
                "source_url": source_url,
                "match_detail": multi_match_detail,
                "match_count": len(journal_line_ids),
            },
            status_code=200,
        )

    async def bulk_delete_statements_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
    ):
        """Handle bulk delete statements request."""
        from app.schemas.bulk_actions import BulkActionRequest
        from app.services.finance.banking.bulk import get_statement_bulk_service

        body = await request.json()
        req = BulkActionRequest(**body)
        service = get_statement_bulk_service(
            db,
            coerce_uuid(auth.organization_id),
            coerce_uuid(auth.user_id),
        )
        return await service.bulk_delete(req.ids)

    async def bulk_export_statements_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
    ):
        """Handle bulk export statements request."""
        from app.schemas.bulk_actions import BulkExportRequest
        from app.services.finance.banking.bulk import get_statement_bulk_service

        body = await request.json()
        req = BulkExportRequest(**body)
        service = get_statement_bulk_service(
            db,
            coerce_uuid(auth.organization_id),
            coerce_uuid(auth.user_id),
        )
        return await service.bulk_export(req.ids, req.format)

    async def bulk_export_accounts_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
    ) -> Response:
        """Handle bulk export bank accounts request."""
        from app.schemas.bulk_actions import BulkExportRequest
        from app.services.finance.banking.bulk import get_account_bulk_service

        body = await request.json()
        req = BulkExportRequest(**body)
        service = get_account_bulk_service(
            db,
            coerce_uuid(auth.organization_id),
            coerce_uuid(auth.user_id),
        )
        return await service.bulk_export(req.ids, req.format)

    async def export_all_accounts_response(
        self,
        auth: WebAuthContext,
        db: Session,
        search: str = "",
        status: str = "",
    ) -> Response:
        """Export all bank accounts matching filters."""
        from app.services.finance.banking.bulk import get_account_bulk_service

        service = get_account_bulk_service(
            db,
            coerce_uuid(auth.organization_id),
            coerce_uuid(auth.user_id),
        )
        return await service.export_all(search=search, status=status)

    async def bulk_export_payees_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
    ) -> Response:
        """Handle bulk export payees request."""
        from app.schemas.bulk_actions import BulkExportRequest
        from app.services.finance.banking.bulk import get_payee_bulk_service

        body = await request.json()
        req = BulkExportRequest(**body)
        service = get_payee_bulk_service(
            db,
            coerce_uuid(auth.organization_id),
            coerce_uuid(auth.user_id),
        )
        return await service.bulk_export(req.ids, req.format)

    async def export_all_payees_response(
        self,
        auth: WebAuthContext,
        db: Session,
        search: str = "",
        payee_type: str = "",
    ) -> Response:
        """Export all payees matching filters."""
        from app.services.finance.banking.bulk import get_payee_bulk_service

        service = get_payee_bulk_service(
            db,
            coerce_uuid(auth.organization_id),
            coerce_uuid(auth.user_id),
        )
        extra_filters: dict[str, str] = {}
        if payee_type:
            extra_filters["payee_type"] = payee_type
        return await service.export_all(
            search=search, extra_filters=extra_filters if extra_filters else None
        )

    def list_reconciliations_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        account_id: str | None,
        status: str | None,
        start_date: str | None,
        end_date: str | None,
        page: int,
        limit: int = 50,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Bank Reconciliations", "banking", db=db)
        context.update(
            self.list_reconciliations_context(
                db,
                str(auth.organization_id),
                account_id=account_id,
                status=status,
                start_date=start_date,
                end_date=end_date,
                page=page,
                limit=limit,
            )
        )
        return templates.TemplateResponse(
            request,
            "finance/banking/reconciliations.html",
            context,
        )

    def reconciliation_new_form_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        *,
        account_id: str | None = None,
    ) -> HTMLResponse:
        context = base_context(request, auth, "New Reconciliation", "banking", db=db)
        context.update(
            self.reconciliation_form_context(
                db, str(auth.organization_id), account_id=account_id
            )
        )
        return templates.TemplateResponse(
            request,
            "finance/banking/reconciliation_form.html",
            context,
        )

    def reconciliation_detail_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        reconciliation_id: str,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Bank Reconciliation", "banking", db=db)
        context.update(
            self.reconciliation_detail_context(
                db,
                str(auth.organization_id),
                reconciliation_id,
            )
        )
        return templates.TemplateResponse(
            request, "finance/banking/reconciliation.html", context
        )

    def reconciliation_report_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        reconciliation_id: str,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Reconciliation Report", "banking", db=db)
        context.update(
            self.reconciliation_report_context(
                db,
                str(auth.organization_id),
                reconciliation_id,
            )
        )
        return templates.TemplateResponse(
            request,
            "finance/banking/reconciliation_report.html",
            context,
        )

    def list_payees_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        search: str | None,
        payee_type: str | None,
        page: int,
        limit: int = 25,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Payees", "banking", db=db)
        context.update(
            self.list_payees_context(
                db,
                str(auth.organization_id),
                search=search,
                payee_type=payee_type,
                page=page,
                per_page=limit,
            )
        )
        return templates.TemplateResponse(
            request, "finance/banking/payees.html", context
        )

    def payee_new_form_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
    ) -> HTMLResponse:
        context = base_context(request, auth, "New Payee", "banking", db=db)
        context.update(self.payee_form_context(db, str(auth.organization_id)))
        return templates.TemplateResponse(
            request, "finance/banking/payee_form.html", context
        )

    def payee_detail_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        payee_id: str,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Edit Payee", "banking", db=db)
        context.update(
            self.payee_form_context(
                db,
                str(auth.organization_id),
                payee_id=payee_id,
            )
        )
        return templates.TemplateResponse(
            request, "finance/banking/payee_form.html", context
        )

    def list_rules_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        rule_type: str | None,
        page: int,
        limit: int = 25,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Transaction Rules", "banking", db=db)
        context.update(
            self.list_rules_context(
                db,
                str(auth.organization_id),
                rule_type=rule_type,
                page=page,
                per_page=limit,
            )
        )
        return templates.TemplateResponse(
            request, "finance/banking/rules.html", context
        )

    def rule_new_form_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
    ) -> HTMLResponse:
        context = base_context(request, auth, "New Transaction Rule", "banking", db=db)
        context.update(self.rule_form_context(db, str(auth.organization_id)))
        return templates.TemplateResponse(
            request, "finance/banking/rule_form.html", context
        )

    def rule_detail_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        rule_id: str,
    ) -> HTMLResponse:
        context = base_context(request, auth, "Edit Transaction Rule", "banking", db=db)
        context.update(
            self.rule_form_context(
                db,
                str(auth.organization_id),
                rule_id=rule_id,
            )
        )
        return templates.TemplateResponse(
            request, "finance/banking/rule_form.html", context
        )

    def rule_duplicate_form_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        rule_id: str,
    ) -> HTMLResponse:
        """Render duplicate-rule form."""
        context = base_context(request, auth, "Duplicate Rule", "banking", db=db)
        context.update(
            self.rule_duplicate_form_context(
                db,
                str(auth.organization_id),
                rule_id=rule_id,
            )
        )
        return templates.TemplateResponse(
            request, "finance/banking/rule_duplicate.html", context
        )

    def duplicate_rule_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        rule_id: str,
        bank_account_ids: list[str],
        include_global: bool,
    ) -> HTMLResponse | RedirectResponse:
        """Handle POST duplicate-rule action."""
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        target_ids = [bid for bid in bank_account_ids if bid]
        try:
            service = TransactionCategorizationService()
            copies = service.duplicate_rule_to_accounts(
                db=db,
                organization_id=org_id,
                source_rule_id=UUID(rule_id),
                bank_account_ids=[UUID(v) for v in target_ids],
                include_global=include_global,
                created_by=auth.person_id,
            )
            db.flush()
        except (ValueError, TypeError) as exc:
            context = base_context(request, auth, "Duplicate Rule", "banking", db=db)
            context.update(
                self.rule_duplicate_form_context(
                    db,
                    str(auth.organization_id),
                    rule_id=rule_id,
                )
            )
            context["error"] = str(exc)
            context["selected_bank_account_ids"] = target_ids
            context["include_global"] = include_global
            return templates.TemplateResponse(
                request, "finance/banking/rule_duplicate.html", context
            )

        return RedirectResponse(
            url=f"/finance/banking/rules?success=Rule+duplicated+({len(copies)}+copy)",
            status_code=303,
        )

    def bulk_rule_duplicate_form_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        rule_ids: list[str],
    ) -> HTMLResponse:
        """Render bulk duplicate-rule form."""
        context = base_context(request, auth, "Bulk Duplicate Rules", "banking", db=db)
        try:
            context.update(
                self.bulk_rule_duplicate_form_context(
                    db,
                    str(auth.organization_id),
                    rule_ids=rule_ids,
                )
            )
        except ValueError as exc:
            return RedirectResponse(
                url=f"/finance/banking/rules?error={str(exc).replace(' ', '+')}",
                status_code=303,
            )
        return templates.TemplateResponse(
            request, "finance/banking/rule_duplicate_bulk.html", context
        )

    def bulk_duplicate_rules_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        rule_ids: list[str],
        bank_account_ids: list[str],
        include_global: bool,
    ) -> HTMLResponse | RedirectResponse:
        """Handle bulk duplicate-rules action."""
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        unique_rule_ids = list(dict.fromkeys(rid for rid in rule_ids if rid))
        unique_bank_ids = list(dict.fromkeys(bid for bid in bank_account_ids if bid))
        service = TransactionCategorizationService()

        try:
            total_copies = 0
            for rid in unique_rule_ids:
                copies = service.duplicate_rule_to_accounts(
                    db=db,
                    organization_id=org_id,
                    source_rule_id=UUID(rid),
                    bank_account_ids=[UUID(v) for v in unique_bank_ids],
                    include_global=include_global,
                    created_by=auth.person_id,
                )
                total_copies += len(copies)
            db.flush()
        except (ValueError, TypeError) as exc:
            context = base_context(
                request,
                auth,
                "Bulk Duplicate Rules",
                "banking",
                db=db,
            )
            context.update(
                self.bulk_rule_duplicate_form_context(
                    db,
                    str(auth.organization_id),
                    rule_ids=unique_rule_ids,
                )
            )
            context["error"] = str(exc)
            context["selected_bank_account_ids"] = unique_bank_ids
            context["include_global"] = include_global
            return templates.TemplateResponse(
                request, "finance/banking/rule_duplicate_bulk.html", context
            )

        return RedirectResponse(
            url=(
                "/finance/banking/rules?"
                f"success=Rules+duplicated+({total_copies}+copies)"
            ),
            status_code=303,
        )

    def reorder_rules_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        rule_id: str,
        direction: str,
    ) -> HTMLResponse | RedirectResponse:
        """Handle POST to reorder a rule up or down.

        Returns the #results-container partial for HTMX swap.
        """
        from app.services.finance.banking.categorization import (
            TransactionCategorizationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        service = TransactionCategorizationService()
        service.swap_rule_order(db, org_id, UUID(rule_id), direction)
        db.flush()

        # Check if this is an HTMX request — return partial
        if request.headers.get("HX-Request"):
            context = base_context(request, auth, "Transaction Rules", "banking", db=db)
            context.update(self.list_rules_context(db, str(org_id)))
            return templates.TemplateResponse(
                request,
                "finance/banking/_rules_table.html",
                context,
            )

        # Fallback: full redirect
        return RedirectResponse(
            url="/finance/banking/rules?success=Record+saved+successfully",
            status_code=303,
        )

    # ─── Reconciliation POST handlers ───────────────────────────────────

    async def create_reconciliation_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        redirect_cls: type,
    ) -> Any:
        """Create a reconciliation from the web form and redirect to detail page."""
        from decimal import Decimal, InvalidOperation
        from uuid import UUID as _UUID

        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
            ReconciliationInput,
        )

        form = await request.form()
        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        try:
            bank_account_id = _UUID(str(form.get("bank_account_id", "")))
            inp = ReconciliationInput(
                reconciliation_date=date.fromisoformat(
                    str(form.get("reconciliation_date", ""))
                ),
                period_start=date.fromisoformat(str(form.get("period_start", ""))),
                period_end=date.fromisoformat(str(form.get("period_end", ""))),
                statement_opening_balance=Decimal(
                    str(form.get("statement_opening_balance", "0"))
                ),
                statement_closing_balance=Decimal(
                    str(form.get("statement_closing_balance", "0"))
                ),
                notes=str(form.get("notes", "")) or None,
            )
        except (ValueError, InvalidOperation) as e:
            logger.warning("Invalid reconciliation form data: %s", e)
            # Re-render form with error
            context = base_context(
                request, auth, "New Reconciliation", "banking", db=db
            )
            context.update(self.reconciliation_form_context(db, str(org_id)))
            context["error"] = f"Invalid form data: {e}"
            return templates.TemplateResponse(
                request,
                "finance/banking/reconciliation_form.html",
                context,
            )

        svc = BankReconciliationService()
        try:
            user_id = getattr(auth, "user_id", None) or getattr(auth, "person_id", None)
            recon = svc.create_reconciliation(
                db=db,
                organization_id=org_id,
                bank_account_id=bank_account_id,
                input=inp,
                prepared_by=user_id,
            )
            db.flush()
        except HTTPException:
            raise
        except (ValueError, RuntimeError) as e:
            logger.warning("Reconciliation creation failed: %s", e)
            context = base_context(
                request, auth, "New Reconciliation", "banking", db=db
            )
            context.update(self.reconciliation_form_context(db, str(org_id)))
            context["error"] = str(e)
            return templates.TemplateResponse(
                request,
                "finance/banking/reconciliation_form.html",
                context,
            )

        return redirect_cls(
            url=f"/finance/banking/reconciliations/{recon.reconciliation_id}",
            status_code=303,
        )

    async def reconciliation_action_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        reconciliation_id: str,
        action: str,
    ) -> Response:
        """Handle reconciliation lifecycle actions (auto-match, submit, approve, reject)."""
        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
        )

        await request.form()  # consume form body for CSRF validation
        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        svc = BankReconciliationService()
        recon_uuid = UUID(reconciliation_id)
        user_id = getattr(auth, "user_id", None) or getattr(auth, "person_id", None)

        try:
            if action == "auto_match":
                svc.auto_match(
                    db=db,
                    organization_id=org_id,
                    reconciliation_id=recon_uuid,
                    created_by=user_id,
                )
            elif action == "submit":
                svc.submit_for_review(db, org_id, recon_uuid)
            elif action == "approve":
                if not user_id:
                    raise HTTPException(status_code=401, detail="User ID required")
                svc.approve(
                    db=db,
                    organization_id=org_id,
                    reconciliation_id=recon_uuid,
                    approved_by=user_id,
                )
            elif action == "reject":
                if not user_id:
                    raise HTTPException(status_code=401, detail="User ID required")
                notes = request.query_params.get("notes", "Rejected via UI")
                svc.reject(
                    db=db,
                    organization_id=org_id,
                    reconciliation_id=recon_uuid,
                    rejected_by=user_id,
                    notes=notes,
                )
            db.flush()
        except HTTPException:
            raise
        except (ValueError, RuntimeError) as e:
            logger.warning("Reconciliation %s failed: %s", action, e)
            raise HTTPException(status_code=400, detail=str(e))

        # HTMX requests get a 200 + HX-Refresh header
        if request.headers.get("HX-Request"):
            return Response(
                content="",
                status_code=200,
                headers={"HX-Refresh": "true"},
            )
        return RedirectResponse(
            url=f"/finance/banking/reconciliations/{reconciliation_id}",
            status_code=303,
        )

    async def reconciliation_match_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        reconciliation_id: str,
    ) -> Response:
        """Add a single match from Alpine.js fetch (JSON body)."""
        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        body = await request.json()
        svc = BankReconciliationService()
        user_id = getattr(auth, "user_id", None) or getattr(auth, "person_id", None)
        force_match = bool(body.get("force_match", False))

        try:
            from app.models.finance.banking.bank_reconciliation import (
                ReconciliationMatchType,
            )
            from app.services.finance.banking.bank_reconciliation import (
                ReconciliationMatchInput,
            )

            match_type_str = body.get("match_type", "manual")
            try:
                match_type = ReconciliationMatchType(match_type_str)
            except ValueError:
                match_type = ReconciliationMatchType.manual

            match_input = ReconciliationMatchInput(
                statement_line_id=UUID(str(body["statement_line_id"])),
                journal_line_id=UUID(str(body["journal_line_id"])),
                match_type=match_type,
            )
            svc.add_match(
                db=db,
                organization_id=org_id,
                reconciliation_id=UUID(reconciliation_id),
                input=match_input,
                created_by=user_id,
                force_match=force_match,
            )
            db.flush()
        except HTTPException:
            raise
        except (ValueError, RuntimeError, KeyError) as e:
            logger.warning("Match creation failed: %s", e)
            return JSONResponse(content={"detail": str(e)}, status_code=400)

        return JSONResponse(content={"status": "ok"}, status_code=200)

    async def reconciliation_multi_match_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        reconciliation_id: str,
    ) -> Response:
        """Add a multi-match from Alpine.js fetch (JSON body)."""
        from app.services.finance.banking.bank_reconciliation import (
            BankReconciliationService,
        )

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        body = await request.json()
        svc = BankReconciliationService()
        user_id = getattr(auth, "user_id", None) or getattr(auth, "person_id", None)

        try:
            from decimal import Decimal

            svc.add_multi_match(
                db=db,
                organization_id=org_id,
                reconciliation_id=UUID(reconciliation_id),
                statement_line_ids=[UUID(s) for s in body["statement_line_ids"]],
                journal_line_ids=[UUID(s) for s in body["journal_line_ids"]],
                tolerance=Decimal(str(body.get("tolerance", "0.01"))),
                notes=body.get("notes"),
                created_by=user_id,
            )
            db.flush()
        except HTTPException:
            raise
        except (ValueError, RuntimeError, KeyError) as e:
            logger.warning("Multi-match creation failed: %s", e)
            return JSONResponse(content={"detail": str(e)}, status_code=400)

        return JSONResponse(content={"status": "ok"}, status_code=200)

    # ─────────────────────────────────────────────────────────────
    # Bank Account Create / Update (form POST handlers)
    # ─────────────────────────────────────────────────────────────

    async def create_account_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
    ) -> Response:
        """Handle POST to create a new bank account from form data."""
        from app.models.finance.banking.bank_account import BankAccountType
        from app.services.finance.banking.bank_account import (
            BankAccountInput,
            bank_account_service,
        )

        org_id = auth.organization_id
        user_id = auth.user_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        form = await request.form()
        try:
            data = BankAccountInput(
                bank_name=str(form.get("bank_name", "")),
                account_number=str(form.get("account_number", "")),
                account_name=str(form.get("account_name", "")),
                gl_account_id=UUID(str(form["gl_account_id"])),
                currency_code=(
                    str(form.get("currency_code", "")).strip()
                    or org_context_service.get_functional_currency(db, org_id)
                ),
                account_type=BankAccountType(str(form.get("account_type", "checking"))),
                bank_code=str(form.get("bank_code", "")) or None,
                branch_code=str(form.get("branch_code", "")) or None,
                branch_name=str(form.get("branch_name", "")) or None,
                iban=str(form.get("iban", "")) or None,
                contact_name=str(form.get("contact_name", "")) or None,
                contact_phone=str(form.get("contact_phone", "")) or None,
                contact_email=str(form.get("contact_email", "")) or None,
                notes=str(form.get("notes", "")) or None,
                allow_overdraft="allow_overdraft" in form,
                overdraft_limit=Decimal(str(form["overdraft_limit"]))
                if form.get("overdraft_limit")
                else None,
            )
            account = bank_account_service.create(
                db, org_id, data, coerce_uuid(user_id) if user_id else None
            )
            db.flush()
            return RedirectResponse(
                url=f"/finance/banking/accounts/{account.bank_account_id}",
                status_code=303,
            )
        except (ValueError, RuntimeError) as exc:
            logger.warning("Bank account creation failed: %s", exc)
            return RedirectResponse(
                url=f"/finance/banking/accounts/new?error={exc}",
                status_code=303,
            )

    async def update_account_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        account_id: str,
    ) -> Response:
        """Handle POST to update an existing bank account from form data."""
        from app.models.finance.banking.bank_account import BankAccountType
        from app.services.finance.banking.bank_account import (
            BankAccountInput,
            bank_account_service,
        )

        org_id = auth.organization_id
        user_id = auth.user_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        # Load existing account to preserve account_number (read-only in edit)
        existing = db.get(BankAccount, coerce_uuid(account_id))
        if not existing or existing.organization_id != coerce_uuid(org_id):
            raise HTTPException(status_code=404, detail="Bank account not found")

        form = await request.form()
        try:
            data = BankAccountInput(
                bank_name=str(form.get("bank_name", "")),
                account_number=existing.account_number,
                account_name=str(form.get("account_name", "")),
                gl_account_id=UUID(str(form["gl_account_id"])),
                currency_code=str(form.get("currency_code", existing.currency_code)),
                account_type=BankAccountType(str(form.get("account_type", "checking"))),
                bank_code=str(form.get("bank_code", "")) or None,
                branch_code=str(form.get("branch_code", "")) or None,
                branch_name=str(form.get("branch_name", "")) or None,
                iban=str(form.get("iban", "")) or None,
                contact_name=str(form.get("contact_name", "")) or None,
                contact_phone=str(form.get("contact_phone", "")) or None,
                contact_email=str(form.get("contact_email", "")) or None,
                notes=str(form.get("notes", "")) or None,
                allow_overdraft="allow_overdraft" in form,
                overdraft_limit=Decimal(str(form["overdraft_limit"]))
                if form.get("overdraft_limit")
                else None,
            )
            bank_account_service.update(
                db,
                org_id,
                coerce_uuid(account_id),
                data,
                coerce_uuid(user_id) if user_id else None,
            )
            db.flush()
            return RedirectResponse(
                url=f"/finance/banking/accounts/{account_id}",
                status_code=303,
            )
        except (ValueError, RuntimeError) as exc:
            logger.warning("Bank account update failed: %s", exc)
            return RedirectResponse(
                url=f"/finance/banking/accounts/{account_id}/edit?error={exc}",
                status_code=303,
            )

    # ─────────────────────────────────────────────────────────────
    # Payee Create / Update (form POST handlers)
    # ─────────────────────────────────────────────────────────────

    async def create_payee_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
    ) -> Response:
        """Handle POST to create a new payee from form data."""
        from app.models.finance.banking.payee import Payee, PayeeType

        org_id = auth.organization_id
        user_id = auth.user_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        form = await request.form()
        try:
            payee = Payee(
                organization_id=coerce_uuid(org_id),
                payee_name=str(form.get("payee_name", "")),
                payee_type=PayeeType(str(form.get("payee_type", "OTHER"))),
                is_active="is_active" in form,
                name_patterns=str(form.get("name_patterns", "")) or None,
                default_account_id=UUID(str(form["default_account_id"]))
                if form.get("default_account_id")
                else None,
                default_tax_code_id=UUID(str(form["default_tax_code_id"]))
                if form.get("default_tax_code_id")
                else None,
                supplier_id=UUID(str(form["supplier_id"]))
                if form.get("supplier_id")
                else None,
                customer_id=UUID(str(form["customer_id"]))
                if form.get("customer_id")
                else None,
                notes=str(form.get("notes", "")) or None,
                created_by=coerce_uuid(user_id) if user_id else None,
            )
            db.add(payee)
            db.flush()
            db.flush()
            logger.info("Created payee %s: %s", payee.payee_id, payee.payee_name)
            return RedirectResponse(
                url="/finance/banking/payees",
                status_code=303,
            )
        except (ValueError, RuntimeError) as exc:
            logger.warning("Payee creation failed: %s", exc)
            return RedirectResponse(
                url=f"/finance/banking/payees/new?error={exc}",
                status_code=303,
            )

    async def update_payee_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
        payee_id: str,
    ) -> Response:
        """Handle POST to update an existing payee from form data."""
        from app.models.finance.banking.payee import Payee, PayeeType

        org_id = auth.organization_id
        if org_id is None:
            raise HTTPException(status_code=401, detail="Authentication required")

        form = await request.form()
        payee = db.get(Payee, coerce_uuid(payee_id))
        if not payee or payee.organization_id != coerce_uuid(org_id):
            raise HTTPException(status_code=404, detail="Payee not found")

        try:
            payee.payee_name = str(form.get("payee_name", payee.payee_name))
            ptype_raw = str(form.get("payee_type", ""))
            if ptype_raw:
                payee.payee_type = PayeeType(ptype_raw)
            payee.is_active = "is_active" in form
            payee.name_patterns = str(form.get("name_patterns", "")) or None
            payee.default_account_id = (
                UUID(str(form["default_account_id"]))
                if form.get("default_account_id")
                else None
            )
            payee.default_tax_code_id = (
                UUID(str(form["default_tax_code_id"]))
                if form.get("default_tax_code_id")
                else None
            )
            payee.supplier_id = (
                UUID(str(form["supplier_id"])) if form.get("supplier_id") else None
            )
            payee.customer_id = (
                UUID(str(form["customer_id"])) if form.get("customer_id") else None
            )
            payee.notes = str(form.get("notes", "")) or None
            db.flush()
            db.flush()
            logger.info("Updated payee %s: %s", payee.payee_id, payee.payee_name)
            return RedirectResponse(
                url="/finance/banking/payees",
                status_code=303,
            )
        except (ValueError, RuntimeError) as exc:
            logger.warning("Payee update failed: %s", exc)
            return RedirectResponse(
                url=f"/finance/banking/payees/{payee_id}?error={exc}",
                status_code=303,
            )

    # =========================================================================
    # Dashboard
    # =========================================================================

    @staticmethod
    def dashboard_context(
        db: Session,
        organization_id: str,
    ) -> dict[str, Any]:
        """Build context for the banking dashboard page."""
        org_id = coerce_uuid(organization_id)

        # ── Account totals ──
        accounts = list(
            db.scalars(
                select(BankAccount)
                .where(
                    BankAccount.organization_id == org_id,
                    BankAccount.status == BankAccountStatus.active,
                )
                .order_by(BankAccount.bank_name, BankAccount.account_name)
            ).all()
        )

        total_balance = sum(
            (a.last_statement_balance or Decimal("0") for a in accounts),
            Decimal("0"),
        )

        # ── Unreconciled transaction count ──
        gl_account_ids = [a.gl_account_id for a in accounts if a.gl_account_id]

        unreconciled_count = 0
        if gl_account_ids:
            # Count posted GL lines against bank accounts that are unmatched
            # (no matching bank statement line)
            unreconciled_count = (
                db.scalar(
                    select(func.count(JournalEntryLine.line_id))
                    .join(
                        JournalEntry,
                        JournalEntryLine.journal_entry_id
                        == JournalEntry.journal_entry_id,
                    )
                    .where(
                        JournalEntry.organization_id == org_id,
                        JournalEntry.status == JournalStatus.POSTED,
                        JournalEntryLine.account_id.in_(gl_account_ids),
                        ~JournalEntryLine.line_id.in_(
                            select(BankStatementLine.matched_journal_line_id).where(
                                BankStatementLine.matched_journal_line_id.isnot(None)
                            )
                        ),
                    )
                )
                or 0
            )

        # ── MTD inflows/outflows ──
        today = date.today()
        month_start = today.replace(day=1)
        inflows_mtd = Decimal("0")
        outflows_mtd = Decimal("0")

        if gl_account_ids:
            mtd_row = db.execute(
                select(
                    func.coalesce(
                        func.sum(JournalEntryLine.debit_amount), Decimal("0")
                    ).label("total_debits"),
                    func.coalesce(
                        func.sum(JournalEntryLine.credit_amount), Decimal("0")
                    ).label("total_credits"),
                )
                .join(
                    JournalEntry,
                    JournalEntryLine.journal_entry_id == JournalEntry.journal_entry_id,
                )
                .where(
                    JournalEntry.organization_id == org_id,
                    JournalEntry.status == JournalStatus.POSTED,
                    JournalEntryLine.account_id.in_(gl_account_ids),
                    JournalEntry.entry_date >= month_start,
                    JournalEntry.entry_date <= today,
                )
            ).one()
            # For bank asset accounts: debit = money in, credit = money out
            inflows_mtd = mtd_row.total_debits or Decimal("0")
            outflows_mtd = mtd_row.total_credits or Decimal("0")

        # ── Reconciliation status ──
        recon_counts: dict[ReconciliationStatus, int] = {
            row[0]: row[1]
            for row in db.execute(
                select(
                    BankReconciliation.status,
                    func.count(BankReconciliation.reconciliation_id),
                )
                .where(BankReconciliation.organization_id == org_id)
                .group_by(BankReconciliation.status)
            ).all()
        }

        # ── Recent transactions (last 10) ──
        recent_transactions: list[dict[str, Any]] = []
        if gl_account_ids:
            gl_to_bank: dict[UUID, BankAccount] = {}
            for acct in accounts:
                if acct.gl_account_id and acct.gl_account_id not in gl_to_bank:
                    gl_to_bank[acct.gl_account_id] = acct

            txn_stmt = (
                select(JournalEntryLine, JournalEntry)
                .join(
                    JournalEntry,
                    JournalEntryLine.journal_entry_id == JournalEntry.journal_entry_id,
                )
                .where(
                    JournalEntry.organization_id == org_id,
                    JournalEntry.status == JournalStatus.POSTED,
                    JournalEntryLine.account_id.in_(gl_account_ids),
                )
                .order_by(JournalEntry.entry_date.desc())
                .limit(10)
            )
            rows = db.execute(txn_stmt).all()

            # Batch-resolve payment metadata
            metadata_pairs = [
                (
                    getattr(entry, "source_document_type", None),
                    getattr(entry, "source_document_id", None),
                )
                for _line, entry in rows
            ]
            metadata_map = resolve_payment_metadata_batch(db, metadata_pairs)

            for line, entry in rows:
                bank_acct = gl_to_bank.get(line.account_id)
                if not bank_acct:
                    continue
                currency = (
                    bank_acct.currency_code
                    or org_context_service.get_functional_currency(db, org_id)
                )
                doc_id = getattr(entry, "source_document_id", None)
                meta = metadata_map.get(doc_id) if doc_id else None
                txn = _gl_line_as_transaction(line, entry, bank_acct, currency, meta)
                recent_transactions.append(txn)

        # ── Account balances for display ──
        account_balances = [
            {
                "bank_account_id": a.bank_account_id,
                "bank_name": a.bank_name,
                "account_name": a.account_name,
                "account_number": a.account_number,
                "currency_code": a.currency_code,
                "balance": _format_currency(a.last_statement_balance, a.currency_code),
                "last_reconciled_date": _format_date(a.last_reconciled_date),
                "status": a.status.value if a.status else "",
            }
            for a in accounts
        ]

        org_currency = (
            accounts[0].currency_code
            if accounts
            else org_context_service.get_functional_currency(db, org_id)
        )

        return {
            "total_balance": _format_currency(total_balance, org_currency),
            "unreconciled_count": unreconciled_count,
            "inflows_mtd": _format_currency(inflows_mtd, org_currency),
            "outflows_mtd": _format_currency(outflows_mtd, org_currency),
            "recent_transactions": recent_transactions,
            "account_balances": account_balances,
            "account_count": len(accounts),
            "recon_draft": recon_counts.get(ReconciliationStatus.draft, 0),
            "recon_pending_review": recon_counts.get(
                ReconciliationStatus.pending_review, 0
            ),
            "recon_approved": recon_counts.get(ReconciliationStatus.approved, 0),
            "recon_rejected": recon_counts.get(ReconciliationStatus.rejected, 0),
        }

    def dashboard_response(
        self,
        request: Request,
        auth: WebAuthContext,
        db: Session,
    ) -> HTMLResponse:
        """Render the banking dashboard page."""
        context = base_context(request, auth, "Banking", "banking", db=db)
        context.update(self.dashboard_context(db, str(auth.organization_id)))
        return templates.TemplateResponse(
            request, "finance/banking/dashboard.html", context
        )


banking_web_service = BankingWebService()
