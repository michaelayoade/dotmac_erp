"""
Bank Statement Service.

Provides statement import and management functionality.
"""

import builtins
import csv
import logging
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO, StringIO
from uuid import UUID

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import and_, func, select
from sqlalchemy.orm import Session

from app.models.finance.banking.bank_account import BankAccount
from app.models.finance.banking.bank_statement import (
    BankStatement,
    BankStatementLine,
    BankStatementStatus,
    StatementLineType,
)

logger = logging.getLogger(__name__)


@dataclass
class StatementLineInput:
    """Input for a bank statement line."""

    line_number: int
    transaction_date: date
    transaction_type: StatementLineType
    amount: Decimal
    description: str | None = None
    reference: str | None = None
    payee_payer: str | None = None
    bank_reference: str | None = None
    check_number: str | None = None
    bank_category: str | None = None
    bank_code: str | None = None
    value_date: date | None = None
    running_balance: Decimal | None = None
    transaction_id: str | None = None
    raw_data: dict | None = None


@dataclass
class DuplicateLineInfo:
    """Information about a duplicate line."""

    line_number: int
    transaction_date: date
    amount: Decimal
    description: str | None
    original_statement_id: UUID
    original_line_id: UUID


@dataclass
class StatementImportResult:
    """Result of a statement import operation."""

    statement: BankStatement
    lines_imported: int
    lines_skipped: int
    duplicates_found: int = 0
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    duplicate_lines: list[DuplicateLineInfo] = field(default_factory=list)


class BankStatementService:
    """Service for managing bank statements."""

    @staticmethod
    def _normalize_header(value: str) -> str:
        return value.strip().lower().replace(" ", "_")

    @staticmethod
    def _parse_decimal(value: str | None, field: str, row: int) -> Decimal | None:
        if value is None:
            return None
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))
        text = str(value).strip()
        if not text:
            return None
        # Allow comma thousand separators (e.g., "1,234.56")
        text = text.replace(",", "")
        try:
            amount = Decimal(text)
        except Exception as exc:  # noqa: BLE001 - report parsing errors
            raise ValueError(f"Row {row}: invalid {field} '{text}'") from exc
        if amount < 0:
            raise ValueError(f"Row {row}: {field} must be positive")
        return amount

    @staticmethod
    def _parse_date(
        value: str | None,
        field: str,
        row: int,
        date_format: str | None = None,
    ) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = "" if value is None else str(value).strip()
        if not text:
            raise ValueError(f"Row {row}: {field} is required")
        # 1. Try org's configured date format first (e.g. DD/MM/YYYY → %d/%m/%Y)
        if date_format:
            try:
                return datetime.strptime(text, date_format).date()
            except ValueError:
                pass
        # 2. Fallback: ISO 8601 (YYYY-MM-DD and YYYY/MM/DD)
        try:
            return date.fromisoformat(text.replace("/", "-"))
        except ValueError:
            pass
        fmt_hint = f" (expected format: {date_format})" if date_format else ""
        raise ValueError(f"Row {row}: invalid {field} '{text}'{fmt_hint}")

    def parse_csv_rows(
        self,
        content: bytes,
        csv_format: str,
        date_format: str | None = None,
    ) -> tuple[list[dict], list[str]]:
        errors: list[str] = []
        if csv_format not in ("type", "debit_credit"):
            return [], [f"Unsupported CSV format '{csv_format}'."]
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("utf-8", errors="replace")

        delimiter = ","
        try:
            header_line = next((line for line in text.splitlines() if line.strip()), "")
            candidates = [",", "\t", ";", "|"]
            counts = {d: header_line.count(d) for d in candidates}
            best = max(counts, key=counts.__getitem__)
            if counts.get(best, 0) > 0:
                delimiter = best
            else:
                sniff = csv.Sniffer().sniff(text[:2048])
                if sniff and getattr(sniff, "delimiter", None):
                    delimiter = sniff.delimiter
        except Exception:
            # Fall back to comma if delimiter detection fails.
            delimiter = ","

        reader = csv.DictReader(StringIO(text), delimiter=delimiter)
        if not reader.fieldnames:
            return [], ["CSV file must include a header row."]

        header_map = {
            name: self._normalize_header(name) for name in reader.fieldnames if name
        }
        fields = set(header_map.values())
        required = (
            {"transaction_date", "debit", "credit"}
            if csv_format == "debit_credit"
            else {"transaction_date", "transaction_type", "amount"}
        )
        missing = [name for name in required if name not in fields]
        if missing:
            return [], [f"Missing required column(s): {', '.join(sorted(missing))}"]

        rows: list[dict] = []
        for row_index, row in enumerate(reader, start=2):
            if not any(
                str(value).strip() for value in row.values() if value is not None
            ):
                continue
            normalized_row = {
                header_map.get(key, ""): value for key, value in row.items() if key
            }
            try:
                line_data: dict = {
                    "line_number": row_index - 1,
                    "transaction_date": self._parse_date(
                        normalized_row.get("transaction_date"),
                        "transaction_date",
                        row_index,
                        date_format=date_format,
                    ),
                    "description": (normalized_row.get("description") or "").strip()
                    or None,
                    "reference": (normalized_row.get("reference") or "").strip()
                    or None,
                    "payee_payer": (normalized_row.get("payee_payer") or "").strip()
                    or None,
                    "bank_reference": (
                        normalized_row.get("bank_reference") or ""
                    ).strip()
                    or None,
                    "check_number": (normalized_row.get("check_number") or "").strip()
                    or None,
                    "bank_category": (normalized_row.get("bank_category") or "").strip()
                    or None,
                    "bank_code": (normalized_row.get("bank_code") or "").strip()
                    or None,
                    "transaction_id": (
                        normalized_row.get("transaction_id") or ""
                    ).strip()
                    or None,
                }
                value_date = normalized_row.get("value_date")
                if value_date:
                    line_data["value_date"] = self._parse_date(
                        value_date, "value_date", row_index, date_format=date_format
                    )
                running_balance = self._parse_decimal(
                    normalized_row.get("running_balance"), "running_balance", row_index
                )
                if running_balance is not None:
                    line_data["running_balance"] = running_balance

                if csv_format == "debit_credit":
                    line_data["debit"] = self._parse_decimal(
                        normalized_row.get("debit"), "debit", row_index
                    )
                    line_data["credit"] = self._parse_decimal(
                        normalized_row.get("credit"), "credit", row_index
                    )
                else:
                    transaction_type = (
                        (normalized_row.get("transaction_type") or "").strip().lower()
                    )
                    if transaction_type not in ("debit", "credit"):
                        raise ValueError(
                            f"Row {row_index}: transaction_type must be 'debit' or 'credit'"
                        )
                    line_data["transaction_type"] = transaction_type
                    line_data["amount"] = self._parse_decimal(
                        normalized_row.get("amount"), "amount", row_index
                    )
            except ValueError as exc:
                errors.append(str(exc))
                continue

            rows.append(line_data)

        return rows, errors

    def parse_xlsx_rows(
        self,
        content: bytes,
        csv_format: str,
        date_format: str | None = None,
    ) -> tuple[list[dict], list[str]]:
        errors: list[str] = []
        if csv_format not in ("type", "debit_credit"):
            return [], [f"Unsupported CSV format '{csv_format}'."]

        workbook = load_workbook(
            filename=BytesIO(content), read_only=True, data_only=True
        )
        try:
            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                headers = next(rows_iter)
            except StopIteration:
                return [], ["XLSX file must include a header row."]

            if not headers:
                return [], ["XLSX file must include a header row."]

            header_map = [
                self._normalize_header(str(h)) if h is not None else "" for h in headers
            ]
            fields = {name for name in header_map if name}
            required = (
                {"transaction_date", "debit", "credit"}
                if csv_format == "debit_credit"
                else {"transaction_date", "transaction_type", "amount"}
            )
            missing = [name for name in required if name not in fields]
            if missing:
                return [], [f"Missing required column(s): {', '.join(sorted(missing))}"]

            rows: list[dict] = []
            for row_index, values in enumerate(rows_iter, start=2):
                if not values or not any(
                    value is not None and str(value).strip() for value in values
                ):
                    continue
                row = {
                    header_map[i]: values[i] if i < len(values) else None
                    for i in range(len(header_map))
                    if header_map[i]
                }
                try:
                    line_data: dict = {
                        "line_number": row_index - 1,
                        "transaction_date": self._parse_date(
                            row.get("transaction_date"),
                            "transaction_date",
                            row_index,
                            date_format=date_format,
                        ),
                        "description": (
                            (row.get("description") or "").strip()
                            if isinstance(row.get("description"), str)
                            else row.get("description")
                        )
                        or None,
                        "reference": (
                            (row.get("reference") or "").strip()
                            if isinstance(row.get("reference"), str)
                            else row.get("reference")
                        )
                        or None,
                        "payee_payer": (
                            (row.get("payee_payer") or "").strip()
                            if isinstance(row.get("payee_payer"), str)
                            else row.get("payee_payer")
                        )
                        or None,
                        "bank_reference": (
                            (row.get("bank_reference") or "").strip()
                            if isinstance(row.get("bank_reference"), str)
                            else row.get("bank_reference")
                        )
                        or None,
                        "check_number": (
                            (row.get("check_number") or "").strip()
                            if isinstance(row.get("check_number"), str)
                            else row.get("check_number")
                        )
                        or None,
                        "bank_category": (
                            (row.get("bank_category") or "").strip()
                            if isinstance(row.get("bank_category"), str)
                            else row.get("bank_category")
                        )
                        or None,
                        "bank_code": (
                            (row.get("bank_code") or "").strip()
                            if isinstance(row.get("bank_code"), str)
                            else row.get("bank_code")
                        )
                        or None,
                        "transaction_id": (
                            (row.get("transaction_id") or "").strip()
                            if isinstance(row.get("transaction_id"), str)
                            else row.get("transaction_id")
                        )
                        or None,
                    }
                    value_date = row.get("value_date")
                    if value_date:
                        line_data["value_date"] = self._parse_date(
                            value_date, "value_date", row_index, date_format=date_format
                        )
                    running_balance = self._parse_decimal(
                        row.get("running_balance"), "running_balance", row_index
                    )
                    if running_balance is not None:
                        line_data["running_balance"] = running_balance

                    if csv_format == "debit_credit":
                        line_data["debit"] = self._parse_decimal(
                            row.get("debit"), "debit", row_index
                        )
                        line_data["credit"] = self._parse_decimal(
                            row.get("credit"), "credit", row_index
                        )
                    else:
                        transaction_type = (
                            str(row.get("transaction_type")).strip().lower()
                            if row.get("transaction_type") is not None
                            else ""
                        )
                        if transaction_type not in ("debit", "credit"):
                            raise ValueError(
                                f"Row {row_index}: transaction_type must be 'debit' or 'credit'"
                            )
                        line_data["transaction_type"] = transaction_type
                        line_data["amount"] = self._parse_decimal(
                            row.get("amount"), "amount", row_index
                        )
                except ValueError as exc:
                    errors.append(str(exc))
                    continue

                rows.append(line_data)

            return rows, errors
        finally:
            workbook.close()

    def build_line_inputs(
        self, lines: list
    ) -> tuple[list[StatementLineInput], list[str]]:
        results: list[StatementLineInput] = []
        errors: list[str] = []

        for idx, line in enumerate(lines, start=1):
            line_number = getattr(line, "line_number", idx) or idx
            transaction_type = getattr(line, "transaction_type", None)
            amount = getattr(line, "amount", None)
            debit = getattr(line, "debit", None)
            credit = getattr(line, "credit", None)

            if isinstance(transaction_type, str):
                try:
                    transaction_type = StatementLineType(transaction_type)
                except ValueError:
                    errors.append(
                        f"Line {line_number}: transaction_type must be 'debit' or 'credit'"
                    )
                    continue

            if transaction_type is None:
                if debit and debit > 0:
                    transaction_type = StatementLineType.debit
                    amount = debit
                elif credit and credit > 0:
                    transaction_type = StatementLineType.credit
                    amount = credit
                else:
                    errors.append(
                        f"Line {line_number}: provide transaction_type+amount or debit/credit"
                    )
                    continue
            elif amount is None:
                errors.append(f"Line {line_number}: amount is required")
                continue

            results.append(
                StatementLineInput(
                    line_number=line_number,
                    transaction_date=line.transaction_date,
                    transaction_type=transaction_type,
                    amount=amount,
                    description=getattr(line, "description", None),
                    reference=getattr(line, "reference", None),
                    payee_payer=getattr(line, "payee_payer", None),
                    bank_reference=getattr(line, "bank_reference", None),
                    check_number=getattr(line, "check_number", None),
                    bank_category=getattr(line, "bank_category", None),
                    bank_code=getattr(line, "bank_code", None),
                    value_date=getattr(line, "value_date", None),
                    running_balance=getattr(line, "running_balance", None),
                    transaction_id=getattr(line, "transaction_id", None),
                    raw_data=getattr(line, "raw_data", None),
                )
            )

        return results, errors

    @staticmethod
    def build_sample_csv(format: str = "type") -> tuple[bytes, str]:
        if format == "debit_credit":
            header = (
                "transaction_date,debit,credit,description,reference,"
                "payee_payer,check_number,value_date,running_balance\n"
            )
            rows = [
                "2026-01-15,,50000.00,Salary deposit from ABC Corp,TRF-001,ABC Corp Ltd,,2026-01-15,50000.00\n",
                "2026-01-16,2500.00,,Office rent payment,CHQ-4521,Landlord Properties,4521,2026-01-16,47500.00\n",
                "2026-01-17,150.75,,ATM withdrawal - Lekki,ATM-889,,,,47349.25\n",
                "2026-01-18,,12000.00,Customer payment - Invoice INV-0042,TRF-002,XYZ Trading,,2026-01-18,59349.25\n",
                "2026-01-20,8500.00,,Vendor payment - PO-0015,TRF-003,Office Supplies Ltd,,2026-01-20,50849.25\n",
            ]
        else:
            header = (
                "transaction_date,transaction_type,amount,description,"
                "reference,payee_payer,check_number,value_date,running_balance\n"
            )
            rows = [
                "2026-01-15,credit,50000.00,Salary deposit from ABC Corp,TRF-001,ABC Corp Ltd,,2026-01-15,50000.00\n",
                "2026-01-16,debit,2500.00,Office rent payment,CHQ-4521,Landlord Properties,4521,2026-01-16,47500.00\n",
                "2026-01-17,debit,150.75,ATM withdrawal - Lekki,ATM-889,,,,47349.25\n",
                "2026-01-18,credit,12000.00,Customer payment - Invoice INV-0042,TRF-002,XYZ Trading,,2026-01-18,59349.25\n",
                "2026-01-20,debit,8500.00,Vendor payment - PO-0015,TRF-003,Office Supplies Ltd,,2026-01-20,50849.25\n",
            ]
        content = header + "".join(rows)
        return content.encode("utf-8"), "bank_statement_sample.csv"

    def _check_duplicate_line(
        self,
        db: Session,
        bank_account_id: UUID,
        line: StatementLineInput,
        organization_id: UUID | None = None,
    ) -> BankStatementLine | None:
        """
        Check if a transaction line is a potential duplicate.

        Matches on: same account, date, amount, and transaction type.
        Scoped to organization_id to prevent cross-tenant matches.
        """
        # Find existing lines with same date/amount/type
        stmt = (
            select(BankStatementLine)
            .join(BankStatement)
            .where(
                BankStatement.bank_account_id == bank_account_id,
                BankStatementLine.transaction_date == line.transaction_date,
                BankStatementLine.amount == line.amount,
                BankStatementLine.transaction_type == line.transaction_type,
            )
        )
        if organization_id is not None:
            stmt = stmt.where(BankStatement.organization_id == organization_id)
        existing = db.execute(stmt).scalars().first()

        if existing:
            # Additional check: if bank_reference matches, it's definitely a duplicate
            if line.bank_reference and existing.bank_reference:
                if line.bank_reference == existing.bank_reference:
                    return existing

            # If transaction_id matches, it's definitely a duplicate
            if line.transaction_id and existing.transaction_id:
                if line.transaction_id == existing.transaction_id:
                    return existing

            # Check description similarity
            if line.description and existing.description:
                # Simple word overlap check
                words1 = set(line.description.upper().split())
                words2 = set(existing.description.upper().split())
                if words1 and words2:
                    overlap = len(words1 & words2) / len(words1 | words2)
                    if overlap > 0.7:
                        return existing

        return None

    def import_statement(
        self,
        db: Session,
        organization_id: UUID,
        bank_account_id: UUID,
        statement_number: str,
        statement_date: date,
        period_start: date,
        period_end: date,
        opening_balance: Decimal,
        closing_balance: Decimal,
        lines: list[StatementLineInput],
        import_source: str | None = None,
        import_filename: str | None = None,
        imported_by: UUID | None = None,
        check_duplicates: bool = True,
        skip_duplicates: bool = True,
    ) -> StatementImportResult:
        """Import a bank statement with lines."""
        # Validate bank account
        bank_account = db.get(BankAccount, bank_account_id)
        if not bank_account:
            raise HTTPException(
                status_code=404, detail=f"Bank account {bank_account_id} not found"
            )

        if bank_account.organization_id != organization_id:
            raise HTTPException(
                status_code=403,
                detail="Bank account does not belong to this organization",
            )

        # Check for duplicate statement
        existing = db.execute(
            select(BankStatement).where(
                and_(
                    BankStatement.bank_account_id == bank_account_id,
                    BankStatement.statement_number == statement_number,
                )
            )
        ).scalar_one_or_none()

        if existing:
            raise HTTPException(
                status_code=400,
                detail=f"Statement {statement_number} already exists for this account",
            )

        # Calculate totals
        total_credits = Decimal("0")
        total_debits = Decimal("0")
        for line_input in lines:
            if line_input.transaction_type == StatementLineType.credit:
                total_credits += line_input.amount
            else:
                total_debits += line_input.amount

        # Create statement
        statement = BankStatement(
            organization_id=organization_id,
            bank_account_id=bank_account_id,
            statement_number=statement_number,
            statement_date=statement_date,
            period_start=period_start,
            period_end=period_end,
            opening_balance=opening_balance,
            closing_balance=closing_balance,
            total_credits=total_credits,
            total_debits=total_debits,
            currency_code=bank_account.currency_code,
            status=BankStatementStatus.imported,
            import_source=import_source,
            import_filename=import_filename,
            imported_by=imported_by,
            total_lines=len(lines),
            unmatched_lines=len(lines),
        )
        db.add(statement)
        db.flush()

        # Import lines
        result = StatementImportResult(
            statement=statement,
            lines_imported=0,
            lines_skipped=0,
        )

        for line_input in lines:
            try:
                # Check for duplicates
                if check_duplicates:
                    duplicate = self._check_duplicate_line(
                        db, bank_account_id, line_input, organization_id
                    )
                    if duplicate:
                        result.duplicates_found += 1
                        result.duplicate_lines.append(
                            DuplicateLineInfo(
                                line_number=line_input.line_number,
                                transaction_date=line_input.transaction_date,
                                amount=line_input.amount,
                                description=line_input.description,
                                original_statement_id=duplicate.statement_id,
                                original_line_id=duplicate.line_id,
                            )
                        )
                        if skip_duplicates:
                            result.lines_skipped += 1
                            result.warnings.append(
                                f"Line {line_input.line_number}: Skipped as duplicate of existing transaction"
                            )
                            continue

                line = BankStatementLine(
                    statement_id=statement.statement_id,
                    line_number=line_input.line_number,
                    transaction_id=line_input.transaction_id,
                    transaction_date=line_input.transaction_date,
                    value_date=line_input.value_date,
                    transaction_type=line_input.transaction_type,
                    amount=line_input.amount,
                    running_balance=line_input.running_balance,
                    description=line_input.description,
                    reference=line_input.reference,
                    payee_payer=line_input.payee_payer,
                    bank_reference=line_input.bank_reference,
                    check_number=line_input.check_number,
                    bank_category=line_input.bank_category,
                    bank_code=line_input.bank_code,
                    raw_data=line_input.raw_data,
                    is_matched=False,
                )
                db.add(line)
                result.lines_imported += 1
            except Exception as e:
                result.lines_skipped += 1
                result.errors.append(f"Line {line_input.line_number}: {str(e)}")

        # Update bank account with latest statement info
        bank_account.last_statement_date = datetime.combine(
            statement_date,
            datetime.min.time(),
            tzinfo=UTC,
        )
        bank_account.last_statement_balance = closing_balance

        db.flush()

        # Validate statement balance
        if not statement.is_balanced:
            result.warnings.append(
                f"Statement does not balance: "
                f"Opening ({opening_balance}) + Credits ({total_credits}) - "
                f"Debits ({total_debits}) != Closing ({closing_balance})"
            )

        db.commit()
        return result

    def get(
        self,
        db: Session,
        organization_id: UUID,
        statement_id: UUID,
    ) -> BankStatement | None:
        """Get a statement by ID within an organization."""
        statement = db.get(BankStatement, statement_id)
        if not statement or statement.organization_id != organization_id:
            return None
        return statement

    def get_with_lines(
        self,
        db: Session,
        organization_id: UUID,
        statement_id: UUID,
    ) -> BankStatement | None:
        """Get a statement with all lines loaded within an organization."""
        statement = self.get(db, organization_id, statement_id)
        if statement:
            _ = statement.lines
        return statement

    def list(
        self,
        db: Session,
        organization_id: UUID,
        bank_account_id: UUID | None = None,
        status: BankStatementStatus | None = None,
        start_date: date | None = None,
        end_date: date | None = None,
        limit: int = 100,
        offset: int = 0,
    ) -> list[BankStatement]:
        """List statements with optional filters."""
        query = select(BankStatement).where(
            BankStatement.organization_id == organization_id
        )

        if bank_account_id:
            query = query.where(BankStatement.bank_account_id == bank_account_id)
        if status:
            query = query.where(BankStatement.status == status)
        if start_date:
            query = query.where(BankStatement.statement_date >= start_date)
        if end_date:
            query = query.where(BankStatement.statement_date <= end_date)

        query = query.order_by(BankStatement.statement_date.desc())
        query = query.offset(offset).limit(limit)

        return list(db.execute(query).scalars().all())

    def count(
        self,
        db: Session,
        organization_id: UUID,
        bank_account_id: UUID | None = None,
        status: BankStatementStatus | None = None,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> int:
        """Count statements matching filters (for pagination)."""
        query = select(func.count(BankStatement.statement_id)).where(
            BankStatement.organization_id == organization_id
        )

        if bank_account_id:
            query = query.where(BankStatement.bank_account_id == bank_account_id)
        if status:
            query = query.where(BankStatement.status == status)
        if start_date:
            query = query.where(BankStatement.statement_date >= start_date)
        if end_date:
            query = query.where(BankStatement.statement_date <= end_date)

        return db.execute(query).scalar() or 0

    def get_unmatched_lines(
        self,
        db: Session,
        statement_id: UUID,
    ) -> builtins.list[BankStatementLine]:
        """Get all unmatched lines for a statement."""
        query = (
            select(BankStatementLine)
            .where(
                and_(
                    BankStatementLine.statement_id == statement_id,
                    BankStatementLine.is_matched == False,
                )
            )
            .order_by(BankStatementLine.transaction_date, BankStatementLine.line_number)
        )

        return list(db.execute(query).scalars().all())

    def mark_line_matched(
        self,
        db: Session,
        line_id: UUID,
        journal_line_id: UUID,
        matched_by: UUID | None = None,
    ) -> BankStatementLine:
        """Mark a statement line as matched to a GL entry."""
        line = db.get(BankStatementLine, line_id)
        if not line:
            raise HTTPException(
                status_code=404, detail=f"Statement line {line_id} not found"
            )

        line.is_matched = True
        line.matched_at = datetime.utcnow()
        line.matched_by = matched_by
        line.matched_journal_line_id = journal_line_id

        # Update statement counts
        statement = line.statement
        statement.matched_lines += 1
        statement.unmatched_lines -= 1

        # Check if fully reconciled
        if statement.unmatched_lines == 0:
            statement.status = BankStatementStatus.reconciled

        db.flush()
        db.commit()
        db.refresh(line)
        return line

    def unmatch_line(
        self,
        db: Session,
        line_id: UUID,
    ) -> BankStatementLine:
        """Unmatch a statement line."""
        line = db.get(BankStatementLine, line_id)
        if not line:
            raise HTTPException(
                status_code=404, detail=f"Statement line {line_id} not found"
            )

        if not line.is_matched:
            return line

        line.is_matched = False
        line.matched_at = None
        line.matched_by = None
        line.matched_journal_line_id = None

        # Update statement counts
        statement = line.statement
        statement.matched_lines -= 1
        statement.unmatched_lines += 1

        if statement.status == BankStatementStatus.reconciled:
            statement.status = BankStatementStatus.processing

        db.flush()
        db.commit()
        db.refresh(line)
        return line

    def update_status(
        self,
        db: Session,
        statement_id: UUID,
        status: BankStatementStatus,
    ) -> BankStatement:
        """Update statement status."""
        statement = db.get(BankStatement, statement_id)
        if not statement:
            raise HTTPException(
                status_code=404, detail=f"Statement {statement_id} not found"
            )

        statement.status = status
        db.flush()
        db.commit()
        db.refresh(statement)
        return statement

    def delete(self, db: Session, statement_id: UUID) -> bool:
        """Delete a statement and its lines."""
        statement = db.get(BankStatement, statement_id)
        if not statement:
            return False

        if statement.status in [
            BankStatementStatus.reconciled,
            BankStatementStatus.closed,
        ]:
            raise HTTPException(
                status_code=400, detail="Cannot delete a reconciled or closed statement"
            )

        db.delete(statement)
        db.flush()
        db.commit()
        return True

    def get_statement_summary(
        self,
        db: Session,
        organization_id: UUID,
        bank_account_id: UUID,
    ) -> dict:
        """Get summary statistics for statements of an account."""
        bank_account = db.get(BankAccount, bank_account_id)
        if not bank_account or bank_account.organization_id != organization_id:
            raise HTTPException(status_code=404, detail="Bank account not found")

        query = select(
            func.count(BankStatement.statement_id).label("total_statements"),
            func.sum(BankStatement.total_lines).label("total_lines"),
            func.sum(BankStatement.matched_lines).label("matched_lines"),
            func.sum(BankStatement.unmatched_lines).label("unmatched_lines"),
        ).where(BankStatement.bank_account_id == bank_account_id)

        result = db.execute(query).one()

        return {
            "total_statements": result.total_statements or 0,
            "total_lines": result.total_lines or 0,
            "matched_lines": result.matched_lines or 0,
            "unmatched_lines": result.unmatched_lines or 0,
            "match_rate": (
                (result.matched_lines / result.total_lines * 100)
                if result.total_lines
                else 0
            ),
        }


# Singleton instance
bank_statement_service = BankStatementService()
